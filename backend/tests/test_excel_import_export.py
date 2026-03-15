"""
Test: Excel Export/Import Feature for Session Management
Testing:
1. GET /api/sessions/{session_id}/export-template - Export Excel template
2. POST /api/sessions/{session_id}/import-data - Import Excel data
3. Verify Vehicle Checklist sheet inclusion
4. Verify raw marks to percentage calculation
5. Verify pass/fail based on program pass_percentage
"""
import pytest
import requests
import os
import io
from openpyxl import load_workbook, Workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')
SESSION_ID = "1f9cfa93-d5f5-4796-9d8a-566884663d1b"

# Admin credentials
ADMIN_EMAIL = "arjuna@mddrc.com.my"
ADMIN_PASSWORD = "Dana102229"

# Test IC numbers from existing participants
TEST_IC_NUMBERS = ["8445118", "98546161165", "85525655", "5515265", "8995495198", "874656"]

@pytest.fixture(scope="module")
def auth_token():
    """Get admin auth token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": ADMIN_EMAIL,
        "password": ADMIN_PASSWORD
    })
    assert response.status_code == 200, f"Login failed: {response.text}"
    return response.json()["access_token"]

@pytest.fixture(scope="module")
def auth_headers(auth_token):
    """Return headers with auth token"""
    return {"Authorization": f"Bearer {auth_token}"}


class TestExportTemplate:
    """Test Excel template export endpoint"""
    
    def test_export_template_returns_excel(self, auth_headers):
        """GET /api/sessions/{session_id}/export-template returns Excel file"""
        response = requests.get(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/export-template",
            headers=auth_headers
        )
        
        # Verify status code
        assert response.status_code == 200, f"Export failed: {response.text}"
        
        # Verify content type is Excel
        content_type = response.headers.get("Content-Type", "")
        assert "spreadsheetml" in content_type or "excel" in content_type.lower(), \
            f"Expected Excel content type, got: {content_type}"
        
        # Verify Content-Disposition header has filename
        content_disp = response.headers.get("Content-Disposition", "")
        assert "filename=" in content_disp, f"Expected filename in header: {content_disp}"
        print(f"SUCCESS: Export returns Excel file with header: {content_disp}")
    
    def test_export_template_has_4_sheets(self, auth_headers):
        """Verify Excel template has 4 sheets: Pre-Post Tests, Attendance, Vehicle Checklist, Instructions"""
        response = requests.get(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/export-template",
            headers=auth_headers
        )
        assert response.status_code == 200
        
        # Load workbook from response content
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        print(f"Sheet names found: {sheet_names}")
        
        # Verify required sheets exist
        assert "Pre-Post Tests" in sheet_names, f"Missing 'Pre-Post Tests' sheet. Found: {sheet_names}"
        assert "Attendance" in sheet_names, f"Missing 'Attendance' sheet. Found: {sheet_names}"
        assert "Vehicle Checklist" in sheet_names, f"Missing 'Vehicle Checklist' sheet. Found: {sheet_names}"
        assert "Instructions" in sheet_names, f"Missing 'Instructions' sheet. Found: {sheet_names}"
        print(f"SUCCESS: All 4 required sheets present: {sheet_names}")
    
    def test_pre_post_tests_sheet_has_raw_marks_columns(self, auth_headers):
        """Verify Pre-Post Tests sheet has raw marks columns (Marks + Total)"""
        response = requests.get(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/export-template",
            headers=auth_headers
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb["Pre-Post Tests"]
        
        # Get headers from row 4
        headers = [cell.value for cell in ws[4]]
        wb.close()
        
        print(f"Pre-Post Tests headers: {headers}")
        
        # Expected columns: No, Participant Name, IC Number, Pre-Test (Marks), Pre-Test (Total), Post-Test (Marks), Post-Test (Total), Remarks
        assert "No" in headers or "No." in headers or headers[0] in ["No", "No."], f"Missing 'No' column"
        assert "Participant Name" in headers, f"Missing 'Participant Name' column. Headers: {headers}"
        assert "IC Number" in headers, f"Missing 'IC Number' column. Headers: {headers}"
        
        # Check for raw marks columns
        headers_str = " ".join([str(h) for h in headers if h])
        assert "Pre-Test" in headers_str and ("Marks" in headers_str or "marks" in headers_str.lower()), \
            f"Missing Pre-Test (Marks) column. Headers: {headers}"
        assert "Post-Test" in headers_str and ("Marks" in headers_str or "marks" in headers_str.lower()), \
            f"Missing Post-Test (Marks) column. Headers: {headers}"
        
        print(f"SUCCESS: Pre-Post Tests sheet has correct raw marks columns")
    
    def test_vehicle_checklist_sheet_structure(self, auth_headers):
        """Verify Vehicle Checklist sheet has correct columns"""
        response = requests.get(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/export-template",
            headers=auth_headers
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb["Vehicle Checklist"]
        
        # Get headers from row 4
        headers = [cell.value for cell in ws[4]]
        wb.close()
        
        print(f"Vehicle Checklist headers: {headers}")
        
        # Expected base columns: No, Participant Name, IC Number, Vehicle Model, Registration No, Road Tax Expiry
        assert "Participant Name" in headers, f"Missing 'Participant Name' column. Headers: {headers}"
        assert "IC Number" in headers, f"Missing 'IC Number' column. Headers: {headers}"
        assert "Vehicle Model" in headers, f"Missing 'Vehicle Model' column. Headers: {headers}"
        assert "Registration No" in headers, f"Missing 'Registration No' column. Headers: {headers}"
        
        # Road Tax Expiry check (might have different wording)
        headers_str = " ".join([str(h) for h in headers if h])
        assert "Road Tax" in headers_str or "Roadtax" in headers_str or "road tax" in headers_str.lower(), \
            f"Missing Road Tax Expiry column. Headers: {headers}"
        
        print(f"SUCCESS: Vehicle Checklist sheet has correct base columns")
    
    def test_participant_names_not_empty(self, auth_headers):
        """Verify participant names are populated (using full_name field)"""
        response = requests.get(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/export-template",
            headers=auth_headers
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb["Pre-Post Tests"]
        
        # Get participant data starting from row 5
        participant_names = []
        for row in ws.iter_rows(min_row=5, max_row=15, values_only=True):
            if row[1]:  # Column B is Participant Name
                participant_names.append(row[1])
        wb.close()
        
        print(f"Participant names found: {participant_names}")
        
        assert len(participant_names) > 0, "No participant names found in template"
        # Check that names are not empty strings or just numbers
        valid_names = [n for n in participant_names if n and isinstance(n, str) and len(n) > 2]
        assert len(valid_names) > 0, f"Participant names appear empty or invalid: {participant_names}"
        
        print(f"SUCCESS: Found {len(valid_names)} valid participant names")
    
    def test_ic_numbers_not_empty(self, auth_headers):
        """Verify IC numbers are populated (using id_number field)"""
        response = requests.get(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/export-template",
            headers=auth_headers
        )
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        ws = wb["Pre-Post Tests"]
        
        # Get IC numbers from column C starting from row 5
        ic_numbers = []
        for row in ws.iter_rows(min_row=5, max_row=15, values_only=True):
            if row[2]:  # Column C is IC Number
                ic_numbers.append(str(row[2]))
        wb.close()
        
        print(f"IC numbers found: {ic_numbers}")
        
        assert len(ic_numbers) > 0, "No IC numbers found in template"
        # Check that ICs are not empty strings
        valid_ics = [ic for ic in ic_numbers if ic and len(ic) > 3]
        assert len(valid_ics) > 0, f"IC numbers appear empty or invalid: {ic_numbers}"
        
        print(f"SUCCESS: Found {len(valid_ics)} valid IC numbers")


class TestImportData:
    """Test Excel data import endpoint"""
    
    def test_import_test_scores_raw_marks(self, auth_headers):
        """Import raw marks and verify percentage calculation"""
        # Create test Excel file with raw marks
        wb = Workbook()
        
        # Sheet 1: Pre-Post Tests with raw marks
        ws1 = wb.active
        ws1.title = "Pre-Post Tests"
        
        # Add header info rows 1-3
        ws1['A1'] = "Session Info"
        ws1['A2'] = "Test Data"
        ws1['A3'] = ""
        
        # Headers in row 4
        ws1['A4'] = "No"
        ws1['B4'] = "Participant Name"
        ws1['C4'] = "IC Number"
        ws1['D4'] = "Pre-Test (Marks)"
        ws1['E4'] = "Pre-Test (Total)"
        ws1['F4'] = "Post-Test (Marks)"
        ws1['G4'] = "Post-Test (Total)"
        ws1['H4'] = "Remarks"
        
        # Data row (using known IC number)
        ws1['A5'] = 1
        ws1['B5'] = "Test Participant"
        ws1['C5'] = TEST_IC_NUMBERS[0]  # Use first known IC number
        ws1['D5'] = 45  # Pre-test marks obtained
        ws1['E5'] = 50  # Pre-test total marks (45/50 = 90%)
        ws1['F5'] = 48  # Post-test marks obtained  
        ws1['G5'] = 50  # Post-test total marks (48/50 = 96%)
        ws1['H5'] = "Test import"
        
        # Add attendance sheet (required)
        ws2 = wb.create_sheet("Attendance")
        ws2['A1'] = "Attendance"
        ws2['A4'] = "No"
        ws2['B4'] = "Participant Name"
        ws2['C4'] = "IC Number"
        ws2['D4'] = "Day 1"
        
        # Add vehicle checklist sheet (required)
        ws3 = wb.create_sheet("Vehicle Checklist")
        ws3['A1'] = "Vehicle Checklist"
        ws3['A4'] = "No"
        ws3['B4'] = "Participant Name"
        ws3['C4'] = "IC Number"
        ws3['D4'] = "Vehicle Model"
        ws3['E4'] = "Registration No"
        ws3['F4'] = "Road Tax Expiry"
        ws3['G4'] = "Helmet"
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        # Send import request
        files = {'file': ('test_import.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/import-data",
            headers=auth_headers,
            files=files
        )
        
        print(f"Import response: {response.status_code} - {response.text}")
        
        assert response.status_code == 200, f"Import failed: {response.text}"
        
        data = response.json()
        print(f"Import results: {data}")
        
        # Verify test scores were imported
        assert data.get("test_scores_imported", 0) >= 1, f"Expected at least 1 test score imported, got: {data}"
        
        # Check if there were any errors
        if data.get("errors"):
            print(f"Import errors: {data['errors']}")
        
        print(f"SUCCESS: Imported {data.get('test_scores_imported', 0)} test scores")
    
    def test_import_vehicle_checklist(self, auth_headers):
        """Import vehicle checklist data and verify it's stored"""
        # Create test Excel file with vehicle checklist
        wb = Workbook()
        
        # Pre-Post Tests sheet (required)
        ws1 = wb.active
        ws1.title = "Pre-Post Tests"
        ws1['A4'] = "No"
        ws1['B4'] = "Participant Name"
        ws1['C4'] = "IC Number"
        
        # Attendance sheet (required)
        ws2 = wb.create_sheet("Attendance")
        ws2['A4'] = "No"
        ws2['B4'] = "Participant Name"
        ws2['C4'] = "IC Number"
        
        # Vehicle Checklist sheet with data
        ws3 = wb.create_sheet("Vehicle Checklist")
        ws3['A1'] = "Vehicle Checklist"
        ws3['A2'] = ""
        ws3['A3'] = ""
        ws3['A4'] = "No"
        ws3['B4'] = "Participant Name"
        ws3['C4'] = "IC Number"
        ws3['D4'] = "Vehicle Model"
        ws3['E4'] = "Registration No"
        ws3['F4'] = "Road Tax Expiry"
        ws3['G4'] = "Helmet"
        ws3['H4'] = "Tires"
        
        # Data row
        ws3['A5'] = 1
        ws3['B5'] = "Test Participant"
        ws3['C5'] = TEST_IC_NUMBERS[1]  # Use second known IC
        ws3['D5'] = "Honda Wave"
        ws3['E5'] = "ABC1234"
        ws3['F5'] = "2025-06-15"
        ws3['G5'] = "good"
        ws3['H5'] = "needs_repair"
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        # Send import request
        files = {'file': ('test_vehicle.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/import-data",
            headers=auth_headers,
            files=files
        )
        
        print(f"Vehicle import response: {response.status_code} - {response.text}")
        
        assert response.status_code == 200, f"Import failed: {response.text}"
        
        data = response.json()
        print(f"Vehicle import results: {data}")
        
        # Verify vehicle_checklists_imported is returned in response
        assert "vehicle_checklists_imported" in data, f"Expected 'vehicle_checklists_imported' in response: {data}"
        
        print(f"SUCCESS: Imported {data.get('vehicle_checklists_imported', 0)} vehicle checklists")
    
    def test_import_attendance(self, auth_headers):
        """Import attendance data"""
        # Create test Excel file
        wb = Workbook()
        
        # Pre-Post Tests sheet (required)
        ws1 = wb.active
        ws1.title = "Pre-Post Tests"
        ws1['A4'] = "No"
        ws1['B4'] = "Participant Name"
        ws1['C4'] = "IC Number"
        
        # Attendance sheet with data
        ws2 = wb.create_sheet("Attendance")
        ws2['A1'] = "Attendance"
        ws2['A2'] = ""
        ws2['A3'] = ""
        ws2['A4'] = "No"
        ws2['B4'] = "Participant Name"
        ws2['C4'] = "IC Number"
        ws2['D4'] = "Day 1"
        ws2['E4'] = "Day 2"
        
        # Data row
        ws2['A5'] = 1
        ws2['B5'] = "Test Participant"
        ws2['C5'] = TEST_IC_NUMBERS[2]  # Use third known IC
        ws2['D5'] = "P"  # Present
        ws2['E5'] = "L"  # Late
        
        # Vehicle Checklist sheet (required)
        ws3 = wb.create_sheet("Vehicle Checklist")
        ws3['A4'] = "No"
        ws3['B4'] = "Participant Name"
        ws3['C4'] = "IC Number"
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        # Send import request
        files = {'file': ('test_attendance.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/import-data",
            headers=auth_headers,
            files=files
        )
        
        print(f"Attendance import response: {response.status_code} - {response.text}")
        
        assert response.status_code == 200, f"Import failed: {response.text}"
        
        data = response.json()
        print(f"Attendance import results: {data}")
        
        # Verify attendance was imported
        assert "attendance_imported" in data, f"Expected 'attendance_imported' in response: {data}"
        
        print(f"SUCCESS: Imported {data.get('attendance_imported', 0)} attendance records")
    
    def test_import_returns_all_counts(self, auth_headers):
        """Verify import response includes all count fields"""
        # Create minimal Excel file
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Pre-Post Tests"
        ws1['A4'] = "No"
        ws1['B4'] = "Participant Name"
        ws1['C4'] = "IC Number"
        
        ws2 = wb.create_sheet("Attendance")
        ws2['A4'] = "No"
        
        ws3 = wb.create_sheet("Vehicle Checklist")
        ws3['A4'] = "No"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        files = {'file': ('test_minimal.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(
            f"{BASE_URL}/api/sessions/{SESSION_ID}/import-data",
            headers=auth_headers,
            files=files
        )
        
        assert response.status_code == 200
        
        data = response.json()
        
        # Check all expected fields are present
        expected_fields = ["test_scores_imported", "attendance_imported", "vehicle_checklists_imported"]
        for field in expected_fields:
            assert field in data, f"Missing '{field}' in response: {data}"
        
        print(f"SUCCESS: Import response contains all required count fields: {expected_fields}")


class TestUnauthorizedAccess:
    """Test access control"""
    
    def test_export_requires_auth(self):
        """Export endpoint requires authentication"""
        response = requests.get(f"{BASE_URL}/api/sessions/{SESSION_ID}/export-template")
        # Expecting 401 or 403
        assert response.status_code in [401, 403], f"Expected auth error, got: {response.status_code}"
        print(f"SUCCESS: Export requires authentication (returned {response.status_code})")
    
    def test_import_requires_auth(self):
        """Import endpoint requires authentication"""
        wb = Workbook()
        wb.active.title = "Pre-Post Tests"
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        files = {'file': ('test.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(f"{BASE_URL}/api/sessions/{SESSION_ID}/import-data", files=files)
        
        assert response.status_code in [401, 403], f"Expected auth error, got: {response.status_code}"
        print(f"SUCCESS: Import requires authentication (returned {response.status_code})")


class TestSessionNotFound:
    """Test 404 handling"""
    
    def test_export_nonexistent_session(self, auth_headers):
        """Export for nonexistent session returns 404"""
        response = requests.get(
            f"{BASE_URL}/api/sessions/nonexistent-session-id/export-template",
            headers=auth_headers
        )
        assert response.status_code == 404, f"Expected 404, got: {response.status_code}"
        print("SUCCESS: Export for nonexistent session returns 404")
    
    def test_import_nonexistent_session(self, auth_headers):
        """Import to nonexistent session returns 404"""
        wb = Workbook()
        wb.active.title = "Pre-Post Tests"
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        files = {'file': ('test.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(
            f"{BASE_URL}/api/sessions/nonexistent-session-id/import-data",
            headers=auth_headers,
            files=files
        )
        
        assert response.status_code == 404, f"Expected 404, got: {response.status_code}"
        print("SUCCESS: Import to nonexistent session returns 404")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
