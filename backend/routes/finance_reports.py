"""
Finance Reports routes - P&L, Subledgers, Chart of Accounts, General Ledger
Stage F5: ~12 endpoints
"""
from fastapi import APIRouter, HTTPException, Depends
from typing import Optional

from core import db, get_current_user, get_malaysia_time
from models import User

router = APIRouter(prefix="/finance", tags=["finance-reports"])


# Chart of Accounts - Static configuration
CHART_OF_ACCOUNTS = {
    # Assets (1xxx)
    "1001": {"name": "Cash at Bank", "type": "Asset"},
    "1002": {"name": "Petty Cash", "type": "Asset"},
    "1100": {"name": "Accounts Receivable", "type": "Asset"},
    
    # Liabilities (2xxx)
    "2001": {"name": "Accounts Payable", "type": "Liability"},
    "2100": {"name": "Trainer Payable", "type": "Liability"},
    "2101": {"name": "Coordinator Payable", "type": "Liability"},
    "2102": {"name": "Marketing Commission Payable", "type": "Liability"},
    "2200": {"name": "EPF Payable", "type": "Liability"},
    "2201": {"name": "SOCSO Payable", "type": "Liability"},
    "2202": {"name": "EIS Payable", "type": "Liability"},
    "2210": {"name": "Salary Payable", "type": "Liability"},
    
    # Income (4xxx)
    "4000": {"name": "Training Income - General", "type": "Income"},
    "4001": {"name": "Training Income - Cars", "type": "Income"},
    "4002": {"name": "Training Income - Motorcycles", "type": "Income"},
    "4003": {"name": "Training Income - Heavy Vehicles", "type": "Income"},
    "4004": {"name": "Training Income - Bus", "type": "Income"},
    "4100": {"name": "Other Income", "type": "Income"},
    
    # Expenses (5xxx)
    "5001": {"name": "Trainer Fees", "type": "Expense"},
    "5002": {"name": "Coordinator Fees", "type": "Expense"},
    "5003": {"name": "Marketing Commission", "type": "Expense"},
    "5100": {"name": "Staff Salaries", "type": "Expense"},
    "5101": {"name": "EPF - Employer", "type": "Expense"},
    "5102": {"name": "SOCSO - Employer", "type": "Expense"},
    "5103": {"name": "EIS - Employer", "type": "Expense"},
    "5200": {"name": "F&B Expenses", "type": "Expense"},
    "5201": {"name": "Venue Expenses", "type": "Expense"},
    "5202": {"name": "HRDCorp Levy", "type": "Expense"},
    "5300": {"name": "Petty Cash Expenses", "type": "Expense"},
    "5400": {"name": "Other Expenses", "type": "Expense"},
}


@router.get("/profit-loss")
async def get_profit_loss_report(
    year: int = None,
    month: int = None,
    current_user: User = Depends(get_current_user)
):
    """Get Profit/Loss report - monthly breakdown and YTD
    
    REVENUE RECOGNITION:
    - Revenue is recognized when invoice is issued:
      * invoice.status IN ["issued", "partial", "paid"]
    - SST/Tax is excluded from revenue (gross revenue = total - tax)
    - Session completion status is NOT required for revenue recognition
    
    EXPENSE POLICY (Improvement 2 - Actuals Only):
    - Session expenses use ONLY actual_amount
    - No fallback to estimated_amount
    - This may show higher profit if actuals not entered (expected behavior)
    
    All session-related expenses are attributed to the month based on 
    the SESSION'S START DATE, not the record's created_at date.
    """
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    now = get_malaysia_time()
    year = year or now.year
    
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    # Get sessions for the year to map expenses to correct months
    # Include completion_status for revenue recognition
    sessions = await db.sessions.find({
        "start_date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0, "id": 1, "start_date": 1, "invoice_id": 1, "completion_status": 1}).to_list(10000)
    
    session_date_map = {}
    session_completion_map = {}  # Track completion status
    # ========== MULTI-INVOICE SUPPORT (Improvement) ==========
    # Build session lookup by ID - invoices will be mapped via invoice.session_id
    for s in sessions:
        session_date_map[s.get("id")] = s.get("start_date", "")
        # Backward compatibility: missing completion_status is treated as eligible (completed)
        session_completion_map[s.get("id")] = s.get("completion_status", "completed")
    # ========== END MULTI-INVOICE SUPPORT ==========
    
    # Get all data sources
    manual_income = await db.manual_income.find({"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(1000)
    payslips = await db.hr_payslips.find({"year": year}, {"_id": 0}).to_list(1000)
    pay_advice = await db.hr_pay_advices.find({"year": year}, {"_id": 0}).to_list(1000)
    all_trainer_fees = await db.trainer_fees.find({}, {"_id": 0}).to_list(10000)
    all_coordinator_fees = await db.coordinator_fees.find({}, {"_id": 0}).to_list(10000)
    all_session_expenses = await db.session_expenses.find({}, {"_id": 0}).to_list(10000)
    all_marketing_commissions = await db.marketing_commissions.find({"status": {"$in": ["pending", "approved", "paid"]}}, {"_id": 0}).to_list(10000)
    petty_cash = await db.petty_cash_transactions.find({"date": {"$gte": start_date, "$lte": end_date}, "type": "expense", "status": "approved"}, {"_id": 0}).to_list(1000)
    manual_expenses = await db.manual_expenses.find({"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(1000)
    
    # Build monthly breakdown
    monthly_data = {}
    for m in range(1, 13):
        monthly_data[m] = {
            "month": m,
            "month_name": ["", "January", "February", "March", "April", "May", "June", 
                          "July", "August", "September", "October", "November", "December"][m],
            "income": {"invoices": 0, "manual": 0, "total": 0},
            "expenses": {"payroll": 0, "session_workers": 0, "marketing_commissions": 0, 
                        "session_expenses": 0, "petty_cash": 0, "manual": 0, "total": 0},
            "net_profit": 0
        }
    
    # ELIGIBLE invoice statuses for revenue recognition
    REVENUE_INVOICE_STATUSES = ["issued", "partial", "paid"]
    
    # ========== MULTI-INVOICE SUPPORT ==========
    # Build a comprehensive invoice-to-session map using both directions
    invoice_session_map = {}
    all_invoices = await db.invoices.find({}, {"_id": 0}).to_list(10000)
    for inv in all_invoices:
        inv_id = inv.get("id")
        # Direct session_id on invoice (multi-invoice support)
        if inv.get("session_id") and inv["session_id"] in session_date_map:
            invoice_session_map[inv_id] = inv["session_id"]
    # Also add from session.invoice_id (legacy single invoice)
    for s in sessions:
        if s.get("invoice_id") and s["invoice_id"] not in invoice_session_map:
            invoice_session_map[s["invoice_id"]] = s["id"]
    
    # Process invoices (income) - revenue recognized when payment received (cash-basis)
    for inv in all_invoices:
        try:
            if inv.get("status") not in REVENUE_INVOICE_STATUSES:
                continue
            
            inv_id = inv.get("id")
            session_id = invoice_session_map.get(inv_id) or inv.get("session_id")
            
            # Calculate gross revenue (exclude SST/tax)
            total_amount = float(inv.get("total_amount") or inv.get("amount") or 0)
            tax_amount = float(inv.get("tax_amount") or inv.get("sst_amount") or 0)
            gross_revenue = total_amount - tax_amount
            
            # Attribute to month based on session start_date or invoice created_at
            if session_id and session_id in session_date_map:
                session_date = session_date_map[session_id]
                if session_date.startswith(str(year)):
                    inv_month = int(session_date[5:7])
                    monthly_data[inv_month]["income"]["invoices"] += gross_revenue
            else:
                # Fallback to invoice created_at for non-session invoices
                inv_date = inv.get("created_at", "")[:10]
                if inv_date.startswith(str(year)):
                    inv_month = int(inv_date[5:7]) if len(inv_date) >= 7 else 1
                    monthly_data[inv_month]["income"]["invoices"] += gross_revenue
        except Exception:
            pass
    # ========== END MULTI-INVOICE SUPPORT ==========
    
    # Process manual income
    for inc in manual_income:
        try:
            inc_month = int(inc.get("date", "")[5:7])
            monthly_data[inc_month]["income"]["manual"] += float(inc.get("amount", 0))
        except Exception:
            pass
    
    # Process payroll
    for ps in payslips:
        try:
            ps_month = ps.get("month", 1)
            gross = float(ps.get("gross_salary", 0))
            epf_er = float(ps.get("epf_employer", 0))
            socso_er = float(ps.get("socso_employer", 0))
            eis_er = float(ps.get("eis_employer", 0))
            monthly_data[ps_month]["expenses"]["payroll"] += gross + epf_er + socso_er + eis_er
        except Exception:
            pass
    
    # Process pay advice
    for pa in pay_advice:
        try:
            pa_month = pa.get("month", 1)
            monthly_data[pa_month]["expenses"]["session_workers"] += float(pa.get("amount", 0))
        except Exception:
            pass
    
    # Process trainer fees
    for tf in all_trainer_fees:
        try:
            session_id = tf.get("session_id")
            session_date = session_date_map.get(session_id, "")
            if not session_date or not session_date.startswith(str(year)):
                continue
            tf_month = int(session_date[5:7]) if len(session_date) >= 7 else 1
            monthly_data[tf_month]["expenses"]["session_workers"] += float(tf.get("fee_amount") or 0)
        except Exception:
            pass
    
    # Process coordinator fees
    for cf in all_coordinator_fees:
        try:
            session_id = cf.get("session_id")
            session_date = session_date_map.get(session_id, "")
            if not session_date or not session_date.startswith(str(year)):
                continue
            cf_month = int(session_date[5:7]) if len(session_date) >= 7 else 1
            monthly_data[cf_month]["expenses"]["session_workers"] += float(cf.get("total_fee") or 0)
        except Exception:
            pass
    
    # Process session expenses - ACTUALS ONLY (Improvement 2)
    for exp in all_session_expenses:
        try:
            session_id = exp.get("session_id")
            session_date = session_date_map.get(session_id, "")
            if not session_date or not session_date.startswith(str(year)):
                continue
            exp_month = int(session_date[5:7]) if len(session_date) >= 7 else 1
            # ACTUALS ONLY: Only use actual_amount, never fall back to estimated
            # If actual_amount is 0 or missing, expense is not counted
            # This is expected behavior per user requirement
            amount = float(exp.get("actual_amount") or 0)
            monthly_data[exp_month]["expenses"]["session_expenses"] += amount
        except Exception:
            pass
    
    # Process marketing commissions
    for mc in all_marketing_commissions:
        try:
            session_id = mc.get("session_id")
            session_date = session_date_map.get(session_id, "")
            if not session_date or not session_date.startswith(str(year)):
                continue
            mc_month = int(session_date[5:7]) if len(session_date) >= 7 else 1
            monthly_data[mc_month]["expenses"]["marketing_commissions"] += float(mc.get("calculated_amount") or 0)
        except Exception:
            pass
    
    # Process petty cash
    for pc in petty_cash:
        try:
            pc_month = int(pc.get("date", "")[5:7])
            monthly_data[pc_month]["expenses"]["petty_cash"] += float(pc.get("amount", 0))
        except Exception:
            pass
    
    # Process manual expenses
    for exp in manual_expenses:
        try:
            exp_month = int(exp.get("date", "")[5:7])
            monthly_data[exp_month]["expenses"]["manual"] += float(exp.get("amount", 0))
        except Exception:
            pass
    
    # Calculate totals
    ytd_income = 0
    ytd_expenses = 0
    
    for m in range(1, 13):
        md = monthly_data[m]
        md["income"]["invoices"] = round(md["income"]["invoices"], 2)
        md["income"]["manual"] = round(md["income"]["manual"], 2)
        md["income"]["total"] = round(md["income"]["invoices"] + md["income"]["manual"], 2)
        md["expenses"]["payroll"] = round(md["expenses"]["payroll"], 2)
        md["expenses"]["session_workers"] = round(md["expenses"]["session_workers"], 2)
        md["expenses"]["marketing_commissions"] = round(md["expenses"]["marketing_commissions"], 2)
        md["expenses"]["session_expenses"] = round(md["expenses"]["session_expenses"], 2)
        md["expenses"]["petty_cash"] = round(md["expenses"]["petty_cash"], 2)
        md["expenses"]["manual"] = round(md["expenses"]["manual"], 2)
        md["expenses"]["total"] = round(sum([md["expenses"]["payroll"], md["expenses"]["session_workers"],
                                       md["expenses"]["marketing_commissions"], md["expenses"]["session_expenses"],
                                       md["expenses"]["petty_cash"], md["expenses"]["manual"]]), 2)
        md["net_profit"] = round(md["income"]["total"] - md["expenses"]["total"], 2)
        ytd_income += md["income"]["total"]
        ytd_expenses += md["expenses"]["total"]
    
    ytd_income = round(ytd_income, 2)
    ytd_expenses = round(ytd_expenses, 2)
    
    return {
        "year": year,
        "monthly_breakdown": list(monthly_data.values()),
        "ytd_summary": {
            "total_income": ytd_income,
            "total_expenses": ytd_expenses,
            "net_profit": round(ytd_income - ytd_expenses, 2),
            "profit_margin": round((ytd_income - ytd_expenses) / ytd_income * 100, 2) if ytd_income > 0 else 0
        },
        "expense_breakdown": {
            "payroll": round(sum(md["expenses"]["payroll"] for md in monthly_data.values()), 2),
            "session_workers": round(sum(md["expenses"]["session_workers"] for md in monthly_data.values()), 2),
            "marketing_commissions": round(sum(md["expenses"]["marketing_commissions"] for md in monthly_data.values()), 2),
            "session_expenses": round(sum(md["expenses"]["session_expenses"] for md in monthly_data.values()), 2),
            "petty_cash": round(sum(md["expenses"]["petty_cash"] for md in monthly_data.values()), 2),
            "manual": round(sum(md["expenses"]["manual"] for md in monthly_data.values()), 2)
        }
    }


@router.get("/profit-loss/by-programme")
async def get_profit_loss_by_programme(
    year: int = None,
    current_user: User = Depends(get_current_user)
):
    """Get Profit/Loss report broken down by programme (dynamic)."""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    now = get_malaysia_time()
    year = year or now.year
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    programmes = await db.programs.find({}, {"_id": 0, "id": 1, "name": 1, "category": 1}).to_list(100)
    
    sessions = await db.sessions.find({"start_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(10000)
    
    session_to_programme = {}
    invoice_to_session = {}
    
    for s in sessions:
        sid = s.get("id")
        session_to_programme[sid] = s.get("program_id")
        # Legacy: session.invoice_id (single invoice)
        if s.get("invoice_id"):
            invoice_to_session[s.get("invoice_id")] = sid
    
    # ========== MULTI-INVOICE SUPPORT ==========
    # Also map invoices that have session_id on the invoice document
    # This catches additional invoices not stored in session.invoice_id
    all_invoices = await db.invoices.find({}, {"_id": 0}).to_list(10000)
    for inv in all_invoices:
        inv_session_id = inv.get("session_id")
        if inv_session_id and inv_session_id in session_to_programme:
            inv_id = inv.get("id")
            if inv_id and inv_id not in invoice_to_session:
                invoice_to_session[inv_id] = inv_session_id
    # ========== END MULTI-INVOICE SUPPORT ==========
    
    programme_data = {}
    for prog in programmes:
        prog_name = prog.get("name") or prog.get("programme_name") or "Unknown Programme"
        programme_data[prog["id"]] = {
            "programme_id": prog["id"],
            "programme_name": prog_name,
            "category": prog.get("category", ""),
            "income": 0,
            "expenses": {"trainer_fees": 0, "coordinator_fees": 0, "marketing_commissions": 0, "session_expenses": 0, "total": 0},
            "gross_profit": 0,
            "gross_margin_pct": 0,
            "session_count": 0
        }
    
    programme_data["_other"] = {
        "programme_id": "_other",
        "programme_name": "Other / Unassigned",
        "category": "Other",
        "income": 0,
        "expenses": {"trainer_fees": 0, "coordinator_fees": 0, "marketing_commissions": 0, "session_expenses": 0, "total": 0},
        "gross_profit": 0,
        "gross_margin_pct": 0,
        "session_count": 0
    }
    
    for s in sessions:
        prog_id = s.get("program_id") or "_other"
        # Create entry for orphaned programmes (programme was deleted from DB)
        if prog_id not in programme_data and prog_id != "_other":
            programme_data[prog_id] = {
                "programme_id": prog_id,
                "programme_name": s.get("name") or s.get("programme_name") or "Unknown Programme",
                "category": "",
                "income": 0,
                "expenses": {"trainer_fees": 0, "coordinator_fees": 0, "marketing_commissions": 0, "session_expenses": 0, "total": 0},
                "gross_profit": 0,
                "gross_margin_pct": 0,
                "session_count": 0
            }
        if prog_id in programme_data:
            programme_data[prog_id]["session_count"] += 1
    
    invoices = await db.invoices.find({"status": {"$in": ["approved", "issued", "paid"]}}, {"_id": 0}).to_list(10000)
    
    for inv in invoices:
        try:
            amount = float(inv.get("total_amount") or inv.get("amount") or 0)
            inv_id = inv.get("id")
            
            # Try multiple ways to find the session for this invoice
            session_id = invoice_to_session.get(inv_id)
            if not session_id:
                # Direct session_id on the invoice document (multi-invoice support)
                session_id = inv.get("session_id")
            
            prog_id = session_to_programme.get(session_id, "_other") if session_id else "_other"
            
            if session_id and session_id in session_to_programme:
                if prog_id in programme_data:
                    programme_data[prog_id]["income"] += amount
                else:
                    programme_data["_other"]["income"] += amount
            else:
                inv_date = inv.get("created_at", "")[:10]
                if inv_date.startswith(str(year)):
                    programme_data["_other"]["income"] += amount
        except Exception:
            pass
    
    trainer_fees = await db.trainer_fees.find({}, {"_id": 0}).to_list(10000)
    for tf in trainer_fees:
        try:
            session_id = tf.get("session_id")
            if session_id not in session_to_programme:
                continue
            prog_id = session_to_programme.get(session_id) or "_other"
            if prog_id not in programme_data:
                prog_id = "_other"
            programme_data[prog_id]["expenses"]["trainer_fees"] += float(tf.get("fee_amount") or 0)
        except Exception:
            pass
    
    coordinator_fees = await db.coordinator_fees.find({}, {"_id": 0}).to_list(10000)
    for cf in coordinator_fees:
        try:
            session_id = cf.get("session_id")
            if session_id not in session_to_programme:
                continue
            prog_id = session_to_programme.get(session_id) or "_other"
            if prog_id not in programme_data:
                prog_id = "_other"
            programme_data[prog_id]["expenses"]["coordinator_fees"] += float(cf.get("total_fee") or 0)
        except Exception:
            pass
    
    marketing_comms = await db.marketing_commissions.find({"status": {"$in": ["pending", "approved", "paid"]}}, {"_id": 0}).to_list(10000)
    for mc in marketing_comms:
        try:
            session_id = mc.get("session_id")
            if session_id not in session_to_programme:
                continue
            prog_id = session_to_programme.get(session_id) or "_other"
            if prog_id not in programme_data:
                prog_id = "_other"
            programme_data[prog_id]["expenses"]["marketing_commissions"] += float(mc.get("calculated_amount") or 0)
        except Exception:
            pass
    
    session_expenses = await db.session_expenses.find({}, {"_id": 0}).to_list(10000)
    for exp in session_expenses:
        try:
            session_id = exp.get("session_id")
            if session_id not in session_to_programme:
                continue
            prog_id = session_to_programme.get(session_id) or "_other"
            if prog_id not in programme_data:
                prog_id = "_other"
            amount = float(exp.get("actual_amount") or exp.get("estimated_amount") or exp.get("amount") or 0)
            programme_data[prog_id]["expenses"]["session_expenses"] += amount
        except Exception:
            pass
    
    total_income = 0
    total_direct_expenses = 0
    
    for prog_id, data in programme_data.items():
        data["income"] = round(data["income"], 2)
        data["expenses"]["trainer_fees"] = round(data["expenses"]["trainer_fees"], 2)
        data["expenses"]["coordinator_fees"] = round(data["expenses"]["coordinator_fees"], 2)
        data["expenses"]["marketing_commissions"] = round(data["expenses"]["marketing_commissions"], 2)
        data["expenses"]["session_expenses"] = round(data["expenses"]["session_expenses"], 2)
        data["expenses"]["total"] = round(sum([data["expenses"]["trainer_fees"], data["expenses"]["coordinator_fees"],
                                         data["expenses"]["marketing_commissions"], data["expenses"]["session_expenses"]]), 2)
        data["gross_profit"] = round(data["income"] - data["expenses"]["total"], 2)
        data["gross_margin_pct"] = round((data["gross_profit"] / data["income"] * 100), 2) if data["income"] > 0 else 0
        total_income += data["income"]
        total_direct_expenses += data["expenses"]["total"]
    
    total_income = round(total_income, 2)
    total_direct_expenses = round(total_direct_expenses, 2)
    
    payslips = await db.hr_payslips.find({"year": year}, {"_id": 0}).to_list(1000)
    overhead_payroll = round(sum(float(ps.get("gross_salary", 0)) + float(ps.get("epf_employer", 0)) + 
                          float(ps.get("socso_employer", 0)) + float(ps.get("eis_employer", 0)) for ps in payslips), 2)
    
    petty_cash = await db.petty_cash_transactions.find({"date": {"$gte": start_date, "$lte": end_date}, "type": "expense", "status": "approved"}, {"_id": 0}).to_list(1000)
    overhead_petty_cash = round(sum(float(pc.get("amount", 0)) for pc in petty_cash), 2)
    
    manual_expenses = await db.manual_expenses.find({"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(1000)
    overhead_manual = round(sum(float(exp.get("amount", 0)) for exp in manual_expenses), 2)
    
    manual_income = await db.manual_income.find({"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(1000)
    other_income = round(sum(float(inc.get("amount", 0)) for inc in manual_income), 2)
    
    total_overhead = round(overhead_payroll + overhead_petty_cash + overhead_manual, 2)
    total_expenses = round(total_direct_expenses + total_overhead, 2)
    net_profit = round(total_income + other_income - total_expenses, 2)
    
    active_programmes = [data for data in programme_data.values() if data["income"] > 0 or data["expenses"]["total"] > 0]
    active_programmes.sort(key=lambda x: x["income"], reverse=True)
    
    return {
        "year": year,
        "programmes": active_programmes,
        "summary": {
            "total_programme_income": round(total_income, 2),
            "other_income": round(other_income, 2),
            "total_income": round(total_income + other_income, 2),
            "total_direct_costs": round(total_direct_expenses, 2),
            "gross_profit": round(total_income - total_direct_expenses, 2),
            "gross_margin_pct": round((total_income - total_direct_expenses) / total_income * 100, 2) if total_income > 0 else 0,
            "overhead": {"payroll": overhead_payroll, "petty_cash": overhead_petty_cash, "manual": overhead_manual, "total": total_overhead},
            "total_expenses": round(total_expenses, 2),
            "net_profit": round(net_profit, 2),
            "net_margin_pct": round(net_profit / (total_income + other_income) * 100, 2) if (total_income + other_income) > 0 else 0
        }
    }


@router.get("/subledger/trainers")
async def get_trainer_subledger(year: int = None, current_user: User = Depends(get_current_user)):
    """Get Trainer & Coordinator Sub-ledger - aggregated by person"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    now = get_malaysia_time()
    year = year or now.year
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    sessions = await db.sessions.find({"start_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0, "id": 1, "start_date": 1, "program_id": 1}).to_list(10000)
    session_ids = {s["id"] for s in sessions}
    session_map = {s["id"]: s for s in sessions}
    
    programmes = await db.programs.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    programme_map = {p["id"]: p.get("name", "Unknown") for p in programmes}
    
    users = await db.users.find({}, {"_id": 0, "id": 1, "full_name": 1}).to_list(1000)
    user_map = {u["id"]: u.get("full_name", "Unknown") for u in users}
    
    trainer_fees = await db.trainer_fees.find({}, {"_id": 0}).to_list(10000)
    
    trainer_data = {}
    for tf in trainer_fees:
        session_id = tf.get("session_id")
        if session_id not in session_ids:
            continue
        trainer_id = tf.get("trainer_id")
        if not trainer_id:
            continue
        
        if trainer_id not in trainer_data:
            trainer_data[trainer_id] = {"user_id": trainer_id, "name": user_map.get(trainer_id, tf.get("trainer_name", "Unknown")),
                                        "role": "Trainer", "total_earned": 0, "total_paid": 0, "balance": 0, "sessions": []}
        
        session = session_map.get(session_id, {})
        amount = float(tf.get("fee_amount") or 0)
        is_paid = tf.get("status") == "paid"
        
        trainer_data[trainer_id]["total_earned"] += amount
        if is_paid:
            trainer_data[trainer_id]["total_paid"] += amount
        
        trainer_data[trainer_id]["sessions"].append({
            "session_id": session_id, "date": session.get("start_date", ""),
            "programme": programme_map.get(session.get("program_id"), "Unknown"),
            "amount": amount, "status": tf.get("status", "pending")
        })
    
    coordinator_fees = await db.coordinator_fees.find({}, {"_id": 0}).to_list(10000)
    
    coordinator_data = {}
    for cf in coordinator_fees:
        session_id = cf.get("session_id")
        if session_id not in session_ids:
            continue
        coord_id = cf.get("coordinator_id")
        if not coord_id:
            continue
        
        if coord_id not in coordinator_data:
            coordinator_data[coord_id] = {"user_id": coord_id, "name": user_map.get(coord_id, cf.get("coordinator_name", "Unknown")),
                                          "role": "Coordinator", "total_earned": 0, "total_paid": 0, "balance": 0, "sessions": []}
        
        session = session_map.get(session_id, {})
        amount = float(cf.get("total_fee") or 0)
        is_paid = cf.get("status") == "paid"
        
        coordinator_data[coord_id]["total_earned"] += amount
        if is_paid:
            coordinator_data[coord_id]["total_paid"] += amount
        
        coordinator_data[coord_id]["sessions"].append({
            "session_id": session_id, "date": session.get("start_date", ""),
            "programme": programme_map.get(session.get("program_id"), "Unknown"),
            "amount": amount, "status": cf.get("status", "pending")
        })
    
    for data in trainer_data.values():
        data["balance"] = data["total_earned"] - data["total_paid"]
        data["sessions"].sort(key=lambda x: x["date"], reverse=True)
    
    for data in coordinator_data.values():
        data["balance"] = data["total_earned"] - data["total_paid"]
        data["sessions"].sort(key=lambda x: x["date"], reverse=True)
    
    trainers = sorted(trainer_data.values(), key=lambda x: x["total_earned"], reverse=True)
    coordinators = sorted(coordinator_data.values(), key=lambda x: x["total_earned"], reverse=True)
    
    return {
        "year": year, "trainers": trainers, "coordinators": coordinators,
        "totals": {
            "trainer_earned": sum(t["total_earned"] for t in trainers),
            "trainer_paid": sum(t["total_paid"] for t in trainers),
            "trainer_balance": sum(t["balance"] for t in trainers),
            "coordinator_earned": sum(c["total_earned"] for c in coordinators),
            "coordinator_paid": sum(c["total_paid"] for c in coordinators),
            "coordinator_balance": sum(c["balance"] for c in coordinators)
        }
    }


@router.get("/subledger/marketing")
async def get_marketing_subledger(year: int = None, current_user: User = Depends(get_current_user)):
    """Get Marketing Commission Sub-ledger - aggregated by marketer"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    now = get_malaysia_time()
    year = year or now.year
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    sessions = await db.sessions.find({"start_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0, "id": 1, "start_date": 1, "program_id": 1, "company_name": 1}).to_list(10000)
    session_ids = {s["id"] for s in sessions}
    session_map = {s["id"]: s for s in sessions}
    
    programmes = await db.programs.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    programme_map = {p["id"]: p.get("name", "Unknown") for p in programmes}
    
    users = await db.users.find({}, {"_id": 0, "id": 1, "full_name": 1}).to_list(1000)
    user_map = {u["id"]: u.get("full_name", "Unknown") for u in users}
    
    commissions = await db.marketing_commissions.find({}, {"_id": 0}).to_list(10000)
    
    marketer_data = {}
    for mc in commissions:
        session_id = mc.get("session_id")
        if session_id not in session_ids:
            continue
        marketer_id = mc.get("marketing_user_id") or mc.get("user_id")
        if not marketer_id:
            continue
        
        if marketer_id not in marketer_data:
            marketer_data[marketer_id] = {"user_id": marketer_id, "name": user_map.get(marketer_id, mc.get("marketer_name", "Unknown")),
                                          "total_commission": 0, "total_paid": 0, "balance": 0, "clients": []}
        
        session = session_map.get(session_id, {})
        amount = float(mc.get("calculated_amount") or 0)
        is_paid = mc.get("status") == "paid"
        
        marketer_data[marketer_id]["total_commission"] += amount
        if is_paid:
            marketer_data[marketer_id]["total_paid"] += amount
        
        marketer_data[marketer_id]["clients"].append({
            "session_id": session_id, "date": session.get("start_date", ""),
            "client": session.get("company_name", "Unknown"),
            "programme": programme_map.get(session.get("program_id"), "Unknown"),
            "commission_rate": mc.get("commission_rate", 0), "amount": amount, "status": mc.get("status", "pending")
        })
    
    for data in marketer_data.values():
        data["balance"] = data["total_commission"] - data["total_paid"]
        data["clients"].sort(key=lambda x: x["date"], reverse=True)
    
    marketers = sorted(marketer_data.values(), key=lambda x: x["total_commission"], reverse=True)
    
    return {
        "year": year, "marketers": marketers,
        "totals": {"total_commission": sum(m["total_commission"] for m in marketers),
                   "total_paid": sum(m["total_paid"] for m in marketers),
                   "total_balance": sum(m["balance"] for m in marketers)}
    }


@router.get("/subledger/payroll")
async def get_payroll_subledger(year: int = None, current_user: User = Depends(get_current_user)):
    """Get Staff Payroll Register - aggregated by employee"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    now = get_malaysia_time()
    year = year or now.year
    
    payslips = await db.hr_payslips.find({"year": year}, {"_id": 0}).to_list(1000)
    staff = await db.hr_staff.find({}, {"_id": 0}).to_list(1000)
    staff_map = {s["id"]: s for s in staff}
    
    employee_data = {}
    for ps in payslips:
        staff_id = ps.get("staff_id")
        if not staff_id:
            continue
        
        if staff_id not in employee_data:
            staff_info = staff_map.get(staff_id, {})
            employee_data[staff_id] = {
                "staff_id": staff_id, "name": ps.get("full_name") or staff_info.get("full_name", "Unknown"),
                "employee_id": staff_info.get("employee_id", ""), "designation": staff_info.get("designation", ""),
                "total_gross": 0, "total_epf": 0, "total_socso": 0, "total_eis": 0, "total_net": 0, "months": []
            }
        
        employee_data[staff_id]["total_gross"] += float(ps.get("gross_salary", 0))
        employee_data[staff_id]["total_epf"] += float(ps.get("epf_employee", 0))
        employee_data[staff_id]["total_socso"] += float(ps.get("socso_employee", 0))
        employee_data[staff_id]["total_eis"] += float(ps.get("eis_employee", 0))
        employee_data[staff_id]["total_net"] += float(ps.get("nett_pay", 0))
        
        month_names = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        employee_data[staff_id]["months"].append({
            "month": ps.get("month"), "month_name": month_names[ps.get("month", 1)],
            "gross": float(ps.get("gross_salary", 0)), "epf": float(ps.get("epf_employee", 0)),
            "socso": float(ps.get("socso_employee", 0)), "eis": float(ps.get("eis_employee", 0)),
            "net": float(ps.get("nett_pay", 0))
        })
    
    for data in employee_data.values():
        data["months"].sort(key=lambda x: x["month"])
    
    employees = sorted(employee_data.values(), key=lambda x: x["name"])
    
    return {
        "year": year, "employees": employees,
        "totals": {"total_gross": sum(e["total_gross"] for e in employees),
                   "total_epf": sum(e["total_epf"] for e in employees),
                   "total_socso": sum(e["total_socso"] for e in employees),
                   "total_eis": sum(e["total_eis"] for e in employees),
                   "total_net": sum(e["total_net"] for e in employees)}
    }


@router.get("/chart-of-accounts")
async def get_chart_of_accounts(current_user: User = Depends(get_current_user)):
    """Get the Chart of Accounts"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    accounts = []
    for code, info in sorted(CHART_OF_ACCOUNTS.items()):
        accounts.append({"code": code, "name": info["name"], "type": info["type"]})
    return accounts


@router.get("/general-ledger")
async def get_general_ledger(year: int = None, month: int = None, current_user: User = Depends(get_current_user)):
    """Get General Ledger with double-entry transactions."""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    now = get_malaysia_time()
    year = year or now.year
    
    if month:
        start_date = f"{year}-{month:02d}-01"
        end_date = f"{year + 1}-01-01" if month == 12 else f"{year}-{month + 1:02d}-01"
    else:
        start_date = f"{year}-01-01"
        end_date = f"{year}-12-31"
    
    gl_entries = []
    entry_id = 1
    
    programmes = await db.programs.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    programme_map = {p["id"]: p.get("name", "Unknown") for p in programmes}
    
    sessions = await db.sessions.find({"start_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(10000)
    session_map = {s.get("id"): s for s in sessions}
    session_ids = set(session_map.keys())
    
    # Pre-load invoice numbers for all sessions (for TF/CF/MC references)
    all_invoices = await db.invoices.find({}, {"_id": 0, "id": 1, "invoice_number": 1, "session_id": 1}).to_list(10000)
    session_invoice_map = {}
    for inv in all_invoices:
        sid = inv.get("session_id")
        if sid and inv.get("invoice_number"):
            session_invoice_map[sid] = inv["invoice_number"]
    # Also check session.invoice_id mappings
    for s in sessions:
        sid = s.get("id")
        if sid not in session_invoice_map and s.get("invoice_id"):
            for inv in all_invoices:
                if inv.get("id") == s["invoice_id"] and inv.get("invoice_number"):
                    session_invoice_map[sid] = inv["invoice_number"]
                    break
    
    # 1. INVOICES - DR Accounts Receivable, CR Training Income
    invoices = await db.invoices.find({"status": {"$in": ["approved", "issued", "paid"]}}, {"_id": 0}).to_list(10000)
    
    for inv in invoices:
        try:
            session_id = None
            session = None
            
            for s in sessions:
                if s.get("invoice_id") == inv.get("id"):
                    session_id = s.get("id")
                    session = s
                    break
            
            if not session:
                inv_date = inv.get("created_at", "")[:10]
                if not (inv_date >= start_date and inv_date <= end_date):
                    continue
            
            amount = float(inv.get("total_amount") or inv.get("amount") or 0)
            if amount <= 0:
                continue
            
            trans_date = session.get("start_date") if session else inv.get("created_at", "")[:10]
            ref = inv.get("invoice_number", f"INV-{inv.get('id', '')[:8]}")
            programme = programme_map.get(session.get("program_id"), "General") if session else "General"
            
            income_account = "4000"
            prog_name = programme.lower() if programme else ""
            if "car" in prog_name:
                income_account = "4001"
            elif "motorcycle" in prog_name or "motor" in prog_name:
                income_account = "4002"
            elif "heavy" in prog_name or "truck" in prog_name:
                income_account = "4003"
            elif "bus" in prog_name:
                income_account = "4004"
            
            gl_entries.append({"entry_id": entry_id, "date": trans_date, "reference": ref,
                              "description": f"Invoice issued - {inv.get('company_name', 'Customer')}",
                              "account_code": "1100", "account_name": CHART_OF_ACCOUNTS["1100"]["name"],
                              "debit": amount, "credit": 0, "tags": {"session_id": session_id, "programme": programme}})
            gl_entries.append({"entry_id": entry_id, "date": trans_date, "reference": ref,
                              "description": f"Invoice issued - {inv.get('company_name', 'Customer')}",
                              "account_code": income_account, "account_name": CHART_OF_ACCOUNTS[income_account]["name"],
                              "debit": 0, "credit": amount, "tags": {"session_id": session_id, "programme": programme}})
            entry_id += 1
        except Exception:
            pass
    
    # 2. PAYMENTS RECEIVED - DR Bank, CR Accounts Receivable
    payments = await db.payments.find({}, {"_id": 0}).to_list(10000)
    for pmt in payments:
        try:
            pmt_date = pmt.get("payment_date", pmt.get("created_at", ""))[:10]
            if not (pmt_date >= start_date and pmt_date <= end_date):
                continue
            amount = float(pmt.get("amount", 0))
            if amount <= 0:
                continue
            ref = pmt.get("reference", f"PMT-{pmt.get('id', '')[:8]}")
            
            gl_entries.append({"entry_id": entry_id, "date": pmt_date, "reference": ref,
                              "description": f"Payment received - {pmt.get('payment_method', 'Bank')}",
                              "account_code": "1001", "account_name": CHART_OF_ACCOUNTS["1001"]["name"],
                              "debit": amount, "credit": 0, "tags": {}})
            gl_entries.append({"entry_id": entry_id, "date": pmt_date, "reference": ref,
                              "description": f"Payment received - {pmt.get('payment_method', 'Bank')}",
                              "account_code": "1100", "account_name": CHART_OF_ACCOUNTS["1100"]["name"],
                              "debit": 0, "credit": amount, "tags": {}})
            entry_id += 1
        except Exception:
            pass
    
    # 3. TRAINER FEES - DR Trainer Fees Expense, CR Trainer Payable
    trainer_fees = await db.trainer_fees.find({}, {"_id": 0}).to_list(10000)
    for tf in trainer_fees:
        try:
            session_id = tf.get("session_id")
            if session_id not in session_ids:
                continue
            session = session_map.get(session_id, {})
            amount = float(tf.get("fee_amount", 0))
            if amount <= 0:
                continue
            
            trans_date = session.get("start_date", tf.get("created_at", "")[:10])
            programme = programme_map.get(session.get("program_id"), "Unknown")
            inv_num = session_invoice_map.get(session_id)
            tf_ref = f"TF-{inv_num}" if inv_num else f"TF-{session_id[:8]}"
            
            gl_entries.append({"entry_id": entry_id, "date": trans_date, "reference": tf_ref,
                              "description": f"Trainer fee accrual - {tf.get('trainer_name', 'Trainer')}",
                              "account_code": "5001", "account_name": CHART_OF_ACCOUNTS["5001"]["name"],
                              "debit": amount, "credit": 0, "tags": {"session_id": session_id, "programme": programme}})
            gl_entries.append({"entry_id": entry_id, "date": trans_date, "reference": tf_ref,
                              "description": f"Trainer fee accrual - {tf.get('trainer_name', 'Trainer')}",
                              "account_code": "2100", "account_name": CHART_OF_ACCOUNTS["2100"]["name"],
                              "debit": 0, "credit": amount, "tags": {"session_id": session_id, "programme": programme}})
            entry_id += 1
        except Exception:
            pass
    
    # 4. COORDINATOR FEES
    coordinator_fees = await db.coordinator_fees.find({}, {"_id": 0}).to_list(10000)
    for cf in coordinator_fees:
        try:
            session_id = cf.get("session_id")
            if session_id not in session_ids:
                continue
            session = session_map.get(session_id, {})
            amount = float(cf.get("total_fee", 0))
            if amount <= 0:
                continue
            
            trans_date = session.get("start_date", cf.get("created_at", "")[:10])
            programme = programme_map.get(session.get("program_id"), "Unknown")
            inv_num = session_invoice_map.get(session_id)
            cf_ref = f"CF-{inv_num}" if inv_num else f"CF-{session_id[:8]}"
            
            gl_entries.append({"entry_id": entry_id, "date": trans_date, "reference": cf_ref,
                              "description": f"Coordinator fee accrual - {cf.get('coordinator_name', 'Coordinator')}",
                              "account_code": "5002", "account_name": CHART_OF_ACCOUNTS["5002"]["name"],
                              "debit": amount, "credit": 0, "tags": {"session_id": session_id, "programme": programme}})
            gl_entries.append({"entry_id": entry_id, "date": trans_date, "reference": cf_ref,
                              "description": f"Coordinator fee accrual - {cf.get('coordinator_name', 'Coordinator')}",
                              "account_code": "2101", "account_name": CHART_OF_ACCOUNTS["2101"]["name"],
                              "debit": 0, "credit": amount, "tags": {"session_id": session_id, "programme": programme}})
            entry_id += 1
        except Exception:
            pass
    
    # 5. MARKETING COMMISSIONS
    marketing_comms = await db.marketing_commissions.find({"status": {"$in": ["approved", "paid"]}}, {"_id": 0}).to_list(10000)
    for mc in marketing_comms:
        try:
            session_id = mc.get("session_id")
            if session_id not in session_ids:
                continue
            session = session_map.get(session_id, {})
            amount = float(mc.get("calculated_amount", 0))
            if amount <= 0:
                continue
            
            trans_date = session.get("start_date", mc.get("created_at", "")[:10])
            programme = programme_map.get(session.get("program_id"), "Unknown")
            inv_num = session_invoice_map.get(session_id)
            mc_ref = f"MC-{inv_num}" if inv_num else f"MC-{session_id[:8]}"
            
            gl_entries.append({"entry_id": entry_id, "date": trans_date, "reference": mc_ref,
                              "description": f"Marketing commission - {mc.get('marketer_name', 'Marketer')}",
                              "account_code": "5003", "account_name": CHART_OF_ACCOUNTS["5003"]["name"],
                              "debit": amount, "credit": 0, "tags": {"session_id": session_id, "programme": programme}})
            gl_entries.append({"entry_id": entry_id, "date": trans_date, "reference": mc_ref,
                              "description": f"Marketing commission - {mc.get('marketer_name', 'Marketer')}",
                              "account_code": "2102", "account_name": CHART_OF_ACCOUNTS["2102"]["name"],
                              "debit": 0, "credit": amount, "tags": {"session_id": session_id, "programme": programme}})
            entry_id += 1
        except Exception:
            pass
    
    # Sort by date
    gl_entries.sort(key=lambda x: x["date"])
    
    # Round all amounts to 2 decimal places
    for entry in gl_entries:
        entry["debit"] = round(entry["debit"], 2)
        entry["credit"] = round(entry["credit"], 2)
    
    # Summary by account
    account_summary = {}
    for entry in gl_entries:
        code = entry["account_code"]
        if code not in account_summary:
            account_summary[code] = {"account_code": code, "account_name": entry["account_name"], "total_debit": 0, "total_credit": 0}
        account_summary[code]["total_debit"] += entry["debit"]
        account_summary[code]["total_credit"] += entry["credit"]
    
    # Round summary totals
    for code in account_summary:
        account_summary[code]["total_debit"] = round(account_summary[code]["total_debit"], 2)
        account_summary[code]["total_credit"] = round(account_summary[code]["total_credit"], 2)
    
    return {
        "year": year, "month": month,
        "entries": gl_entries,
        "account_summary": sorted(account_summary.values(), key=lambda x: x["account_code"]),
        "totals": {"total_debit": round(sum(e["debit"] for e in gl_entries), 2), "total_credit": round(sum(e["credit"] for e in gl_entries), 2)}
    }


# ==================== PHASE B: JOURNAL-BASED P&L ====================

@router.get("/pnl-journal")
async def get_journal_based_pnl(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    posted_only: Optional[bool] = True,
    current_user: User = Depends(get_current_user)
):
    """
    AUDITOR-GRADE Profit & Loss Statement derived entirely from posted journal entries.
    Groups by account code using the DB Chart of Accounts with pnl_section classification.
    
    Supports: date range, year, month filters.
    Returns: structured P&L with drill-down data per account.
    """
    if current_user.role not in ["admin", "super_admin", "finance", "coordinator"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # 1. Determine date range
    from datetime import datetime
    if date_from and date_to:
        start_date = date_from
        end_date = date_to
        period_label = f"{date_from} to {date_to}"
    elif year and month:
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year}-12-31"
        else:
            end_date = f"{year}-{month+1:02d}-01"
        period_label = f"{datetime(year, month, 1).strftime('%B %Y')}"
    elif year:
        start_date = f"{year}-01-01"
        end_date = f"{year}-12-31"
        period_label = f"Year {year}"
    else:
        year = get_malaysia_time().year
        start_date = f"{year}-01-01"
        end_date = f"{year}-12-31"
        period_label = f"Year {year}"
    
    # 2. Load COA from DB (single source of truth)
    coa_list = await db.chart_of_accounts.find({"is_active": True}, {"_id": 0}).to_list(500)
    coa_map = {a["account_code"]: a for a in coa_list}
    
    # 3. Query posted journal entries in date range
    je_query = {"date": {"$gte": start_date, "$lte": end_date}}
    if posted_only:
        je_query["status"] = "posted"
    
    journal_entries = await db.journal_entries.find(je_query, {"_id": 0}).to_list(50000)
    
    # 4. Aggregate balances by account code from journal lines
    account_balances = {}  # code -> {debit_total, credit_total, entries: [...]}
    unclassified_entries = []
    
    for je in journal_entries:
        for line in je.get("lines", []):
            code = line.get("account_code", "")
            debit = round(float(line.get("debit", 0)), 2)
            credit = round(float(line.get("credit", 0)), 2)
            
            if code not in account_balances:
                account_balances[code] = {"debit_total": 0, "credit_total": 0, "entries": [], "entry_count": 0}
            
            account_balances[code]["debit_total"] = round(account_balances[code]["debit_total"] + debit, 2)
            account_balances[code]["credit_total"] = round(account_balances[code]["credit_total"] + credit, 2)
            account_balances[code]["entry_count"] += 1
            account_balances[code]["entries"].append({
                "journal_no": je.get("journal_no"),
                "date": je.get("date"),
                "description": je.get("description", ""),
                "line_memo": line.get("memo", ""),
                "debit": debit,
                "credit": credit,
                "source_module": je.get("source_module"),
                "source_reference": je.get("source_reference"),
            })
    
    # 5. Build P&L sections from COA classification
    sections = {
        "revenue": {"label": "Revenue", "accounts": [], "total": 0},
        "cost_of_sales": {"label": "Cost of Sales / Direct Costs", "accounts": [], "total": 0},
        "other_income": {"label": "Other Income", "accounts": [], "total": 0},
        "operating_expense": {"label": "Operating Expenses", "accounts": [], "total": 0},
        "other_expense": {"label": "Other Expenses", "accounts": [], "total": 0},
    }
    
    for code, bal in account_balances.items():
        acct = coa_map.get(code)
        if not acct:
            # Account not in COA — classify by code range
            if code.startswith("4"):
                pnl_section = "revenue" if code < "4100" else "other_income"
            elif code.startswith("5"):
                pnl_section = "cost_of_sales"
            elif code.startswith("6"):
                pnl_section = "operating_expense"
            else:
                continue  # Balance sheet account — skip for P&L
            acct = {"account_code": code, "account_name": f"Unknown ({code})", "pnl_section": pnl_section, "normal_balance": "credit" if code.startswith("4") else "debit"}
        
        pnl_section = acct.get("pnl_section")
        statement_type = acct.get("statement_type")
        
        if not pnl_section:
            if statement_type == "balance_sheet":
                continue  # Skip balance sheet accounts
            # Fallback classification
            if acct.get("account_type") == "Income":
                pnl_section = "revenue"
            elif acct.get("account_category") == "Direct Cost":
                pnl_section = "cost_of_sales"
            elif acct.get("account_category") == "Operating Expense":
                pnl_section = "operating_expense"
            else:
                continue
        
        if pnl_section not in sections:
            continue
        
        # Calculate balance based on normal_balance
        normal = acct.get("normal_balance", "debit")
        if normal == "credit":
            amount = round(bal["credit_total"] - bal["debit_total"], 2)
        else:
            amount = round(bal["debit_total"] - bal["credit_total"], 2)
        
        if amount == 0 and bal["entry_count"] == 0:
            continue
        
        sections[pnl_section]["accounts"].append({
            "account_code": code,
            "account_name": acct.get("account_name", code),
            "amount": amount,
            "debit_total": bal["debit_total"],
            "credit_total": bal["credit_total"],
            "entry_count": bal["entry_count"],
            "entries": bal["entries"][:50],  # Limit drill-down entries
        })
        sections[pnl_section]["total"] = round(sections[pnl_section]["total"] + amount, 2)
    
    # Sort accounts within sections by code
    for sec in sections.values():
        sec["accounts"].sort(key=lambda a: a["account_code"])
    
    # 6. Calculate P&L totals
    total_revenue = round(sections["revenue"]["total"] + sections["other_income"]["total"], 2)
    total_cos = round(sections["cost_of_sales"]["total"], 2)
    gross_profit = round(sections["revenue"]["total"] - total_cos, 2)
    total_opex = round(sections["operating_expense"]["total"] + sections["other_expense"]["total"], 2)
    net_profit = round(total_revenue - total_cos - total_opex, 2)
    
    # 7. Data quality warnings
    warnings = []
    for code in account_balances:
        if code not in coa_map:
            warnings.append(f"Account {code} used in journals but not in Chart of Accounts")
    
    # Check for unbalanced entries
    for je in journal_entries:
        total_d = sum(float(l.get("debit", 0)) for l in je.get("lines", []))
        total_c = sum(float(l.get("credit", 0)) for l in je.get("lines", []))
        if abs(total_d - total_c) > 0.01:
            warnings.append(f"Unbalanced journal {je.get('journal_no')}: DR {total_d} != CR {total_c}")
    
    return {
        "period": period_label,
        "date_from": start_date,
        "date_to": end_date,
        "posted_only": posted_only,
        "journal_count": len(journal_entries),
        "sections": sections,
        "summary": {
            "total_revenue": sections["revenue"]["total"],
            "other_income": sections["other_income"]["total"],
            "total_income": total_revenue,
            "cost_of_sales": total_cos,
            "gross_profit": gross_profit,
            "gross_margin_pct": round((gross_profit / sections["revenue"]["total"] * 100), 2) if sections["revenue"]["total"] > 0 else 0,
            "operating_expenses": total_opex,
            "net_profit": net_profit,
            "net_margin_pct": round((net_profit / total_revenue * 100), 2) if total_revenue > 0 else 0,
        },
        "warnings": warnings[:20],
    }


@router.get("/pnl-journal/export")
async def export_pnl_journal_excel(
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    year: Optional[int] = None, month: Optional[int] = None,
    posted_only: Optional[bool] = True,
    current_user: User = Depends(get_current_user)
):
    """Export journal-based P&L to Excel"""
    if current_user.role not in ["admin", "super_admin", "finance", "coordinator"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime as dt
    
    # Reuse the P&L logic
    pnl_data = await get_journal_based_pnl(date_from=date_from, date_to=date_to, year=year, month=month, posted_only=posted_only, current_user=current_user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Statement"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    section_font = Font(bold=True, size=11)
    total_font = Font(bold=True, size=12)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = f"PROFIT & LOSS STATEMENT — {pnl_data['period']}"
    ws['A1'].font = Font(bold=True, size=14, color="1a365d")
    ws.merge_cells('A2:C2')
    ws['A2'] = f"{'Posted Only' if posted_only else 'Including Drafts'} | {pnl_data['journal_count']} Journal Entries"
    ws['A2'].font = Font(size=10, color="666666")
    
    row = 4
    for col, header in enumerate(['Code', 'Account Name', 'Amount (RM)'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    section_order = ['revenue', 'other_income', 'cost_of_sales', 'operating_expense', 'other_expense']
    section_colors = {'revenue': '16a34a', 'other_income': '0d9488', 'cost_of_sales': 'ea580c', 'operating_expense': 'dc2626', 'other_expense': '9333ea'}
    calc_rows = {
        'other_income': ('TOTAL INCOME', pnl_data['summary']['total_income'], '16a34a'),
        'cost_of_sales': ('GROSS PROFIT', pnl_data['summary']['gross_profit'], '2563eb'),
    }
    
    row += 1
    for sec_key in section_order:
        sec = pnl_data['sections'].get(sec_key, {})
        if not sec.get('accounts'):
            continue
        
        # Section header
        ws.cell(row=row, column=1, value=sec['label'].upper()).font = Font(bold=True, color=section_colors.get(sec_key, '333333'))
        row += 1
        
        for a in sec.get('accounts', []):
            ws.cell(row=row, column=1, value=a['account_code']).border = thin_border
            ws.cell(row=row, column=2, value=a['account_name']).border = thin_border
            ws.cell(row=row, column=3, value=a['amount']).border = thin_border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
        
        ws.cell(row=row, column=2, value=f"Subtotal {sec['label']}").font = Font(bold=True)
        ws.cell(row=row, column=3, value=sec['total']).font = Font(bold=True)
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        row += 1
        
        if sec_key in calc_rows:
            label, val, color = calc_rows[sec_key]
            ws.cell(row=row, column=2, value=label).font = Font(bold=True, color=color, size=12)
            ws.cell(row=row, column=3, value=val).font = Font(bold=True, color=color, size=12)
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
        
        row += 1
    
    # Net Profit
    ws.cell(row=row, column=2, value='NET PROFIT BEFORE TAX').font = Font(bold=True, size=13, color="1a365d")
    ws.cell(row=row, column=3, value=pnl_data['summary']['net_profit']).font = Font(bold=True, size=13, color="1a365d")
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="PnL_{pnl_data["period"].replace(" ", "_")}.xlsx"'})


@router.get("/pnl-journal/drilldown/{account_code}")
async def pnl_drilldown(
    account_code: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Drill-down: Show all journal entries for a specific account code in a date range."""
    if current_user.role not in ["admin", "super_admin", "finance", "coordinator"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if date_from and date_to:
        start_date, end_date = date_from, date_to
    elif year:
        start_date, end_date = f"{year}-01-01", f"{year}-12-31"
    else:
        y = get_malaysia_time().year
        start_date, end_date = f"{y}-01-01", f"{y}-12-31"
    
    journal_entries = await db.journal_entries.find(
        {"date": {"$gte": start_date, "$lte": end_date}, "status": "posted", "lines.account_code": account_code},
        {"_id": 0}
    ).sort("date", 1).to_list(5000)
    
    result = []
    for je in journal_entries:
        for line in je.get("lines", []):
            if line.get("account_code") == account_code:
                result.append({
                    "journal_no": je.get("journal_no"),
                    "date": je.get("date"),
                    "description": je.get("description"),
                    "line_memo": line.get("memo", ""),
                    "debit": round(float(line.get("debit", 0)), 2),
                    "credit": round(float(line.get("credit", 0)), 2),
                    "source_module": je.get("source_module"),
                    "source_reference": je.get("source_reference"),
                    "source_id": je.get("source_id"),
                })
    
    acct = await db.chart_of_accounts.find_one({"account_code": account_code}, {"_id": 0})
    
    return {
        "account_code": account_code,
        "account_name": acct.get("account_name") if acct else account_code,
        "date_from": start_date,
        "date_to": end_date,
        "entries": result,
        "total_debit": round(sum(e["debit"] for e in result), 2),
        "total_credit": round(sum(e["credit"] for e in result), 2),
    }


# ==================== TRIAL BALANCE ====================

@router.get("/trial-balance")
async def get_trial_balance(
    year: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Trial Balance — COA with actual debit/credit balances from journal entries"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")

    from datetime import datetime

    # Build date filter
    date_filter = {"status": "posted"}
    period_label = "All Time"
    if date_from and date_to:
        date_filter["date"] = {"$gte": date_from, "$lte": date_to}
        period_label = f"{date_from} to {date_to}"
    elif year:
        date_filter["date"] = {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}
        period_label = f"Year {year}"

    # Get all active accounts from COA
    accounts = await db.chart_of_accounts.find(
        {"is_active": True}, {"_id": 0}
    ).sort("account_code", 1).to_list(500)

    # Aggregate journal entry lines by account code
    pipeline = [
        {"$match": date_filter},
        {"$unwind": "$lines"},
        {"$group": {
            "_id": "$lines.account_code",
            "total_debit": {"$sum": {"$toDouble": {"$ifNull": ["$lines.debit", 0]}}},
            "total_credit": {"$sum": {"$toDouble": {"$ifNull": ["$lines.credit", 0]}}},
            "entry_count": {"$sum": 1}
        }}
    ]
    balances_raw = await db.journal_entries.aggregate(pipeline).to_list(500)
    balance_map = {b["_id"]: b for b in balances_raw}

    # Build trial balance rows
    rows = []
    grand_debit = 0
    grand_credit = 0

    for acc in accounts:
        code = acc["account_code"]
        bal = balance_map.get(code, {})
        total_debit = round(bal.get("total_debit", 0), 2)
        total_credit = round(bal.get("total_credit", 0), 2)
        net = round(total_debit - total_credit, 2)

        # Normal balance determines which column shows the net
        normal = acc.get("normal_balance", "debit")
        if normal == "debit":
            debit_balance = round(net, 2) if net > 0 else 0
            credit_balance = round(abs(net), 2) if net < 0 else 0
        else:
            credit_balance = round(abs(net), 2) if net < 0 or (net <= 0) else 0
            debit_balance = round(net, 2) if net > 0 else 0
            # For credit-normal accounts, net negative means credit balance
            credit_balance = round(abs(net), 2) if net <= 0 else 0
            debit_balance = round(net, 2) if net > 0 else 0

        grand_debit += debit_balance
        grand_credit += credit_balance

        rows.append({
            "account_code": code,
            "account_name": acc["account_name"],
            "account_type": acc["account_type"],
            "account_category": acc.get("account_category", ""),
            "normal_balance": normal,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "debit_balance": debit_balance,
            "credit_balance": credit_balance,
            "entry_count": bal.get("entry_count", 0),
        })

    return {
        "period": period_label,
        "accounts": rows,
        "totals": {
            "total_debit": round(grand_debit, 2),
            "total_credit": round(grand_credit, 2),
            "is_balanced": abs(grand_debit - grand_credit) < 0.01,
            "difference": round(grand_debit - grand_credit, 2),
        },
        "account_count": len(rows),
        "accounts_with_activity": sum(1 for r in rows if r["entry_count"] > 0),
    }


# ==================== BALANCE SHEET ====================

@router.get("/balance-sheet")
async def get_balance_sheet(
    as_at: Optional[str] = None,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Balance Sheet — Assets = Liabilities + Equity. Uses journal entries up to the given date."""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")

    from datetime import datetime

    # Determine cut-off date
    if as_at:
        cutoff_date = as_at
    elif year:
        cutoff_date = f"{year}-12-31"
    else:
        cutoff_date = datetime.now().strftime("%Y-%m-%d")

    period_label = f"As at {cutoff_date}"

    # Get COA for balance sheet accounts only
    bs_accounts = await db.chart_of_accounts.find(
        {"is_active": True, "statement_type": "balance_sheet"},
        {"_id": 0}
    ).sort("account_code", 1).to_list(200)

    # Also get P&L accounts (to calculate retained earnings)
    pnl_accounts = await db.chart_of_accounts.find(
        {"is_active": True, "statement_type": "profit_and_loss"},
        {"_id": 0}
    ).sort("account_code", 1).to_list(200)

    all_codes = [a["account_code"] for a in bs_accounts + pnl_accounts]

    # Aggregate all journal entries up to cutoff date
    pipeline = [
        {"$match": {"status": "posted", "date": {"$lte": cutoff_date}}},
        {"$unwind": "$lines"},
        {"$match": {"lines.account_code": {"$in": all_codes}}},
        {"$group": {
            "_id": "$lines.account_code",
            "total_debit": {"$sum": {"$toDouble": {"$ifNull": ["$lines.debit", 0]}}},
            "total_credit": {"$sum": {"$toDouble": {"$ifNull": ["$lines.credit", 0]}}}
        }}
    ]
    balances_raw = await db.journal_entries.aggregate(pipeline).to_list(500)
    balance_map = {b["_id"]: b for b in balances_raw}

    def calc_balance(code, normal):
        bal = balance_map.get(code, {})
        dr = bal.get("total_debit", 0)
        cr = bal.get("total_credit", 0)
        if normal == "debit":
            return round(dr - cr, 2)
        else:
            return round(cr - dr, 2)

    # Build sections
    assets = {"current": [], "non_current": [], "total": 0}
    liabilities = {"current": [], "non_current": [], "total": 0}
    equity = {"accounts": [], "total": 0}

    for acc in bs_accounts:
        code = acc["account_code"]
        normal = acc.get("normal_balance", "debit")
        balance = calc_balance(code, normal)
        row = {
            "account_code": code,
            "account_name": acc["account_name"],
            "account_category": acc.get("account_category", ""),
            "balance": balance,
        }

        atype = acc["account_type"]
        if atype == "Asset":
            # Current assets: codes 1000-1499, Non-current: 1500+
            if int(code) < 1500:
                assets["current"].append(row)
            else:
                assets["non_current"].append(row)
            assets["total"] = round(assets["total"] + balance, 2)
        elif atype == "Liability":
            if int(code) < 2500:
                liabilities["current"].append(row)
            else:
                liabilities["non_current"].append(row)
            liabilities["total"] = round(liabilities["total"] + balance, 2)
        elif atype == "Equity":
            equity["accounts"].append(row)
            equity["total"] = round(equity["total"] + balance, 2)

    # Calculate retained earnings from P&L accounts
    retained_earnings = 0
    for acc in pnl_accounts:
        code = acc["account_code"]
        normal = acc.get("normal_balance", "credit")
        balance = calc_balance(code, normal)
        # Revenue adds to retained earnings, expenses subtract
        if acc["account_type"] in ["Income", "Revenue"]:
            retained_earnings += balance
        else:  # Expense
            retained_earnings -= balance
    retained_earnings = round(retained_earnings, 2)

    equity["accounts"].append({
        "account_code": "RE",
        "account_name": "Retained Earnings (from P&L)",
        "account_category": "Equity",
        "balance": retained_earnings,
    })
    equity["total"] = round(equity["total"] + retained_earnings, 2)

    total_liabilities_equity = round(liabilities["total"] + equity["total"], 2)
    is_balanced = abs(assets["total"] - total_liabilities_equity) < 0.01

    return {
        "period": period_label,
        "as_at": cutoff_date,
        "assets": assets,
        "liabilities": liabilities,
        "equity": equity,
        "summary": {
            "total_assets": assets["total"],
            "total_liabilities": liabilities["total"],
            "total_equity": equity["total"],
            "total_liabilities_equity": total_liabilities_equity,
            "is_balanced": is_balanced,
            "difference": round(assets["total"] - total_liabilities_equity, 2),
        },
    }



@router.get("/ar-aging")
async def get_ar_aging_report(current_user: User = Depends(get_current_user)):
    """Accounts Receivable Aging Report — shows unpaid invoices by age bucket."""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")

    from datetime import datetime, timezone, timedelta

    now = datetime.now(timezone.utc)
    invoices = await db.invoices.find(
        {"status": {"$in": ["issued", "approved", "pending", "partial"]}},
        {"_id": 0}
    ).to_list(None)

    buckets = {"current": [], "days_31_60": [], "days_61_90": [], "days_91_120": [], "days_120_plus": []}
    by_company = {}

    for inv in invoices:
        date_str = inv.get("invoice_date") or inv.get("created_at", "")
        try:
            if "T" in str(date_str):
                inv_date = datetime.fromisoformat(str(date_str).replace("Z", "+00:00"))
            else:
                inv_date = datetime.strptime(str(date_str)[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except Exception:
            inv_date = now

        days = (now - inv_date).days
        amount = inv.get("total_amount", 0) or 0
        company = inv.get("company_name", "Unknown")

        item = {
            "id": inv.get("id"),
            "invoice_number": inv.get("invoice_number", ""),
            "company": company,
            "amount": amount,
            "date": str(date_str)[:10] if date_str else "",
            "days_outstanding": days,
            "status": inv.get("status", ""),
        }

        if days <= 30:
            buckets["current"].append(item)
        elif days <= 60:
            buckets["days_31_60"].append(item)
        elif days <= 90:
            buckets["days_61_90"].append(item)
        elif days <= 120:
            buckets["days_91_120"].append(item)
        else:
            buckets["days_120_plus"].append(item)

        if company not in by_company:
            by_company[company] = {"current": 0, "days_31_60": 0, "days_61_90": 0, "days_91_120": 0, "days_120_plus": 0, "total": 0}
        if days <= 30:
            by_company[company]["current"] += amount
        elif days <= 60:
            by_company[company]["days_31_60"] += amount
        elif days <= 90:
            by_company[company]["days_61_90"] += amount
        elif days <= 120:
            by_company[company]["days_91_120"] += amount
        else:
            by_company[company]["days_120_plus"] += amount
        by_company[company]["total"] += amount

    summary = {
        "current": sum(i["amount"] for i in buckets["current"]),
        "days_31_60": sum(i["amount"] for i in buckets["days_31_60"]),
        "days_61_90": sum(i["amount"] for i in buckets["days_61_90"]),
        "days_91_120": sum(i["amount"] for i in buckets["days_91_120"]),
        "days_120_plus": sum(i["amount"] for i in buckets["days_120_plus"]),
    }
    summary["total"] = sum(summary.values())

    return {
        "as_of": now.isoformat(),
        "summary": summary,
        "buckets": buckets,
        "by_company": by_company,
        "total_invoices": len(invoices),
    }
