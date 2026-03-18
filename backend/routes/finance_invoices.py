"""
Finance Invoices routes
Stage F2: ~15 core invoice endpoints
"""
from fastapi import APIRouter, HTTPException, Depends
from typing import Optional, List
from datetime import datetime
from pydantic import BaseModel, Field, ConfigDict
import uuid
import pytz

from core import db, get_current_user, get_malaysia_time
from models import User
from utils.email_notifications import notify_invoice_issued

# Import accounting auto-posting functions (Phase 2)
from routes.accounting import post_invoice_issued

router = APIRouter(prefix="/finance", tags=["finance-invoices"])

MALAYSIA_TZ = pytz.timezone("Asia/Kuala_Lumpur")


# ============ MODELS ============
class InvoiceUpdate(BaseModel):
    bill_to_name: Optional[str] = None
    bill_to_address: Optional[str] = None
    bill_to_reg_no: Optional[str] = None
    your_reference: Optional[str] = None
    programme_name: Optional[str] = None
    training_dates: Optional[str] = None
    venue: Optional[str] = None
    pax: Optional[int] = None
    num_days: Optional[int] = None
    pricing_type: Optional[str] = None
    line_items: Optional[List[dict]] = None
    subtotal: Optional[float] = None
    mobilisation_fee: Optional[float] = None
    rounding: Optional[float] = None
    tax_rate: Optional[float] = None
    tax_amount: Optional[float] = None
    discount: Optional[float] = None
    total_amount: Optional[float] = None
    status: Optional[str] = None


class InvoiceNumberEditRequest(BaseModel):
    year: int
    month: int
    sequence: int
    reason: str


class VoidInvoiceRequest(BaseModel):
    reason: str


class EditPaidInvoiceRequest(BaseModel):
    bill_to_name: Optional[str] = None
    bill_to_address: Optional[str] = None
    total_amount: Optional[float] = None
    reason: str


class BackdateInvoiceRequest(BaseModel):
    new_date: str
    reason: str


class ResetSequenceRequest(BaseModel):
    year: int
    month: int
    new_sequence: int
    reason: str


class OverrideValidationRequest(BaseModel):
    total_amount: float
    reason: str
    skip_validation: bool = True


# ============ HELPER FUNCTIONS ============
async def generate_invoice_number():
    """Generate unique invoice number: INV/MDDRC/YYYY/MM/0001"""
    now = get_malaysia_time()
    year = now.year
    month = now.month
    prefix = f"INV/MDDRC/{year}/{month:02d}/"
    
    sequence_override = await db.invoice_sequence_settings.find_one(
        {"year": year, "month": month}, {"_id": 0}
    )
    
    last_invoice = await db.invoices.find_one(
        {"invoice_number": {"$regex": f"^INV/MDDRC/{year}/{month:02d}/"}},
        sort=[("invoice_number", -1)]
    )
    
    if sequence_override and sequence_override.get("next_sequence"):
        new_num = sequence_override["next_sequence"]
        await db.invoice_sequence_settings.delete_one({"year": year, "month": month})
    elif last_invoice:
        last_num = int(last_invoice["invoice_number"].split("/")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"{prefix}{new_num:04d}"


async def log_finance_action(entity_type: str, entity_id: str, action: str, 
                             changed_by: str, before_value: dict = None, 
                             after_value: dict = None, reason: str = None):
    log_entry = {
        "id": str(uuid.uuid4()),
        "entity_type": entity_type,
        "entity_id": entity_id,
        "action": action,
        "before_value": before_value,
        "after_value": after_value,
        "changed_by": changed_by,
        "reason": reason,
        "timestamp": get_malaysia_time().isoformat()
    }
    await db.finance_audit_log.insert_one(log_entry)


async def create_audit_trail_entry(
    action: str, record_reference: str, entity_type: str, entity_id: str,
    changed_by: User, reason: str, field_changed: str = None,
    from_value: str = None, to_value: str = None
):
    """Create a detailed audit trail entry"""
    entry = {
        "id": str(uuid.uuid4()),
        "action": action,
        "record_reference": record_reference,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "field_changed": field_changed,
        "from_value": from_value,
        "to_value": to_value,
        "changed_by_name": changed_by.full_name,
        "changed_by_email": changed_by.email,
        "reason": reason,
        "timestamp": get_malaysia_time().isoformat()
    }
    await db.audit_trail.insert_one(entry)
    return entry


# ============ CORE INVOICE ENDPOINTS ============
@router.get("/invoices")
async def get_invoices(
    status: Optional[str] = None,
    company_id: Optional[str] = None,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all invoices with optional year filter"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if status:
        query["status"] = status
    if company_id:
        query["company_id"] = company_id
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    if year:
        def get_invoice_year(inv):
            date_val = inv.get("invoice_date") or inv.get("created_at")
            if not date_val:
                return None
            if isinstance(date_val, str):
                try:
                    return datetime.fromisoformat(date_val.replace('Z', '+00:00')).year
                except:
                    try:
                        return int(date_val[:4])
                    except:
                        return None
            elif hasattr(date_val, 'year'):
                return date_val.year
            return None
        
        invoices = [inv for inv in invoices if get_invoice_year(inv) == year]
    
    return invoices


@router.get("/invoices/export")
async def export_invoices(
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export invoices data as actual Excel (.xlsx) file"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if status:
        query["status"] = status
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", 1).to_list(10000)
    payments = await db.payments.find({}, {"_id": 0}).to_list(10000)
    payment_by_invoice = {p.get("invoice_id"): p for p in payments}
    
    credit_notes = await db.credit_notes.find({}, {"_id": 0}).to_list(10000)
    cn_by_invoice = {}
    for cn in credit_notes:
        inv_id = cn.get("invoice_id")
        if inv_id:
            if inv_id not in cn_by_invoice:
                cn_by_invoice[inv_id] = []
            cn_by_invoice[inv_id].append(cn)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Bil", "Date", "Invoice Number", "Bill To", "Programme", "Company Name", 
               "Venue", "No of Participants", "Invoice Value (RM)", "Invoice Status", 
               "Payment Status", "Credit Note No & Value"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    bil = 1
    for inv in invoices:
        payment = payment_by_invoice.get(inv.get("id"))
        payment_status = "Paid" if payment else "Unpaid"
        
        inv_credit_notes = cn_by_invoice.get(inv.get("id"), [])
        cn_info = ""
        if inv_credit_notes:
            cn_parts = [f"{cn.get('cn_number', 'CN')}: RM{cn.get('amount', 0)}" for cn in inv_credit_notes]
            cn_info = "; ".join(cn_parts)
        
        row = bil + 1
        ws.cell(row=row, column=1, value=bil).border = thin_border
        ws.cell(row=row, column=2, value=str(inv.get("created_at", ""))[:10] if inv.get("created_at") else "").border = thin_border
        ws.cell(row=row, column=3, value=inv.get("invoice_number", "")).border = thin_border
        ws.cell(row=row, column=4, value=inv.get("bill_to_name") or inv.get("company_name", "")).border = thin_border
        ws.cell(row=row, column=5, value=inv.get("programme_name", "")).border = thin_border
        ws.cell(row=row, column=6, value=inv.get("company_name", "")).border = thin_border
        ws.cell(row=row, column=7, value=inv.get("venue", "")).border = thin_border
        ws.cell(row=row, column=8, value=inv.get("pax", 0)).border = thin_border
        ws.cell(row=row, column=9, value=inv.get("total_amount", 0)).border = thin_border
        ws.cell(row=row, column=10, value=inv.get("status", "").replace("_", " ").title()).border = thin_border
        ws.cell(row=row, column=11, value=payment_status).border = thin_border
        ws.cell(row=row, column=12, value=cn_info).border = thin_border
        bil += 1
    
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="invoices_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'}
    )


@router.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Get single invoice"""
    if current_user.role not in ["admin", "super_admin", "finance", "coordinator"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return invoice


@router.put("/invoices/{invoice_id}")
async def update_invoice(
    invoice_id: str,
    update_data: InvoiceUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update invoice (Finance only)"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Finance can update invoices")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if invoice.get("status") in ["issued", "paid"]:
        raise HTTPException(status_code=400, detail="Cannot modify issued/paid invoice")
    
    before_value = dict(invoice)
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = get_malaysia_time().isoformat()
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_dict})
    
    if "status" in update_dict:
        await db.sessions.update_one(
            {"invoice_id": invoice_id},
            {"$set": {"invoice_status": update_dict["status"]}}
        )
    
    await log_finance_action("invoice", invoice_id, "updated", current_user.id, before_value, update_dict)
    
    return await db.invoices.find_one({"id": invoice_id}, {"_id": 0})


@router.post("/invoices/{invoice_id}/approve")
async def approve_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Approve invoice"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Finance can approve invoices")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if invoice.get("status") not in ["auto_draft", "finance_review", "draft"]:
        raise HTTPException(status_code=400, detail="Invoice cannot be approved from current status")
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user.id,
            "approved_at": get_malaysia_time().isoformat(),
            "updated_at": get_malaysia_time().isoformat()
        }}
    )
    
    await db.sessions.update_one({"invoice_id": invoice_id}, {"$set": {"invoice_status": "approved"}})
    await log_finance_action("invoice", invoice_id, "status_changed", current_user.id, 
                            {"status": invoice.get("status")}, {"status": "approved"})
    
    return {"message": "Invoice approved successfully"}


@router.post("/invoices/{invoice_id}/issue")
async def issue_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    """Issue invoice"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Finance can issue invoices")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if invoice.get("status") != "approved":
        raise HTTPException(status_code=400, detail="Only approved invoices can be issued")
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": "issued",
            "issued_by": current_user.id,
            "issued_at": get_malaysia_time().isoformat(),
            "updated_at": get_malaysia_time().isoformat()
        }}
    )
    
    await db.sessions.update_one({"invoice_id": invoice_id}, {"$set": {"invoice_status": "issued"}})
    
    session = await db.sessions.find_one({"invoice_id": invoice_id}, {"_id": 0})
    if session and session.get("marketing_user_id"):
        commission_amount = 0.0
        if session.get("commission_type") == "percentage":
            commission_amount = invoice.get("total_amount", 0) * (session.get("commission_rate", 0) / 100)
        else:
            commission_amount = session.get("commission_fixed_amount", 0)
        
        await db.marketing_commissions.update_one(
            {"session_id": session["id"]},
            {"$set": {
                "calculated_amount": commission_amount,
                "invoice_id": invoice_id,
                "status": "approved",
                "updated_at": get_malaysia_time().isoformat()
            }},
            upsert=True
        )
    
    await log_finance_action("invoice", invoice_id, "status_changed", current_user.id,
                            {"status": invoice.get("status")}, {"status": "issued"})
    
    # ============ ACCOUNTING AUTO-POST (Phase 2) ============
    # Create journal entry for issued invoice
    try:
        updated_invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
        accounting_result = await post_invoice_issued(
            invoice=updated_invoice,
            session=session,
            user_id=current_user.id,
            user_name=current_user.full_name
        )
        if accounting_result.get("error"):
            print(f"Accounting auto-post warning: {accounting_result.get('error')}")
    except Exception as e:
        print(f"Accounting auto-post error: {str(e)}")
    # ============ END ACCOUNTING AUTO-POST ============
    
    # ============ EMAIL NOTIFICATION ============
    try:
        updated_invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
        await notify_invoice_issued(updated_invoice, session)
    except Exception as e:
        print(f"Invoice notification error: {str(e)}")
    # ============ END EMAIL NOTIFICATION ============
    
    return {"message": "Invoice issued successfully"}


@router.post("/invoices/{invoice_id}/cancel")
async def cancel_invoice(invoice_id: str, reason: str = "", current_user: User = Depends(get_current_user)):
    """Cancel invoice"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Finance can cancel invoices")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": "cancelled",
            "cancelled_by": current_user.id,
            "cancelled_at": get_malaysia_time().isoformat(),
            "cancellation_reason": reason,
            "updated_at": get_malaysia_time().isoformat()
        }}
    )
    
    await db.sessions.update_one({"invoice_id": invoice_id}, {"$set": {"invoice_status": "cancelled"}})
    await log_finance_action("invoice", invoice_id, "status_changed", current_user.id,
                            {"status": invoice.get("status")}, {"status": "cancelled", "reason": reason}, reason)
    
    return {"message": "Invoice cancelled successfully"}

@router.post("/invoices/{invoice_id}/revert-status")
async def revert_invoice_status(invoice_id: str, target_status: str = "auto_draft", reason: str = "", current_user: User = Depends(get_current_user)):
    """Revert invoice status (Admin only) - used to undo cancellations"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can revert invoice status")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    allowed_targets = ["auto_draft", "draft", "finance_review"]
    if target_status not in allowed_targets:
        raise HTTPException(status_code=400, detail=f"Target status must be one of: {allowed_targets}")
    
    old_status = invoice.get("status")
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": target_status,
            "updated_at": get_malaysia_time().isoformat(),
            "revert_reason": reason,
            "reverted_by": current_user.id,
            "reverted_at": get_malaysia_time().isoformat()
        },
        "$unset": {
            "cancelled_by": "",
            "cancelled_at": "",
            "cancellation_reason": ""
        }}
    )
    
    await db.sessions.update_one({"invoice_id": invoice_id}, {"$set": {"invoice_status": target_status}})
    await log_finance_action("invoice", invoice_id, "status_reverted", current_user.id,
                            {"status": old_status}, {"status": target_status, "reason": reason})
    
    return {"message": f"Invoice {invoice.get('invoice_number')} reverted from '{old_status}' to '{target_status}'"}


@router.post("/invoices/revert-batch")
async def revert_invoices_batch(data: dict, current_user: User = Depends(get_current_user)):
    """Revert multiple invoices by invoice_number (Admin only)"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can revert invoice status")
    
    invoice_numbers = data.get("invoice_numbers", [])
    target_status = data.get("target_status", "auto_draft")
    
    allowed_targets = ["auto_draft", "draft", "finance_review"]
    if target_status not in allowed_targets:
        raise HTTPException(status_code=400, detail=f"Target status must be one of: {allowed_targets}")
    
    results = []
    for inv_num in invoice_numbers:
        invoice = await db.invoices.find_one({"invoice_number": inv_num}, {"_id": 0})
        if not invoice:
            results.append({"invoice_number": inv_num, "status": "not_found"})
            continue
        
        old_status = invoice.get("status")
        await db.invoices.update_one(
            {"id": invoice["id"]},
            {"$set": {"status": target_status, "updated_at": get_malaysia_time().isoformat()},
             "$unset": {"cancelled_by": "", "cancelled_at": "", "cancellation_reason": ""}}
        )
        await db.sessions.update_one({"invoice_id": invoice["id"]}, {"$set": {"invoice_status": target_status}})
        await log_finance_action("invoice", invoice["id"], "status_reverted", current_user.id,
                                {"status": old_status}, {"status": target_status})
        results.append({"invoice_number": inv_num, "old_status": old_status, "new_status": target_status})
    
    return {"results": results}




@router.post("/invoices/{invoice_id}/create-replacement")
async def create_replacement_invoice(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Create a replacement invoice for a voided invoice"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can create replacement invoices")
    
    voided_invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not voided_invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if voided_invoice.get("status") != "voided":
        raise HTTPException(status_code=400, detail="Replacement invoice can only be created for voided invoices")
    
    existing_replacement = await db.invoices.find_one({
        "replaces_invoice_id": invoice_id,
        "status": {"$ne": "voided"}
    }, {"_id": 0})
    if existing_replacement:
        raise HTTPException(status_code=400, detail=f"A replacement invoice already exists: {existing_replacement.get('invoice_number')}")
    
    new_invoice_number = await generate_invoice_number()
    
    replacement_invoice = {
        "id": str(uuid.uuid4()),
        "invoice_number": new_invoice_number,
        "session_id": voided_invoice.get("session_id"),
        "company_id": voided_invoice.get("company_id"),
        "company_name": voided_invoice.get("company_name"),
        "programme_name": voided_invoice.get("programme_name"),
        "training_dates": voided_invoice.get("training_dates"),
        "venue": voided_invoice.get("venue"),
        "pax": voided_invoice.get("pax"),
        "line_items": voided_invoice.get("line_items", []),
        "subtotal": voided_invoice.get("subtotal", 0),
        "tax_rate": voided_invoice.get("tax_rate", 0),
        "tax_amount": voided_invoice.get("tax_amount", 0),
        "total_amount": voided_invoice.get("total_amount", 0),
        "discount": voided_invoice.get("discount", 0),
        "mobilisation_fee": voided_invoice.get("mobilisation_fee", 0),
        "rounding": voided_invoice.get("rounding", 0),
        "pricing_type": voided_invoice.get("pricing_type"),
        "bill_to_name": voided_invoice.get("bill_to_name"),
        "bill_to_address": voided_invoice.get("bill_to_address"),
        "bill_to_reg_no": voided_invoice.get("bill_to_reg_no"),
        "your_reference": voided_invoice.get("your_reference", ""),
        "status": "auto_draft",
        "replaces_invoice_id": invoice_id,
        "replaces_invoice_number": voided_invoice.get("invoice_number"),
        "created_at": get_malaysia_time().isoformat(),
        "updated_at": get_malaysia_time().isoformat(),
        "version": 1
    }
    
    await db.invoices.insert_one(replacement_invoice)
    
    if voided_invoice.get("session_id"):
        await db.sessions.update_one(
            {"id": voided_invoice.get("session_id")},
            {"$set": {
                "invoice_id": replacement_invoice["id"],
                "invoice_number": new_invoice_number,
                "invoice_status": "auto_draft"
            }}
        )
    
    await log_finance_action(
        entity_type="invoice",
        entity_id=replacement_invoice["id"],
        action="created",
        changed_by=current_user.id,
        after_value={"invoice_number": new_invoice_number, "replaces": voided_invoice.get("invoice_number")}
    )
    
    return {
        "message": "Replacement invoice created successfully",
        "new_invoice_id": replacement_invoice["id"],
        "new_invoice_number": new_invoice_number,
        "replaces_invoice_number": voided_invoice.get("invoice_number")
    }


@router.post("/invoices/{invoice_id}/reverse-void")
async def reverse_voided_invoice(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """Reverse a voided invoice back to draft status"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can reverse voided invoices")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if invoice.get("status") != "voided":
        raise HTTPException(status_code=400, detail="Only voided invoices can be reversed")
    
    await create_audit_trail_entry(
        action="Invoice Void Reversed",
        record_reference=f"{invoice.get('company_name')} - {invoice.get('invoice_number')}",
        entity_type="invoice",
        entity_id=invoice_id,
        changed_by=current_user,
        reason="Void reversed by admin",
        field_changed="status",
        from_value="voided",
        to_value="auto_draft"
    )
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {
            "$set": {
                "status": "auto_draft",
                "updated_at": get_malaysia_time().isoformat()
            },
            "$unset": {
                "void_reason": "",
                "voided_at": "",
                "voided_by": "",
                "approved_at": "",
                "approved_by": "",
                "issued_at": "",
                "issued_by": ""
            }
        }
    )
    
    if invoice.get("session_id"):
        await db.sessions.update_one(
            {"id": invoice.get("session_id")},
            {"$set": {"invoice_status": "auto_draft"}}
        )
    
    return {
        "message": "Invoice void reversed successfully",
        "invoice_number": invoice.get("invoice_number"),
        "new_status": "auto_draft"
    }


# ============ ADMIN INVOICE ENDPOINTS ============
@router.get("/admin/invoices")
async def get_admin_invoices(
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all invoices for admin management"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can access")
    
    query = {}
    if status and status != "all":
        query["status"] = status
    if search:
        query["$or"] = [
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"company_name": {"$regex": search, "$options": "i"}}
        ]
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return invoices


@router.put("/admin/invoices/{invoice_id}/number")
async def edit_invoice_number(
    invoice_id: str,
    request: InvoiceNumberEditRequest,
    current_user: User = Depends(get_current_user)
):
    """Edit invoice number (year/month/sequence)"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can edit invoice numbers")
    
    if not request.reason or len(request.reason.strip()) < 5:
        raise HTTPException(status_code=400, detail="Reason is required (minimum 5 characters)")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    old_invoice_number = invoice["invoice_number"]
    new_invoice_number = f"INV/MDDRC/{request.year}/{request.month:02d}/{request.sequence:04d}"
    
    existing = await db.invoices.find_one({
        "invoice_number": new_invoice_number,
        "id": {"$ne": invoice_id}
    })
    if existing:
        raise HTTPException(status_code=400, detail=f"Invoice number {new_invoice_number} already exists")
    
    company_name = invoice.get("company_name", "Unknown")
    total_amount = invoice.get("total_amount", 0)
    record_ref = f"{company_name} - RM {total_amount:,.2f}"
    
    await create_audit_trail_entry(
        action="Invoice Number Changed",
        record_reference=record_ref,
        entity_type="invoice",
        entity_id=invoice_id,
        changed_by=current_user,
        reason=request.reason,
        field_changed="invoice_number",
        from_value=old_invoice_number,
        to_value=new_invoice_number
    )
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "invoice_number": new_invoice_number,
            "updated_at": get_malaysia_time().isoformat()
        }}
    )
    
    return {
        "message": "Invoice number updated successfully",
        "old_number": old_invoice_number,
        "new_number": new_invoice_number
    }


@router.post("/admin/invoices/{invoice_id}/void")
async def void_invoice(
    invoice_id: str,
    request: VoidInvoiceRequest,
    current_user: User = Depends(get_current_user)
):
    """Void an invoice"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can void invoices")
    
    if not request.reason or len(request.reason.strip()) < 5:
        raise HTTPException(status_code=400, detail="Reason is required (minimum 5 characters)")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if invoice.get("status") == "voided":
        raise HTTPException(status_code=400, detail="Invoice is already voided")
    
    old_status = invoice.get("status")
    company_name = invoice.get("company_name", "Unknown")
    total_amount = invoice.get("total_amount", 0)
    record_ref = f"{company_name} - RM {total_amount:,.2f}"
    
    await create_audit_trail_entry(
        action="Invoice Voided",
        record_reference=record_ref,
        entity_type="invoice",
        entity_id=invoice_id,
        changed_by=current_user,
        reason=request.reason,
        field_changed="status",
        from_value=old_status,
        to_value="voided"
    )
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": "voided",
            "voided_by": current_user.id,
            "voided_at": get_malaysia_time().isoformat(),
            "void_reason": request.reason,
            "updated_at": get_malaysia_time().isoformat()
        }}
    )
    
    return {"message": "Invoice voided successfully", "invoice_number": invoice.get("invoice_number")}


@router.put("/admin/invoices/{invoice_id}/edit-paid")
async def edit_paid_invoice(
    invoice_id: str,
    request: EditPaidInvoiceRequest,
    current_user: User = Depends(get_current_user)
):
    """Edit a paid invoice"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can edit paid invoices")
    
    if not request.reason or len(request.reason.strip()) < 5:
        raise HTTPException(status_code=400, detail="Reason is required (minimum 5 characters)")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    company_name = invoice.get("company_name", "Unknown")
    total_amount = invoice.get("total_amount", 0)
    record_ref = f"{company_name} - RM {total_amount:,.2f}"
    
    update_data = {"updated_at": get_malaysia_time().isoformat()}
    changes = []
    
    if request.bill_to_name is not None and request.bill_to_name != invoice.get("bill_to_name"):
        changes.append(f"Bill To: {invoice.get('bill_to_name')} → {request.bill_to_name}")
        update_data["bill_to_name"] = request.bill_to_name
    
    if request.bill_to_address is not None and request.bill_to_address != invoice.get("bill_to_address"):
        changes.append(f"Address changed")
        update_data["bill_to_address"] = request.bill_to_address
    
    if request.total_amount is not None and request.total_amount != invoice.get("total_amount"):
        changes.append(f"Amount: RM {invoice.get('total_amount'):,.2f} → RM {request.total_amount:,.2f}")
        update_data["total_amount"] = request.total_amount
    
    if not changes:
        return {"message": "No changes detected"}
    
    await create_audit_trail_entry(
        action="Paid Invoice Edited",
        record_reference=record_ref,
        entity_type="invoice",
        entity_id=invoice_id,
        changed_by=current_user,
        reason=request.reason,
        field_changed="multiple",
        from_value="; ".join([c.split(" → ")[0] if " → " in c else c for c in changes]),
        to_value="; ".join([c.split(" → ")[1] if " → " in c else c for c in changes])
    )
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    
    return {"message": "Paid invoice updated successfully", "changes": changes}


@router.put("/admin/invoices/{invoice_id}/backdate")
async def backdate_invoice(
    invoice_id: str,
    request: BackdateInvoiceRequest,
    current_user: User = Depends(get_current_user)
):
    """Backdate an invoice"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can backdate invoices")
    
    if not request.reason or len(request.reason.strip()) < 5:
        raise HTTPException(status_code=400, detail="Reason is required (minimum 5 characters)")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    company_name = invoice.get("company_name", "Unknown")
    total_amount = invoice.get("total_amount", 0)
    record_ref = f"{company_name} - RM {total_amount:,.2f}"
    
    old_created_at = invoice.get("created_at")
    if isinstance(old_created_at, datetime):
        old_date = old_created_at.strftime("%Y-%m-%d")
    elif isinstance(old_created_at, str):
        old_date = old_created_at[:10]
    else:
        old_date = "Unknown"
    
    try:
        new_datetime = datetime.strptime(request.new_date, "%Y-%m-%d")
        new_datetime = new_datetime.replace(tzinfo=MALAYSIA_TZ)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    await create_audit_trail_entry(
        action="Invoice Backdated",
        record_reference=record_ref,
        entity_type="invoice",
        entity_id=invoice_id,
        changed_by=current_user,
        reason=request.reason,
        field_changed="created_at",
        from_value=old_date,
        to_value=request.new_date
    )
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "created_at": new_datetime.isoformat(),
            "invoice_date": request.new_date,
            "updated_at": get_malaysia_time().isoformat()
        }}
    )
    
    return {"message": "Invoice backdated successfully", "old_date": old_date, "new_date": request.new_date}


@router.post("/admin/sequence/reset")
async def reset_invoice_sequence(
    request: ResetSequenceRequest,
    current_user: User = Depends(get_current_user)
):
    """Reset invoice sequence counter"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can reset sequence")
    
    if not request.reason or len(request.reason.strip()) < 5:
        raise HTTPException(status_code=400, detail="Reason is required (minimum 5 characters)")
    
    if request.new_sequence < 1:
        raise HTTPException(status_code=400, detail="Sequence must be at least 1")
    
    prefix = f"INV/MDDRC/{request.year}/{request.month:02d}/"
    last_invoice = await db.invoices.find_one(
        {"invoice_number": {"$regex": f"^{prefix}"}},
        sort=[("invoice_number", -1)]
    )
    
    current_sequence = 0
    if last_invoice:
        current_sequence = int(last_invoice["invoice_number"].split("/")[-1])
    
    await db.invoice_sequence_settings.update_one(
        {"year": request.year, "month": request.month},
        {
            "$set": {
                "next_sequence": request.new_sequence,
                "reset_by": current_user.id,
                "reset_at": get_malaysia_time().isoformat(),
                "reset_reason": request.reason
            }
        },
        upsert=True
    )
    
    await create_audit_trail_entry(
        action="Invoice Sequence Reset",
        record_reference=f"Sequence for {request.year}/{request.month:02d}",
        entity_type="sequence",
        entity_id=f"{request.year}-{request.month:02d}",
        changed_by=current_user,
        reason=request.reason,
        field_changed="next_sequence",
        from_value=str(current_sequence),
        to_value=str(request.new_sequence)
    )
    
    return {
        "message": "Sequence reset successfully",
        "year": request.year,
        "month": request.month,
        "old_sequence": current_sequence,
        "new_sequence": request.new_sequence
    }


@router.put("/admin/invoices/{invoice_id}/override")
async def override_invoice_validation(
    invoice_id: str,
    request: OverrideValidationRequest,
    current_user: User = Depends(get_current_user)
):
    """Override invoice amount without validation"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can override validation")
    
    if not request.reason or len(request.reason.strip()) < 5:
        raise HTTPException(status_code=400, detail="Reason is required (minimum 5 characters)")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    company_name = invoice.get("company_name", "Unknown")
    old_amount = invoice.get("total_amount", 0)
    record_ref = f"{company_name} - RM {old_amount:,.2f}"
    
    await create_audit_trail_entry(
        action="Invoice Amount Override",
        record_reference=record_ref,
        entity_type="invoice",
        entity_id=invoice_id,
        changed_by=current_user,
        reason=request.reason,
        field_changed="total_amount",
        from_value=f"RM {old_amount:,.2f}",
        to_value=f"RM {request.total_amount:,.2f}"
    )
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "total_amount": request.total_amount,
            "validation_overridden": True,
            "override_reason": request.reason,
            "overridden_by": current_user.id,
            "overridden_at": get_malaysia_time().isoformat(),
            "updated_at": get_malaysia_time().isoformat()
        }}
    )
    
    return {
        "message": "Invoice amount overridden successfully",
        "old_amount": old_amount,
        "new_amount": request.total_amount
    }


@router.get("/admin/audit-trail")
async def get_admin_audit_trail(
    entity_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    """Get audit trail with filters"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can access audit trail")
    
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    
    if start_date or end_date:
        query["timestamp"] = {}
        if start_date:
            query["timestamp"]["$gte"] = start_date
        if end_date:
            query["timestamp"]["$lte"] = end_date + "T23:59:59"
    
    entries = await db.audit_trail.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    return entries


@router.get("/admin/audit-trail/export")
async def export_audit_trail(
    entity_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export audit trail as list"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin and Finance can export audit trail")
    
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    
    if start_date or end_date:
        query["timestamp"] = {}
        if start_date:
            query["timestamp"]["$gte"] = start_date
        if end_date:
            query["timestamp"]["$lte"] = end_date + "T23:59:59"
    
    entries = await db.audit_trail.find(query, {"_id": 0}).sort("timestamp", -1).to_list(10000)
    
    export_data = []
    for entry in entries:
        export_data.append({
            "Timestamp": entry.get("timestamp", ""),
            "Action": entry.get("action", ""),
            "Record Reference": entry.get("record_reference", ""),
            "Field Changed": entry.get("field_changed", ""),
            "From Value": entry.get("from_value", ""),
            "To Value": entry.get("to_value", ""),
            "Changed By": entry.get("changed_by_name", ""),
            "Reason": entry.get("reason", "")
        })
    
    return export_data


# ==================== DELETE INVOICE ====================

class DeleteInvoiceRequest(BaseModel):
    reason: str
    reuse_number: bool = True  # Whether to add invoice number to reuse pool


@router.delete("/invoices/{invoice_id}")
async def delete_invoice(
    invoice_id: str,
    request: DeleteInvoiceRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Permanently delete an invoice and sync with related data.
    - Removes invoice from database
    - Updates session to remove invoice_id reference
    - Adds invoice number to reuse pool (if auto_draft)
    - Logs the deletion in audit trail
    """
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin can delete invoices")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    invoice_number = invoice.get("invoice_number", "")
    session_id = invoice.get("session_id")
    company_name = invoice.get("company_name", "Unknown")
    total_amount = invoice.get("total_amount", 0)
    status = invoice.get("status", "")
    is_additional = invoice.get("is_additional_invoice", False)
    
    # Check if this is an issued/paid invoice - warn but allow
    if status in ["issued", "paid"]:
        # For issued/paid invoices, we still delete but log it prominently
        pass
    
    # Create audit trail entry BEFORE deleting
    record_ref = f"{invoice_number} - {company_name} - RM {total_amount:,.2f}"
    await create_audit_trail_entry(
        action="Invoice Deleted",
        record_reference=record_ref,
        entity_type="invoice",
        entity_id=invoice_id,
        changed_by=current_user,
        reason=request.reason,
        field_changed="deleted",
        from_value=f"Status: {status}, Amount: RM {total_amount:,.2f}",
        to_value="DELETED"
    )
    
    # If it's a draft invoice and user wants to reuse the number, add to pool
    if request.reuse_number and status in ["auto_draft", "draft"] and invoice_number:
        # Extract session name for reference
        session = await db.sessions.find_one({"id": session_id}, {"_id": 0, "name": 1}) if session_id else None
        session_name = session.get("name") if session else "Unknown Session"
        
        await db.deleted_invoice_numbers.insert_one({
            "invoice_number": invoice_number,
            "deleted_from_session_name": session_name,
            "deleted_at": get_malaysia_time().isoformat(),
            "deleted_by": current_user.id,
            "deletion_reason": request.reason,
            "original_status": status,
            "original_amount": total_amount,
            "is_reused": False,
            "reused_at": None
        })
    
    # Update session to remove invoice reference
    if session_id:
        if is_additional:
            # For additional invoices, remove from session's additional_invoice_ids array
            await db.sessions.update_one(
                {"id": session_id},
                {"$pull": {"additional_invoice_ids": invoice_id}}
            )
        else:
            # For main invoice, clear the invoice_id and status
            await db.sessions.update_one(
                {"id": session_id},
                {"$set": {
                    "invoice_id": None,
                    "invoice_status": None,
                    "updated_at": get_malaysia_time().isoformat()
                }}
            )
    
    # Delete related credit notes if any
    credit_notes_deleted = await db.credit_notes.delete_many({"invoice_id": invoice_id})
    
    # Delete the invoice
    await db.invoices.delete_one({"id": invoice_id})
    
    # Log finance action
    await log_finance_action(
        "invoice", 
        invoice_id, 
        "deleted", 
        current_user.id,
        invoice,  # before value is the entire invoice
        {"deleted": True, "reason": request.reason},
        request.reason
    )
    
    return {
        "message": "Invoice deleted successfully",
        "invoice_number": invoice_number,
        "session_updated": session_id is not None,
        "credit_notes_deleted": credit_notes_deleted.deleted_count,
        "number_added_to_reuse_pool": request.reuse_number and status in ["auto_draft", "draft"]
    }
