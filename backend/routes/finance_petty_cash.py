"""
Finance Petty Cash & Manual Entries routes
Stage F6: ~14 endpoints
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from typing import Optional
from pydantic import BaseModel
from io import BytesIO
from datetime import datetime
import uuid
import base64
import tempfile
import os as _os

from core import db, get_current_user, get_malaysia_time
from models import User

try:
    from routes.accounting import create_auto_journal_entry
except ImportError:
    create_auto_journal_entry = None

router = APIRouter(prefix="/finance", tags=["finance-petty-cash"])


# ============ MODELS ============
class ManualIncomeEntry(BaseModel):
    description: str
    amount: float
    category: str = "Other Income"
    date: str  # YYYY-MM-DD
    notes: Optional[str] = None


class ManualExpenseEntry(BaseModel):
    description: str
    amount: float
    category: str
    date: str  # YYYY-MM-DD
    notes: Optional[str] = None


class PettyCashSetup(BaseModel):
    float_amount: float
    custodian_id: Optional[str] = None
    custodian_name: Optional[str] = None
    approval_threshold: float = 100.0


class PettyCashTransaction(BaseModel):
    type: str
    amount: float
    description: str
    category: Optional[str] = None
    receipt_url: Optional[str] = None
    date: str
    notes: Optional[str] = None


class PettyCashReconciliation(BaseModel):
    physical_count: float
    notes: Optional[str] = None


# ============ MANUAL INCOME ENDPOINTS ============
@router.post("/manual-income")
async def add_manual_income(entry: ManualIncomeEntry, current_user: User = Depends(get_current_user)):
    """Add a one-off manual income entry"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    record = {
        "id": str(uuid.uuid4()),
        "description": entry.description,
        "amount": entry.amount,
        "category": entry.category,
        "date": entry.date,
        "notes": entry.notes,
        "created_by": current_user.id,
        "created_at": get_malaysia_time().isoformat()
    }
    await db.manual_income.insert_one(record)
    # Auto-post journal entry
    if create_auto_journal_entry:
        try:
            await create_auto_journal_entry(
                entry_date=entry.date, description=f"Manual Income: {entry.description}",
                source_module="manual_income", source_id=record["id"], source_reference=f"MI-{record['id'][:8]}",
                lines=[
                    {"account_code": "1000", "debit": round(entry.amount, 2), "credit": 0, "memo": "Cash received"},
                    {"account_code": "4100", "debit": 0, "credit": round(entry.amount, 2), "memo": entry.description},
                ],
                created_by_id=current_user.id, created_by_name=current_user.full_name, skip_date_check=True
            )
        except Exception:
            pass  # Non-blocking — backfill will catch it
    return {"message": "Income entry added", "id": record["id"]}


@router.get("/manual-income")
async def get_manual_income(year: int = None, current_user: User = Depends(get_current_user)):
    """Get manual income entries"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if year:
        query["date"] = {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}
    
    entries = await db.manual_income.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return entries


@router.delete("/manual-income/{entry_id}")
async def delete_manual_income(entry_id: str, current_user: User = Depends(get_current_user)):
    """Delete a manual income entry"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = await db.manual_income.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted"}


# ============ MANUAL EXPENSE ENDPOINTS ============
@router.post("/manual-expense")
async def add_manual_expense(entry: ManualExpenseEntry, current_user: User = Depends(get_current_user)):
    """Add a one-off manual expense entry"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    record = {
        "id": str(uuid.uuid4()),
        "description": entry.description,
        "amount": entry.amount,
        "category": entry.category,
        "date": entry.date,
        "notes": entry.notes,
        "created_by": current_user.id,
        "created_at": get_malaysia_time().isoformat()
    }
    await db.manual_expenses.insert_one(record)
    # Auto-post journal entry
    if create_auto_journal_entry:
        try:
            await create_auto_journal_entry(
                entry_date=entry.date, description=f"Manual Expense: {entry.description}",
                source_module="manual_expense", source_id=record["id"], source_reference=f"ME-{record['id'][:8]}",
                lines=[
                    {"account_code": "6999", "debit": round(entry.amount, 2), "credit": 0, "memo": entry.description},
                    {"account_code": "1000", "debit": 0, "credit": round(entry.amount, 2), "memo": "Cash paid"},
                ],
                created_by_id=current_user.id, created_by_name=current_user.full_name, skip_date_check=True
            )
        except Exception:
            pass
    return {"message": "Expense entry added", "id": record["id"]}


@router.get("/manual-expenses")
async def get_manual_expenses(year: int = None, current_user: User = Depends(get_current_user)):
    """Get manual expense entries"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if year:
        query["date"] = {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}
    
    entries = await db.manual_expenses.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return entries


@router.delete("/manual-expense/{entry_id}")
async def delete_manual_expense(entry_id: str, current_user: User = Depends(get_current_user)):
    """Delete a manual expense entry"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    result = await db.manual_expenses.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted"}


# ============ PETTY CASH ENDPOINTS ============
@router.get("/petty-cash/settings")
async def get_petty_cash_settings(current_user: User = Depends(get_current_user)):
    """Get petty cash settings"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    settings = await db.petty_cash_settings.find_one({}, {"_id": 0})
    if not settings:
        settings = {
            "float_amount": 500.0,
            "current_balance": 500.0,
            "custodian_id": None,
            "custodian_name": None,
            "approval_threshold": 100.0,
            "last_reconciliation": None
        }
    return settings


@router.post("/petty-cash/setup")
async def setup_petty_cash(setup: PettyCashSetup, current_user: User = Depends(get_current_user)):
    """Setup or update petty cash settings"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    existing = await db.petty_cash_settings.find_one({})
    
    settings = {
        "float_amount": setup.float_amount,
        "current_balance": setup.float_amount if not existing else existing.get("current_balance", setup.float_amount),
        "custodian_id": setup.custodian_id,
        "custodian_name": setup.custodian_name,
        "approval_threshold": setup.approval_threshold,
        "updated_at": get_malaysia_time().isoformat(),
        "updated_by": current_user.id
    }
    
    if existing:
        await db.petty_cash_settings.update_one({}, {"$set": settings})
    else:
        settings["created_at"] = get_malaysia_time().isoformat()
        settings["last_reconciliation"] = None
        await db.petty_cash_settings.insert_one(settings)
    
    return {"message": "Petty cash settings updated", "settings": {k: v for k, v in settings.items() if k != "_id"}}


@router.post("/petty-cash/transaction")
async def add_petty_cash_transaction(txn: PettyCashTransaction, current_user: User = Depends(get_current_user)):
    """Add a petty cash transaction"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    settings = await db.petty_cash_settings.find_one({})
    if not settings:
        raise HTTPException(status_code=400, detail="Petty cash not set up")
    
    current_balance = settings.get("current_balance", 0)
    
    if txn.type == "expense":
        if txn.amount > current_balance:
            raise HTTPException(status_code=400, detail=f"Insufficient balance. Current: RM {current_balance:.2f}")
        new_balance = current_balance - txn.amount
    elif txn.type == "topup":
        new_balance = current_balance + txn.amount
    else:
        raise HTTPException(status_code=400, detail="Invalid type")
    
    requires_approval = txn.type == "expense" and txn.amount > settings.get("approval_threshold", 100)
    
    transaction = {
        "id": str(uuid.uuid4()),
        "type": txn.type,
        "amount": txn.amount,
        "description": txn.description,
        "category": txn.category or "Miscellaneous",
        "receipt_url": txn.receipt_url,
        "date": txn.date,
        "notes": txn.notes,
        "balance_before": current_balance,
        "balance_after": new_balance,
        "status": "pending" if requires_approval else "approved",
        "created_by": current_user.id,
        "created_by_name": current_user.full_name,
        "created_at": get_malaysia_time().isoformat(),
        "approved_by": None if requires_approval else current_user.id,
        "approved_at": None if requires_approval else get_malaysia_time().isoformat()
    }
    
    await db.petty_cash_transactions.insert_one(transaction)
    
    if not requires_approval:
        await db.petty_cash_settings.update_one({}, {"$set": {"current_balance": new_balance}})
        # Auto-post journal entry for approved petty cash
        if create_auto_journal_entry:
            try:
                if txn.type == "expense":
                    lines = [
                        {"account_code": "6600", "debit": round(txn.amount, 2), "credit": 0, "memo": txn.description},
                        {"account_code": "1010", "debit": 0, "credit": round(txn.amount, 2), "memo": "Petty Cash"},
                    ]
                else:  # topup
                    lines = [
                        {"account_code": "1010", "debit": round(txn.amount, 2), "credit": 0, "memo": "Petty Cash Top-up"},
                        {"account_code": "1000", "debit": 0, "credit": round(txn.amount, 2), "memo": "Cash at Bank"},
                    ]
                await create_auto_journal_entry(
                    entry_date=txn.date, description=f"Petty Cash {txn.type}: {txn.description}",
                    source_module="petty_cash", source_id=transaction["id"], source_reference=f"PC-{transaction['id'][:8]}",
                    lines=lines,
                    created_by_id=current_user.id, created_by_name=current_user.full_name, skip_date_check=True
                )
            except Exception:
                pass
    
    return {
        "message": "Transaction added" + (" (pending approval)" if requires_approval else ""),
        "transaction_id": transaction["id"],
        "new_balance": new_balance if not requires_approval else current_balance,
        "requires_approval": requires_approval
    }


@router.get("/petty-cash/transactions")
async def get_petty_cash_transactions(
    year: int = None,
    month: int = None,
    status: str = None,
    current_user: User = Depends(get_current_user)
):
    """Get petty cash transactions"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if year:
        start = f"{year}-01-01"
        end = f"{year}-12-31"
        if month:
            start = f"{year}-{month:02d}-01"
            end = f"{year}-{month:02d}-31"
        query["date"] = {"$gte": start, "$lte": end}
    if status:
        query["status"] = status
    
    transactions = await db.petty_cash_transactions.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return transactions


@router.post("/petty-cash/approve/{transaction_id}")
async def approve_petty_cash_transaction(transaction_id: str, current_user: User = Depends(get_current_user)):
    """Approve a pending transaction"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    txn = await db.petty_cash_transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Not found")
    if txn.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Not pending")
    
    await db.petty_cash_transactions.update_one(
        {"id": transaction_id},
        {"$set": {"status": "approved", "approved_by": current_user.id, "approved_at": get_malaysia_time().isoformat()}}
    )
    
    settings = await db.petty_cash_settings.find_one({})
    new_balance = settings.get("current_balance", 0)
    if txn.get("type") == "expense":
        new_balance -= txn.get("amount", 0)
    elif txn.get("type") == "topup":
        new_balance += txn.get("amount", 0)
    
    await db.petty_cash_settings.update_one({}, {"$set": {"current_balance": new_balance}})
    return {"message": "Approved", "new_balance": new_balance}


@router.post("/petty-cash/reject/{transaction_id}")
async def reject_petty_cash_transaction(transaction_id: str, current_user: User = Depends(get_current_user)):
    """Reject a pending transaction"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    txn = await db.petty_cash_transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Not found")
    if txn.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Not pending")
    
    await db.petty_cash_transactions.update_one(
        {"id": transaction_id},
        {"$set": {"status": "rejected", "rejected_by": current_user.id, "rejected_at": get_malaysia_time().isoformat()}}
    )
    return {"message": "Rejected"}


@router.delete("/petty-cash/transaction/{transaction_id}")
async def delete_petty_cash_transaction(transaction_id: str, current_user: User = Depends(get_current_user)):
    """Delete a transaction"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    txn = await db.petty_cash_transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Not found")
    
    if txn.get("status") == "approved":
        settings = await db.petty_cash_settings.find_one({})
        current_balance = settings.get("current_balance", 0)
        if txn.get("type") == "expense":
            new_balance = current_balance + txn.get("amount", 0)
        elif txn.get("type") == "topup":
            new_balance = current_balance - txn.get("amount", 0)
        else:
            new_balance = current_balance
        await db.petty_cash_settings.update_one({}, {"$set": {"current_balance": new_balance}})
    
    await db.petty_cash_transactions.delete_one({"id": transaction_id})
    return {"message": "Deleted"}


@router.post("/petty-cash/reconcile")
async def reconcile_petty_cash(recon: PettyCashReconciliation, current_user: User = Depends(get_current_user)):
    """Reconcile petty cash"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    settings = await db.petty_cash_settings.find_one({})
    if not settings:
        raise HTTPException(status_code=400, detail="Not set up")
    
    system_balance = settings.get("current_balance", 0)
    variance = recon.physical_count - system_balance
    
    reconciliation = {
        "id": str(uuid.uuid4()),
        "date": get_malaysia_time().isoformat()[:10],
        "system_balance": system_balance,
        "physical_count": recon.physical_count,
        "variance": variance,
        "notes": recon.notes,
        "reconciled_by": current_user.id,
        "reconciled_by_name": current_user.full_name,
        "created_at": get_malaysia_time().isoformat()
    }
    
    await db.petty_cash_reconciliations.insert_one(reconciliation)
    await db.petty_cash_settings.update_one({}, {"$set": {"current_balance": recon.physical_count, "last_reconciliation": reconciliation["date"]}})
    
    if abs(variance) > 0.01:
        adjustment = {
            "id": str(uuid.uuid4()),
            "type": "adjustment",
            "amount": abs(variance),
            "description": f"Reconciliation adjustment",
            "category": "Adjustment",
            "date": get_malaysia_time().isoformat()[:10],
            "notes": recon.notes,
            "balance_before": system_balance,
            "balance_after": recon.physical_count,
            "status": "approved",
            "created_by": current_user.id,
            "created_by_name": current_user.full_name,
            "created_at": get_malaysia_time().isoformat(),
            "approved_by": current_user.id,
            "approved_at": get_malaysia_time().isoformat()
        }
        await db.petty_cash_transactions.insert_one(adjustment)
    
    return {"message": "Complete", "system_balance": system_balance, "physical_count": recon.physical_count, "variance": variance, "new_balance": recon.physical_count}


@router.get("/petty-cash/reconciliations")
async def get_reconciliation_history(current_user: User = Depends(get_current_user)):
    """Get reconciliation history"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    reconciliations = await db.petty_cash_reconciliations.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return reconciliations


@router.get("/petty-cash/summary")
async def get_petty_cash_summary(year: int = None, current_user: User = Depends(get_current_user)):
    """Get petty cash summary"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    now = get_malaysia_time()
    year = year or now.year
    
    query = {"date": {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}, "type": "expense", "status": "approved"}
    transactions = await db.petty_cash_transactions.find(query, {"_id": 0}).to_list(10000)
    
    by_category = {}
    for txn in transactions:
        cat = txn.get("category", "Miscellaneous")
        if cat not in by_category:
            by_category[cat] = {"count": 0, "total": 0}
        by_category[cat]["count"] += 1
        by_category[cat]["total"] += txn.get("amount", 0)
    
    by_month = {}
    for txn in transactions:
        try:
            m = int(txn.get("date", "")[5:7])
            if m not in by_month:
                by_month[m] = 0
            by_month[m] += txn.get("amount", 0)
        except:
            pass
    
    settings = await db.petty_cash_settings.find_one({}, {"_id": 0})
    
    return {
        "year": year,
        "current_balance": settings.get("current_balance", 0) if settings else 0,
        "float_amount": settings.get("float_amount", 0) if settings else 0,
        "by_category": by_category,
        "by_month": by_month,
        "total_expenses": sum(c["total"] for c in by_category.values())
    }



# ============ EXPORT ENDPOINTS ============
async def _load_export_data(start_date: str, end_date: str, current_user: User):
    """Load petty cash transactions + settings + custodian signature for export."""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")

    query = {
        "date": {"$gte": start_date, "$lte": end_date},
        "status": {"$in": ["approved", "pending"]},
    }
    txns = await db.petty_cash_transactions.find(query, {"_id": 0}).sort("date", 1).to_list(5000)

    settings = await db.petty_cash_settings.find_one({}, {"_id": 0}) or {}
    company = await db.company_settings.find_one({}, {"_id": 0}) or {}

    custodian_sig = ""
    approver_sig = current_user.digital_signature or ""
    custodian_id = settings.get("custodian_id")
    if custodian_id:
        cust = await db.users.find_one({"id": custodian_id}, {"_id": 0, "digital_signature": 1})
        if cust:
            custodian_sig = cust.get("digital_signature") or ""

    return txns, settings, company, custodian_sig, approver_sig


@router.get("/petty-cash/export/excel")
async def export_petty_cash_excel(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user),
):
    """Export petty cash transactions to Excel within a date range"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    txns, settings, company, _, _ = await _load_export_data(start_date, end_date, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Petty Cash Log"

    header_fill = PatternFill("solid", fgColor="1A365D")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1A365D")
    total_fill = PatternFill("solid", fgColor="FFF3CD")
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title rows
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{company.get('company_name', 'MDDRC')} — Petty Cash Claim"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Period: {start_date} to {end_date}  |  Custodian: {settings.get('custodian_name', '-')}  |  Float: RM {settings.get('float_amount', 0):,.2f}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, size=10, color="555555")

    ws.append([])  # row 3 blank

    headers = ["BIL", "Date", "Item", "Category", "Description", "Debit (RM)", "Credit (RM)", "Balance (RM)", "Receipt"]
    ws.append(headers)
    header_row_idx = ws.max_row
    for col_idx, _h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Determine opening balance (balance_before of first txn)
    opening_balance = txns[0].get("balance_before", 0) if txns else settings.get("current_balance", 0)
    ws.append(["", "", "Opening Balance", "", "", "", "", round(opening_balance, 2), ""])
    opening_row = ws.max_row
    ws.cell(row=opening_row, column=3).font = Font(bold=True)
    ws.cell(row=opening_row, column=8).font = Font(bold=True)
    for c in range(1, 10):
        ws.cell(row=opening_row, column=c).border = border

    total_debit = 0.0
    total_credit = 0.0

    for idx, t in enumerate(txns, start=1):
        is_expense = t.get("type") == "expense"
        is_topup = t.get("type") == "topup"
        amt = float(t.get("amount", 0) or 0)
        debit = amt if is_expense else 0.0  # expense out of petty cash
        credit = amt if is_topup else 0.0   # top-up into petty cash
        total_debit += debit
        total_credit += credit

        receipt_mark = "YES" if t.get("receipt_url") else ""

        ws.append([
            idx,
            t.get("date", ""),
            t.get("description", ""),
            t.get("category", ""),
            t.get("notes", "") or "",
            round(debit, 2) if debit else "",
            round(credit, 2) if credit else "",
            round(float(t.get("balance_after", 0) or 0), 2),
            receipt_mark,
        ])
        row = ws.max_row
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        for c in (6, 7, 8):
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="right")
        if t.get("status") == "pending":
            ws.cell(row=row, column=3).font = Font(italic=True, color="B58900")

    # Totals row
    closing_balance = txns[-1].get("balance_after", opening_balance) if txns else opening_balance
    ws.append(["", "", "TOTAL", "", "", round(total_debit, 2), round(total_credit, 2), round(float(closing_balance), 2), ""])
    total_row = ws.max_row
    for c in range(1, 10):
        ws.cell(row=total_row, column=c).fill = total_fill
        ws.cell(row=total_row, column=c).font = Font(bold=True)
        ws.cell(row=total_row, column=c).border = border
    for c in (6, 7, 8):
        ws.cell(row=total_row, column=c).alignment = Alignment(horizontal="right")

    # Signature block
    ws.append([])
    ws.append([])
    sig_row = ws.max_row + 1
    ws.cell(row=sig_row, column=1, value="Prepared by (Custodian):").font = Font(bold=True)
    ws.cell(row=sig_row, column=5, value="Approved by:").font = Font(bold=True)
    ws.cell(row=sig_row + 3, column=1, value=f"Name: {settings.get('custodian_name', '__________________________')}")
    ws.cell(row=sig_row + 3, column=5, value=f"Name: {current_user.full_name}")
    ws.cell(row=sig_row + 4, column=1, value="Date: __________________________")
    ws.cell(row=sig_row + 4, column=5, value=f"Date: {get_malaysia_time().strftime('%d %b %Y')}")

    # Column widths
    widths = [6, 12, 32, 18, 30, 13, 13, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"PettyCash_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/petty-cash/export/pdf")
async def export_petty_cash_pdf(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user),
):
    """Export petty cash transactions to a printable PDF within a date range"""
    from fpdf import FPDF

    txns, settings, company, custodian_sig, approver_sig = await _load_export_data(start_date, end_date, current_user)

    def _s(t):
        if t is None:
            return ""
        t = str(t)
        repl = {"\u2013": "-", "\u2014": "-", "\u2018": "'", "\u2019": "'", "\u201c": '"', "\u201d": '"', "\u2022": "*"}
        for k, v in repl.items():
            t = t.replace(k, v)
        return "".join(c if ord(c) < 128 else "?" for c in t)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(26, 54, 93)
    pdf.cell(0, 8, _s(company.get("company_name", "MDDRC")) + " - Petty Cash Claim", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(
        0, 6,
        f"Period: {start_date} to {end_date}   |   Custodian: {_s(settings.get('custodian_name', '-'))}   |   Float: RM {settings.get('float_amount', 0):,.2f}",
        ln=True, align="C",
    )
    pdf.ln(3)

    # Header row
    pdf.set_fill_color(26, 54, 93)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    # widths (total ≈ 277)
    cw = [10, 22, 60, 30, 60, 25, 25, 27, 18]
    headers = ["BIL", "Date", "Item", "Category", "Description", "Debit (RM)", "Credit (RM)", "Balance (RM)", "Receipt"]
    for w, h in zip(cw, headers):
        pdf.cell(w, 7, h, border=1, align="C", fill=True)
    pdf.ln()

    # Opening balance row
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_fill_color(245, 245, 245)
    opening_balance = txns[0].get("balance_before", 0) if txns else settings.get("current_balance", 0)
    pdf.cell(cw[0], 6, "", border=1, fill=True)
    pdf.cell(cw[1], 6, "", border=1, fill=True)
    pdf.cell(cw[2], 6, "Opening Balance", border=1, fill=True)
    pdf.cell(cw[3], 6, "", border=1, fill=True)
    pdf.cell(cw[4], 6, "", border=1, fill=True)
    pdf.cell(cw[5], 6, "", border=1, fill=True)
    pdf.cell(cw[6], 6, "", border=1, fill=True)
    pdf.cell(cw[7], 6, f"{float(opening_balance):,.2f}", border=1, align="R", fill=True)
    pdf.cell(cw[8], 6, "", border=1, fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    total_debit = 0.0
    total_credit = 0.0
    alt = False
    for idx, t in enumerate(txns, start=1):
        is_expense = t.get("type") == "expense"
        is_topup = t.get("type") == "topup"
        amt = float(t.get("amount", 0) or 0)
        debit = amt if is_expense else 0.0
        credit = amt if is_topup else 0.0
        total_debit += debit
        total_credit += credit

        item = _s(t.get("description", ""))[:55]
        category = _s(t.get("category", ""))[:22]
        desc = _s(t.get("notes", "") or "")[:55]

        if alt:
            pdf.set_fill_color(250, 250, 250)
        else:
            pdf.set_fill_color(255, 255, 255)
        alt = not alt

        pdf.cell(cw[0], 6, str(idx), border=1, align="C", fill=True)
        pdf.cell(cw[1], 6, _s(t.get("date", "")), border=1, align="C", fill=True)
        pdf.cell(cw[2], 6, item, border=1, fill=True)
        pdf.cell(cw[3], 6, category, border=1, fill=True)
        pdf.cell(cw[4], 6, desc, border=1, fill=True)
        pdf.cell(cw[5], 6, f"{debit:,.2f}" if debit else "", border=1, align="R", fill=True)
        pdf.cell(cw[6], 6, f"{credit:,.2f}" if credit else "", border=1, align="R", fill=True)
        pdf.cell(cw[7], 6, f"{float(t.get('balance_after', 0) or 0):,.2f}", border=1, align="R", fill=True)
        pdf.cell(cw[8], 6, "YES" if t.get("receipt_url") else "", border=1, align="C", fill=True)
        pdf.ln()

    # Totals
    closing_balance = txns[-1].get("balance_after", opening_balance) if txns else opening_balance
    pdf.set_fill_color(255, 243, 205)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(cw[0] + cw[1] + cw[2] + cw[3] + cw[4], 7, "TOTAL", border=1, align="R", fill=True)
    pdf.cell(cw[5], 7, f"{total_debit:,.2f}", border=1, align="R", fill=True)
    pdf.cell(cw[6], 7, f"{total_credit:,.2f}", border=1, align="R", fill=True)
    pdf.cell(cw[7], 7, f"{float(closing_balance):,.2f}", border=1, align="R", fill=True)
    pdf.cell(cw[8], 7, "", border=1, fill=True)
    pdf.ln(12)

    # Signature blocks
    def _embed_sig(sig_data, x, y, h=16):
        if not sig_data:
            return False
        try:
            data = sig_data
            if data.startswith("data:image"):
                data = data.split(",", 1)[1]
            raw = base64.b64decode(data)
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp.write(raw)
                tmp_path = tmp.name
            pdf.image(tmp_path, x=x, y=y, h=h)
            _os.unlink(tmp_path)
            return True
        except Exception:
            return False

    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(0, 0, 0)
    y_sig = pdf.get_y()
    pdf.set_xy(15, y_sig)
    pdf.cell(120, 6, "Prepared by (Custodian):", ln=False)
    pdf.set_xy(160, y_sig)
    pdf.cell(120, 6, "Approved by:", ln=True)

    sig_y = pdf.get_y() + 2
    _embed_sig(custodian_sig, x=15, y=sig_y)
    _embed_sig(approver_sig, x=160, y=sig_y)

    pdf.set_y(sig_y + 20)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_x(15)
    pdf.cell(120, 5, f"Name: {_s(settings.get('custodian_name', '__________________________'))}", ln=False)
    pdf.set_x(160)
    pdf.cell(120, 5, f"Name: {_s(current_user.full_name)}", ln=True)
    pdf.set_x(15)
    pdf.cell(120, 5, "Date: __________________________", ln=False)
    pdf.set_x(160)
    pdf.cell(120, 5, f"Date: {get_malaysia_time().strftime('%d %b %Y')}", ln=True)

    pdf_bytes = pdf.output(dest="S")
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode("latin-1")
    buf = BytesIO(bytes(pdf_bytes))
    buf.seek(0)
    fname = f"PettyCash_{start_date}_to_{end_date}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/petty-cash/transaction/{transaction_id}/receipt")
async def get_petty_cash_receipt(transaction_id: str, current_user: User = Depends(get_current_user)):
    """Return the receipt image for a petty cash transaction"""
    if current_user.role not in ["admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    txn = await db.petty_cash_transactions.find_one({"id": transaction_id}, {"_id": 0, "receipt_url": 1})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return {"receipt_url": txn.get("receipt_url", "") or ""}
