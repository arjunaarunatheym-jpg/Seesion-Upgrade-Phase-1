"""
Templates routes - Template download endpoints for Excel templates
Endpoints: 6
- GET /templates/pre-post-assessment
- GET /templates/feedback
- GET /templates/checklist
- GET /templates/program-test-questions
- GET /templates/program-feedback-questions
- GET /templates/program-checklist-items
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from typing import List
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from core import db, get_current_user
from models import User

router = APIRouter(prefix="/templates", tags=["templates"])


@router.get("/pre-post-assessment")
async def download_pre_post_assessment_template(current_user: User = Depends(get_current_user)):
    """Download Excel template for pre/post assessment bulk upload"""
    if current_user.role not in ["admin", "assistant_admin", "coordinator"]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pre-Post Assessment"

    headers = ["Participant Name", "IC Number", "Pre-Test Score", "Pre-Test Total", "Post-Test Score", "Post-Test Total"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Sample data
    ws.append(["Ahmad bin Ali", "901234-56-7890", 15, 20, 18, 20])
    ws.append(["Siti binti Hassan", "880101-14-5678", 12, 20, 16, 20])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Pre_Post_Assessment_Template.xlsx"}
    )


@router.get("/feedback")
async def download_feedback_template(current_user: User = Depends(get_current_user)):
    """Download Excel template for feedback bulk upload"""
    if current_user.role not in ["admin", "assistant_admin", "coordinator"]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback"

    headers = ["Participant Name", "IC Number"]

    # Get feedback questions
    questions = await db.feedback_questions.find({}, {"_id": 0}).sort("order", 1).to_list(None)
    if not questions:
        questions = [
            {"id": "A1", "question": "Penganjur menepati jangkaan saya", "type": "rating"},
            {"id": "A2", "question": "Kandungan kursus adalah jelas dan mudah difahami", "type": "rating"},
            {"id": "A3", "question": "Hasil pembelajaran adalah selari dengan objektif", "type": "rating"},
            {"id": "A4", "question": "Bahan pembelajaran sangat jelas", "type": "rating"},
            {"id": "A5", "question": "Tempoh kursus adalah mencukupi", "type": "rating"},
            {"id": "A6", "question": "Berpuas hati dengan kandungan kursus", "type": "rating"},
            {"id": "A7", "question": "Cadangan KUALITI KURSUS", "type": "text"},
        ]

    for q in questions:
        headers.append(f"{q.get('id', '')}: {q.get('question', '')[:40]}")

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Feedback_Template.xlsx"}
    )


@router.get("/checklist")
async def download_checklist_template(current_user: User = Depends(get_current_user)):
    """Download Excel template for vehicle checklist bulk upload"""
    if current_user.role not in ["admin", "assistant_admin", "coordinator", "trainer"]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Checklist"

    headers = ["Participant Name", "IC Number", "Vehicle Number", "Vehicle Type", "Item", "Status (good/needs_repair)", "Comments"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    ws.append(["Ahmad bin Ali", "901234-56-7890", "WMK1234", "Car", "Brakes", "good", ""])
    ws.append(["Ahmad bin Ali", "901234-56-7890", "WMK1234", "Car", "Tires", "needs_repair", "Front left worn"])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Vehicle_Checklist_Template.xlsx"}
    )


@router.get("/program-test-questions")
async def download_test_questions_template(current_user: User = Depends(get_current_user)):
    """Download Excel template for program test questions upload"""
    if current_user.role not in ["admin", "assistant_admin"]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Questions"

    headers = ["Program Name", "Question Text", "Option A", "Option B", "Option C", "Option D", "Correct Answer (A/B/C/D)", "Difficulty (easy/medium/hard)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    ws.append(["Defensive Driving", "What is the safe following distance?", "1 second", "2 seconds", "3 seconds", "4 seconds", "C", "easy"])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Test_Questions_Template.xlsx"}
    )


@router.get("/program-feedback-questions")
async def download_feedback_questions_template(current_user: User = Depends(get_current_user)):
    """Download Excel template for program feedback questions upload"""
    if current_user.role not in ["admin", "assistant_admin"]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedback Questions"

    headers = ["Program Name", "Question Text", "Question Type (rating/multiple_choice/text)", "Options (comma-separated)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    ws.append(["Defensive Driving", "How would you rate the trainer?", "rating", ""])
    ws.append(["Defensive Driving", "Which topic was most useful?", "multiple_choice", "Braking,Steering,Awareness,All"])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Feedback_Questions_Template.xlsx"}
    )


@router.get("/program-checklist-items")
async def download_checklist_items_template(current_user: User = Depends(get_current_user)):
    """Download Excel template for program checklist items upload"""
    if current_user.role not in ["admin", "assistant_admin"]:
        raise HTTPException(status_code=403, detail="Permission denied")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist Items"

    headers = ["Program Name", "Checklist Item", "Category (optional)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    ws.append(["Defensive Driving", "Check brake pads", "Brakes"])
    ws.append(["Defensive Driving", "Check tire pressure", "Tires"])
    ws.append(["Defensive Driving", "Check mirrors", "Visibility"])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Checklist_Items_Template.xlsx"}
    )
