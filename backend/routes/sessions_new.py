"""
Sessions routes - Training session management
This is a PARTIAL extraction - complex endpoints remain in server.py for now.
Endpoints extracted: Core session CRUD and participant management
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from typing import List, Optional
from datetime import datetime
import uuid
import logging

from core import db, get_current_user, get_malaysia_time, find_or_create_user, get_or_create_participant_access
from models import User, Session, SessionCreate
from utils.email_notifications import notify_session_completed

router = APIRouter(prefix="/sessions", tags=["sessions"])


@router.get("")
async def get_sessions(
    company_id: Optional[str] = None,
    program_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get sessions filtered by user role - main sessions endpoint"""
    from datetime import date as dt_date
    current_date_str = dt_date.today().isoformat()
    
    if current_user.role == "admin":
        # Admin sees all non-completed, non-archived sessions
        query = {
            "$and": [
                {"is_archived": {"$ne": True}},
                {
                    "$or": [
                        {"completion_status": {"$exists": False}},
                        {"completion_status": "ongoing"},
                        {"completion_status": {"$nin": ["completed", "archived"]}}
                    ]
                }
            ]
        }
    elif current_user.role == "trainer":
        # Trainer sees sessions they're assigned to (current/future, non-completed)
        query = {
            "$and": [
                {"is_archived": {"$ne": True}},
                {"end_date": {"$gte": current_date_str}},
                {
                    "$or": [
                        {"completed_by_coordinator": {"$exists": False}},
                        {"completed_by_coordinator": False},
                        {"completion_status": {"$exists": False}},
                        {"completion_status": "ongoing"}
                    ]
                },
                {
                    "$or": [
                        {"trainer_assignments.trainer_id": current_user.id},
                        {"assistant_coordinator_ids": current_user.id}
                    ]
                }
            ]
        }
    elif current_user.role == "coordinator":
        # Coordinator sees sessions they're assigned to as coordinator OR assistant coordinator
        query = {
            "$and": [
                {"is_archived": {"$ne": True}},
                {"status": "active"},
                {
                    "$or": [
                        {"completion_status": {"$exists": False}},
                        {"completion_status": "ongoing"},
                        {"completion_status": {"$nin": ["completed", "archived"]}}
                    ]
                },
                {
                    "$or": [
                        {"coordinator_id": current_user.id},
                        {"assistant_coordinator_ids": current_user.id}
                    ]
                }
            ]
        }
    elif current_user.role == "assistant_admin":
        # Assistant admin sees all non-completed sessions
        query = {
            "$and": [
                {"is_archived": {"$ne": True}},
                {
                    "$or": [
                        {"completion_status": {"$exists": False}},
                        {"completion_status": "ongoing"},
                        {"completion_status": {"$nin": ["completed", "archived"]}}
                    ]
                }
            ]
        }
    else:
        # Default: show all non-completed, non-archived sessions
        query = {
            "$and": [
                {"is_archived": {"$ne": True}},
                {"status": "active"},
                {
                    "$or": [
                        {"completion_status": {"$exists": False}},
                        {"completion_status": "ongoing"},
                        {"completion_status": {"$nin": ["completed", "archived"]}}
                    ]
                }
            ]
        }
    
    # Add search filters
    if company_id:
        query["$and"].append({"company_id": company_id})
    if program_id:
        query["$and"].append({"program_id": program_id})
    if start_date:
        query["$and"].append({"start_date": {"$gte": start_date}})
    if end_date:
        query["$and"].append({"end_date": {"$lte": end_date}})
    
    sessions = await db.sessions.find(query, {"_id": 0}).sort("start_date", -1).to_list(1000)
    
    # Enrich with company_name and program_name
    for session in sessions:
        if not session.get("company_name") and session.get("company_id"):
            company = await db.companies.find_one({"id": session["company_id"]}, {"_id": 0, "name": 1})
            if company:
                session["company_name"] = company.get("name")
        
        if not session.get("program_name") and session.get("program_id"):
            program = await db.programs.find_one({"id": session["program_id"]}, {"_id": 0, "name": 1})
            if program:
                session["program_name"] = program.get("name")
        
        # Convert datetime objects to strings for JSON serialization
        if isinstance(session.get("created_at"), datetime):
            session["created_at"] = session["created_at"].isoformat()
    
    return sessions


@router.post("")
async def create_session(session_data: SessionCreate, current_user: User = Depends(get_current_user)):
    """Create a new training session"""
    if current_user.role not in ["admin", "assistant_admin"]:
        raise HTTPException(status_code=403, detail="Only admins and assistant admins can create sessions")

    # Create or link participants
    participant_ids = []
    for p_data in session_data.participants:
        result = await find_or_create_user(
            p_data.model_dump(),
            "participant",
            session_data.company_id
        )
        participant_ids.append(result["user_id"])

    # Add existing participants
    participant_ids.extend(session_data.participant_ids)

    # Create or link supervisors
    supervisor_ids = []
    for s_data in session_data.supervisors:
        result = await find_or_create_user(
            s_data.model_dump(),
            "supervisor",
            session_data.company_id
        )
        supervisor_ids.append(result["user_id"])

    supervisor_ids.extend(session_data.supervisor_ids)

    session_obj = Session(
        name=session_data.name,
        program_id=session_data.program_id,
        company_id=session_data.company_id,
        location=session_data.location,
        start_date=session_data.start_date,
        end_date=session_data.end_date,
        supervisor_ids=supervisor_ids,
        participant_ids=participant_ids,
        trainer_assignments=session_data.trainer_assignments,
        coordinator_id=session_data.coordinator_id,
        cert_show_validity=session_data.cert_show_validity,
        cert_validity_months=session_data.cert_validity_months,
        schedule=session_data.schedule,
    )

    doc = session_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.sessions.insert_one(doc)

    # Create participant access records
    for participant_id in participant_ids:
        await get_or_create_participant_access(participant_id, session_obj.id)

    return {"message": "Session created", "session_id": session_obj.id}


@router.get("/calendar")
async def get_calendar_sessions(current_user: User = Depends(get_current_user)):
    """Get all sessions for calendar view — visible to ALL staff for availability planning"""
    # No role-based filtering — all staff see all sessions
    query = {}
    
    sessions = await db.sessions.find(query, {"_id": 0}).to_list(1000)
    
    # Enrich with names
    for session in sessions:
        if isinstance(session.get('created_at'), str):
            session['created_at'] = datetime.fromisoformat(session['created_at'])
        
        if session.get("company_id"):
            company = await db.companies.find_one({"id": session["company_id"]}, {"_id": 0})
            session["company_name"] = company.get("name", "Unknown") if company else "Unknown"
        
        if session.get("program_id"):
            program = await db.programs.find_one({"id": session["program_id"]}, {"_id": 0})
            session["program_name"] = program.get("name", "Unknown") if program else "Unknown"
    
    return sessions


@router.get("/my-marketing-sessions")
async def get_marketing_sessions(current_user: User = Depends(get_current_user)):
    """Get sessions brought in by a marketing user (via their quotations).
    Returns current (ongoing/draft) and past (completed) sessions separately."""
    if current_user.role not in ["marketing", "admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Marketing access required")

    # Find sessions linked to this marketer
    if current_user.role == "marketing":
        query = {"marketing_user_id": current_user.id}
    else:
        query = {"marketing_user_id": {"$exists": True, "$ne": None}}

    sessions = await db.sessions.find(query, {"_id": 0}).to_list(500)

    current_sessions = []
    past_sessions = []

    for session in sessions:
        # Enrich
        if session.get("company_id"):
            company = await db.companies.find_one({"id": session["company_id"]}, {"_id": 0})
            session["company_name"] = company.get("name", "Unknown") if company else "Unknown"
        if session.get("program_id"):
            program = await db.programs.find_one({"id": session["program_id"]}, {"_id": 0})
            session["program_name"] = program.get("name", "Unknown") if program else "Unknown"
        if session.get("coordinator_id"):
            coord = await db.users.find_one({"id": session["coordinator_id"]}, {"_id": 0, "full_name": 1})
            session["coordinator_name"] = coord.get("full_name", "Unassigned") if coord else "Unassigned"
        # Trainer names
        trainer_names = []
        for ta in session.get("trainer_assignments") or []:
            tid = ta.get("trainer_id") if isinstance(ta, dict) else ta
            if tid:
                trainer = await db.users.find_one({"id": tid}, {"_id": 0, "full_name": 1})
                if trainer:
                    trainer_names.append(trainer["full_name"])
        session["trainer_names"] = trainer_names
        session["participant_count"] = len(session.get("participant_ids", []))

        if session.get("completion_status") in ["completed", "archived"] or session.get("is_archived"):
            past_sessions.append(session)
        else:
            current_sessions.append(session)

    return {"current": current_sessions, "past": past_sessions}



@router.get("/past-training")
async def get_past_training(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Get completed/archived training sessions with optional month/year filter"""
    current_date = get_malaysia_time().date()
    current_date_str = current_date.isoformat()
    
    if current_user.role == "trainer":
        # For trainers: Show sessions where:
        # 1. Trainer is assigned, AND
        # 2. Either completed by coordinator OR end_date has passed
        query = {
            "$and": [
                # Trainer must be assigned
                {
                    "$or": [
                        {"trainer_assignments.trainer_id": current_user.id},
                        {"assistant_coordinator_ids": current_user.id}
                    ]
                },
                # Either completed OR end_date has passed
                {
                    "$or": [
                        {"completed_by_coordinator": True},
                        {"completion_status": "completed"},
                        {"completion_status": "archived"},
                        {"is_archived": True},
                        {"end_date": {"$lt": current_date_str}}  # Past sessions
                    ]
                }
            ]
        }
    elif current_user.role == "coordinator":
        query = {
            "$and": [
                {"coordinator_id": current_user.id},
                {
                    "$or": [
                        {"completion_status": "completed"},
                        {"completion_status": "archived"},
                        {"is_archived": True}
                    ]
                }
            ]
        }
    elif current_user.role == "participant":
        query = {
            "$and": [
                {"participant_ids": current_user.id},
                {
                    "$or": [
                        {"completion_status": "completed"},
                        {"completion_status": "archived"},
                        {"is_archived": True}
                    ]
                }
            ]
        }
    else:
        # Admin/assistant_admin see all completed sessions
        query = {
            "$or": [
                {"completion_status": "completed"},
                {"completion_status": "archived"},
                {"is_archived": True}
            ]
        }
    
    # Add month/year filter if provided
    if month and year:
        start_of_month = f"{year}-{month:02d}-01"
        if month == 12:
            end_of_month = f"{year+1}-01-01"
        else:
            end_of_month = f"{year}-{month+1:02d}-01"
        
        date_filter = {
            "end_date": {
                "$gte": start_of_month,
                "$lt": end_of_month
            }
        }
        
        if "$and" in query:
            query["$and"].append(date_filter)
        else:
            query = {"$and": [query, date_filter]}
    
    sessions = await db.sessions.find(query, {"_id": 0}).to_list(1000)
    
    for session in sessions:
        if isinstance(session.get('created_at'), str):
            session['created_at'] = datetime.fromisoformat(session['created_at'])
        
        if session.get("company_id"):
            company = await db.companies.find_one({"id": session["company_id"]}, {"_id": 0})
            session["company_name"] = company.get("name", "Unknown") if company else "Unknown"
        
        if session.get("program_id"):
            program = await db.programs.find_one({"id": session["program_id"]}, {"_id": 0})
            session["program_name"] = program.get("name", "Unknown") if program else "Unknown"
    
    return sessions


@router.get("/{session_id}", response_model=Session)
async def get_session(session_id: str, current_user: User = Depends(get_current_user)):
    """Get a specific session by ID"""
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if isinstance(session.get('created_at'), str):
        session['created_at'] = datetime.fromisoformat(session['created_at'])
    
    return Session(**session)


@router.get("/{session_id}/participants")
async def get_session_participants(session_id: str, current_user: User = Depends(get_current_user)):
    """Get all participants for a session"""
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    participant_ids = session.get("participant_ids", [])
    participants = []
    
    for pid in participant_ids:
        user = await db.users.find_one({"id": pid}, {"_id": 0, "password": 0})
        if user:
            if isinstance(user.get('created_at'), str):
                user['created_at'] = datetime.fromisoformat(user['created_at'])
            participants.append(user)
    
    return participants


@router.get("/{session_id}/participants/enriched")
async def get_enriched_participants(session_id: str, current_user: User = Depends(get_current_user)):
    """Get participants with all their related data (attendance, test results, etc.) - OPTIMIZED"""
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    participant_ids = session.get("participant_ids", [])
    if not participant_ids:
        return []
    
    # Bulk fetch all data
    users = await db.users.find({"id": {"$in": participant_ids}}, {"_id": 0, "password": 0}).to_list(1000)
    attendances = await db.attendance.find({"session_id": session_id, "participant_id": {"$in": participant_ids}}, {"_id": 0}).to_list(1000)
    accesses = await db.participant_access.find({"session_id": session_id, "participant_id": {"$in": participant_ids}}, {"_id": 0}).to_list(1000)
    test_results = await db.test_results.find({"session_id": session_id, "participant_id": {"$in": participant_ids}}, {"_id": 0}).to_list(2000)
    checklists = await db.checklist_submissions.find({"session_id": session_id, "participant_id": {"$in": participant_ids}}, {"_id": 0}).to_list(1000)
    feedbacks = await db.course_feedback.find({"session_id": session_id, "participant_id": {"$in": participant_ids}}, {"_id": 0}).to_list(1000)
    
    # Create lookup maps
    user_map = {u['id']: u for u in users}
    attendance_map = {a['participant_id']: a for a in attendances}
    access_map = {a['participant_id']: a for a in accesses}
    checklist_map = {c['participant_id']: c for c in checklists}
    feedback_map = {f['participant_id']: f for f in feedbacks}
    
    # Group test results by participant
    test_results_map = {}
    for tr in test_results:
        pid = tr['participant_id']
        if pid not in test_results_map:
            test_results_map[pid] = []
        test_results_map[pid].append(tr)
    
    # Build enriched data
    enriched = []
    for pid in participant_ids:
        user = user_map.get(pid, {})
        if not user:
            continue
        
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
        
        enriched.append({
            "user": user,
            "attendance": attendance_map.get(pid),
            "access": access_map.get(pid),
            "test_results": test_results_map.get(pid, []),
            "checklist": checklist_map.get(pid),
            "feedback": feedback_map.get(pid)
        })
    
    return enriched


@router.get("/{session_id}/status")
async def get_session_status(session_id: str, current_user: User = Depends(get_current_user)):
    """Get session completion status"""
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    participant_ids = session.get("participant_ids", [])
    total_participants = len(participant_ids)
    
    # Count completions
    attendance_complete = await db.attendance.count_documents({
        "session_id": session_id,
        "clock_in": {"$exists": True, "$ne": None}
    })
    
    pre_test_complete = await db.test_results.count_documents({
        "session_id": session_id,
        "test_type": {"$in": ["pre", "pre_test"]}
    })
    
    post_test_complete = await db.test_results.count_documents({
        "session_id": session_id,
        "test_type": {"$in": ["post", "post_test"]}
    })
    
    checklist_complete = await db.checklist_submissions.count_documents({
        "session_id": session_id
    })
    
    feedback_complete = await db.course_feedback.count_documents({
        "session_id": session_id
    })
    
    return {
        "session_id": session_id,
        "total_participants": total_participants,
        "attendance_complete": attendance_complete,
        "pre_test_complete": pre_test_complete,
        "post_test_complete": post_test_complete,
        "checklist_complete": checklist_complete,
        "feedback_complete": feedback_complete,
        "completion_status": session.get("completion_status", "ongoing"),
        "completed_by_coordinator": session.get("completed_by_coordinator", False)
    }


@router.get("/{session_id}/completion-checklist")
async def get_completion_checklist(session_id: str, current_user: User = Depends(get_current_user)):
    """Get completion checklist for session"""
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    participant_ids = session.get("participant_ids", [])
    
    checklist = {
        "all_attendance_recorded": False,
        "all_pre_tests_completed": False,
        "all_post_tests_completed": False,
        "all_checklists_submitted": False,
        "all_feedback_submitted": False,
        "coordinator_feedback_submitted": False,
        "chief_trainer_feedback_submitted": False
    }
    
    if participant_ids:
        total = len(participant_ids)
        
        attendance_count = await db.attendance.count_documents({
            "session_id": session_id,
            "clock_in": {"$exists": True, "$ne": None}
        })
        checklist["all_attendance_recorded"] = attendance_count >= total
        
        pre_test_count = await db.test_results.count_documents({
            "session_id": session_id,
            "test_type": {"$in": ["pre", "pre_test"]}
        })
        checklist["all_pre_tests_completed"] = pre_test_count >= total
        
        post_test_count = await db.test_results.count_documents({
            "session_id": session_id,
            "test_type": {"$in": ["post", "post_test"]}
        })
        checklist["all_post_tests_completed"] = post_test_count >= total
        
        checklist_count = await db.checklist_submissions.count_documents({
            "session_id": session_id
        })
        checklist["all_checklists_submitted"] = checklist_count >= total
        
        feedback_count = await db.course_feedback.count_documents({
            "session_id": session_id
        })
        checklist["all_feedback_submitted"] = feedback_count >= total
    
    coord_feedback = await db.coordinator_feedback.find_one({"session_id": session_id})
    checklist["coordinator_feedback_submitted"] = coord_feedback is not None
    
    trainer_feedback = await db.chief_trainer_feedback.find_one({"session_id": session_id})
    checklist["chief_trainer_feedback_submitted"] = trainer_feedback is not None
    
    return checklist


@router.get("/{session_id}/results-summary")
async def get_results_summary(session_id: str, current_user: User = Depends(get_current_user)):
    """Get summary of test results for a session"""
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    participant_ids = session.get("participant_ids", [])
    
    # Get all test results
    results = await db.test_results.find({"session_id": session_id}, {"_id": 0}).to_list(2000)
    
    # Calculate stats
    pre_scores = [r['score'] for r in results if r.get('test_type') in ['pre', 'pre_test']]
    post_scores = [r['score'] for r in results if r.get('test_type') in ['post', 'post_test']]
    
    return {
        "total_participants": len(participant_ids),
        "pre_test": {
            "completed": len(pre_scores),
            "average_score": sum(pre_scores) / len(pre_scores) if pre_scores else 0,
            "passed": sum(1 for r in results if r.get('test_type') in ['pre', 'pre_test'] and r.get('passed', False))
        },
        "post_test": {
            "completed": len(post_scores),
            "average_score": sum(post_scores) / len(post_scores) if post_scores else 0,
            "passed": sum(1 for r in results if r.get('test_type') in ['post', 'post_test'] and r.get('passed', False))
        }
    }


@router.post("/{session_id}/release-pre-test")
async def release_pre_test(session_id: str, current_user: User = Depends(get_current_user)):
    """Release pre-test for all participants"""
    if current_user.role not in ["coordinator", "admin", "trainer"]:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    for pid in session.get("participant_ids", []):
        await get_or_create_participant_access(pid, session_id)
        await db.participant_access.update_one(
            {"participant_id": pid, "session_id": session_id},
            {"$set": {"can_access_pre_test": True}}
        )
    
    return {"message": "Pre-test released for all participants"}


@router.post("/{session_id}/release-post-test")
async def release_post_test(session_id: str, current_user: User = Depends(get_current_user)):
    """Release post-test for all participants"""
    if current_user.role not in ["coordinator", "admin", "trainer"]:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    for pid in session.get("participant_ids", []):
        await get_or_create_participant_access(pid, session_id)
        await db.participant_access.update_one(
            {"participant_id": pid, "session_id": session_id},
            {"$set": {"can_access_post_test": True}}
        )
    
    return {"message": "Post-test released for all participants"}


@router.post("/{session_id}/release-feedback")
async def release_feedback(session_id: str, current_user: User = Depends(get_current_user)):
    """Release feedback for all participants"""
    if current_user.role not in ["coordinator", "admin", "trainer"]:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    for pid in session.get("participant_ids", []):
        await get_or_create_participant_access(pid, session_id)
        await db.participant_access.update_one(
            {"participant_id": pid, "session_id": session_id},
            {"$set": {"can_access_feedback": True}}
        )
    
    return {"message": "Feedback released for all participants"}


@router.get("/{session_id}/participants/attendance")
async def get_participants_attendance(session_id: str, current_user: User = Depends(get_current_user)):
    """Get attendance for all participants in a session"""
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    participant_ids = session.get("participant_ids", [])
    
    result = []
    for pid in participant_ids:
        user = await db.users.find_one({"id": pid}, {"_id": 0, "password": 0})
        attendance = await db.attendance.find_one({"participant_id": pid, "session_id": session_id}, {"_id": 0})
        
        if user:
            result.append({
                "participant": user,
                "attendance": attendance
            })
    
    return result


@router.get("/{session_id}/tests/available")
async def get_available_tests(session_id: str, current_user: User = Depends(get_current_user)):
    """Get available tests for a participant in a session"""
    if current_user.role != "participant":
        raise HTTPException(status_code=403, detail="Only participants can access this")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    access = await get_or_create_participant_access(current_user.id, session_id)
    tests = await db.tests.find({"program_id": session['program_id']}, {"_id": 0}).to_list(10)
    
    import random
    available_tests = []
    for test in tests:
        if isinstance(test.get('created_at'), str):
            test['created_at'] = datetime.fromisoformat(test['created_at'])
        
        test_type = test['test_type']
        can_access = False
        is_completed = False
        
        if test_type in ["pre", "pre_test"]:
            can_access = access.can_access_pre_test
            is_completed = access.pre_test_completed
        elif test_type in ["post", "post_test"]:
            can_access = access.can_access_post_test
            is_completed = access.post_test_completed
        
        if can_access and not is_completed:
            test_copy = test.copy()
            questions = test['questions'].copy()
            
            if test_type == "post":
                random.shuffle(questions)
            
            test_copy['questions'] = [
                {'question': q['question'], 'options': q['options']}
                for q in questions
            ]
            available_tests.append(test_copy)
    
    return available_tests


# ==================== STAGE 6b: Additional Session Endpoints ====================

@router.get("/{session_id}/indemnity-records")
async def get_session_indemnity_records(session_id: str, current_user: User = Depends(get_current_user)):
    """Get indemnity acceptance records for all participants in a session"""
    if current_user.role not in ["admin", "coordinator", "assistant_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    participant_ids = session.get("participant_ids", [])
    participants = await db.users.find(
        {"id": {"$in": participant_ids}},
        {
            "_id": 0,
            "id": 1,
            "full_name": 1,
            "id_number": 1,
            "email": 1,
            "phone_number": 1,
            "profile_verified": 1,
            "indemnity_accepted": 1,
            "indemnity_accepted_at": 1,
            "indemnity_signature": 1,
            "indemnity_signed_name": 1,
            "indemnity_signed_ic": 1,
            "indemnity_signed_date": 1,
            "emergency_contact_name": 1,
            "emergency_contact_relationship": 1,
            "emergency_contact_phone": 1
        }
    ).to_list(1000)
    
    return {
        "session_id": session_id,
        "session_name": session.get("name"),
        "company_name": session.get("company_name"),
        "training_date": f"{session.get('start_date')} to {session.get('end_date')}",
        "location": session.get("location"),
        "total_participants": len(participants),
        "indemnity_records": participants
    }


@router.get("/{session_id}/indemnity-records/export")
async def export_session_indemnity_records(session_id: str, current_user: User = Depends(get_current_user)):
    """Export indemnity records as Excel file"""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    
    if current_user.role not in ["admin", "coordinator", "assistant_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    participant_ids = session.get("participant_ids", [])
    participants = await db.users.find({"id": {"$in": participant_ids}}, {"_id": 0}).to_list(1000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indemnity Records"
    
    headers = ["No", "Full Name", "IC Number", "Indemnity Accepted", "Signed Name", "Signed IC", "Signed Date", "Accepted At"]
    ws.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    for idx, p in enumerate(participants, 1):
        ws.append([
            idx,
            p.get("full_name", ""),
            p.get("id_number", ""),
            "Yes" if p.get("indemnity_accepted") else "No",
            p.get("indemnity_signed_name", ""),
            p.get("indemnity_signed_ic", ""),
            p.get("indemnity_signed_date", ""),
            p.get("indemnity_accepted_at", "")
        ])
    
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 40)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    session_name = session.get("name", "Session").replace(" ", "_")
    filename = f"Indemnity_Records_{session_name}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.put("/{session_id}/toggle-status")
async def toggle_session_status(session_id: str, current_user: User = Depends(get_current_user)):
    """Toggle session between active and inactive (Admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can change session status")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    new_status = "inactive" if session.get("status", "active") == "active" else "active"
    
    await db.sessions.update_one(
        {"id": session_id},
        {"$set": {"status": new_status}}
    )
    
    return {"message": f"Session marked as {new_status}", "status": new_status}


@router.post("/{session_id}/participants")
async def add_participants_to_session(
    session_id: str,
    participant_ids: dict,
    current_user: User = Depends(get_current_user)
):
    """Add participants to a session by IC number or user ID"""
    if current_user.role not in ["admin", "assistant_admin"]:
        raise HTTPException(status_code=403, detail="Only admins and assistant admins can add participants")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    ids_to_add = participant_ids.get("participant_ids", [])
    if not ids_to_add:
        raise HTTPException(status_code=400, detail="No participant IDs provided")
    
    added_ids = []
    for identifier in ids_to_add:
        user = await db.users.find_one(
            {"$or": [{"id_number": identifier}, {"id": identifier}]},
            {"_id": 0, "id": 1}
        )
        if user:
            added_ids.append(user["id"])
        else:
            raise HTTPException(status_code=404, detail=f"User not found: {identifier}")
    
    current_participants = session.get("participant_ids", [])
    newly_added = []
    for user_id in added_ids:
        if user_id not in current_participants:
            current_participants.append(user_id)
            newly_added.append(user_id)
    
    await db.sessions.update_one(
        {"id": session_id},
        {"$set": {"participant_ids": current_participants}}
    )
    
    for user_id in newly_added:
        await get_or_create_participant_access(user_id, session_id)
    
    return {
        "message": f"Successfully added {len(added_ids)} participant(s)",
        "added_count": len(added_ids)
    }


@router.put("/{session_id}")
async def update_session(session_id: str, session_data: dict, current_user: User = Depends(get_current_user)):
    """Update session details with cascade to invoices"""
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if current_user.role == "coordinator":
        if session.get("coordinator_id") != current_user.id:
            raise HTTPException(status_code=403, detail="You can only update sessions assigned to you")
    elif current_user.role not in ["admin", "assistant_admin"]:
        raise HTTPException(status_code=403, detail="Only admins and coordinators can update sessions")
    
    # Capture old values before update
    old_start_date = session.get("start_date")
    old_end_date = session.get("end_date")
    old_company_name = session.get("company_name")
    old_location = session.get("location")
    old_program_id = session.get("program_id")
    
    # Check if participant_ids changed (new participants added)
    old_participant_ids = set(session.get("participant_ids", []))
    new_participant_ids = set(session_data.get("participant_ids", []))
    
    result = await db.sessions.update_one(
        {"id": session_id},
        {"$set": session_data}
    )
    
    # Cascade company_name update to related invoices and quotations
    new_company_name = session_data.get("company_name")
    if new_company_name and new_company_name != old_company_name:
        await db.invoices.update_many(
            {"session_id": session_id},
            {"$set": {"company_name": new_company_name, "bill_to_name": new_company_name}}
        )
        if session.get("quotation_id"):
            await db.quotations.update_one(
                {"id": session.get("quotation_id")},
                {"$set": {"client_name": new_company_name, "company_name": new_company_name}}
            )
        if session.get("lead_id"):
            await db.leads.update_one(
                {"id": session.get("lead_id")},
                {"$set": {"company_name": new_company_name}}
            )
    
    # Cascade date changes to related invoices
    new_start_date = session_data.get("start_date")
    new_end_date = session_data.get("end_date")
    if (new_start_date and new_start_date != old_start_date) or (new_end_date and new_end_date != old_end_date):
        final_start = new_start_date or old_start_date
        final_end = new_end_date or old_end_date
        if final_start and final_end:
            new_training_dates = f"{final_start} to {final_end}"
            await db.invoices.update_many(
                {"session_id": session_id},
                {"$set": {"training_dates": new_training_dates}}
            )
    
    # Cascade venue/location changes to related invoices
    new_location = session_data.get("location")
    if new_location and new_location != old_location:
        await db.invoices.update_many(
            {"session_id": session_id},
            {"$set": {"venue": new_location}}
        )
    
    # Cascade programme changes to related invoices
    new_program_id = session_data.get("program_id")
    if new_program_id and new_program_id != old_program_id:
        programme = await db.programs.find_one({"id": new_program_id}, {"_id": 0})
        if programme:
            await db.invoices.update_many(
                {"session_id": session_id},
                {"$set": {"programme_name": programme.get("name")}}
            )
    
    updated_session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    return updated_session


@router.delete("/{session_id}")
async def delete_session(session_id: str, current_user: User = Depends(get_current_user)):
    """Delete session and all related data, saving auto-draft invoice numbers for reuse"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete sessions")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Check for auto-draft invoices and save their numbers for reuse
    invoices = await db.invoices.find({"session_id": session_id}, {"_id": 0}).to_list(100)
    saved_invoice_numbers = []
    
    for invoice in invoices:
        # Only save numbers from auto-draft/draft invoices (never issued)
        if invoice.get("status") in ["auto_draft", "draft"]:
            invoice_number = invoice.get("invoice_number")
            if invoice_number:
                # Save to deleted_invoice_numbers collection for reuse
                await db.deleted_invoice_numbers.insert_one({
                    "invoice_number": invoice_number,
                    "original_session_id": session_id,
                    "original_session_name": session.get("name"),
                    "original_company_id": session.get("company_id"),
                    "deleted_at": get_malaysia_time().isoformat(),
                    "deleted_by": current_user.id,
                    "is_available": True
                })
                saved_invoice_numbers.append(invoice_number)
    
    total_deleted = 0
    
    result = await db.sessions.delete_one({"id": session_id})
    total_deleted += result.deleted_count
    
    related_collections = [
        "test_results", "course_feedback", "attendance", "attendance_records",
        "participant_attendance", "vehicle_checklists", "vehicle_details",
        "certificates", "participant_access", "training_reports",
        "chief_trainer_feedback", "coordinator_feedback",
        "trainer_fees", "coordinator_fees", "session_expenses",
        "invoices", "credit_notes", "marketing_commissions",
    ]
    
    for collection_name in related_collections:
        result = await db[collection_name].delete_many({"session_id": session_id})
        total_deleted += result.deleted_count
    
    return {
        "message": "Session and all related data deleted successfully",
        "session_name": session.get("name"),
        "records_deleted": total_deleted,
        "invoice_numbers_saved_for_reuse": saved_invoice_numbers
    }


@router.delete("/bulk/delete-all")
async def delete_all_sessions(current_user: User = Depends(get_current_user)):
    """Delete ALL sessions and related data - for testing/cleanup purposes"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete all sessions")
    
    all_sessions = await db.sessions.find({}, {"_id": 0, "id": 1}).to_list(1000)
    session_ids = [s["id"] for s in all_sessions]
    
    if not session_ids:
        return {"message": "No sessions to delete", "sessions_deleted": 0, "total_records_deleted": 0}
    
    total_deleted = 0
    
    collections_to_clean = [
        "sessions", "test_results", "course_feedback", "attendance",
        "attendance_records", "participant_attendance", "vehicle_checklists",
        "vehicle_details", "certificates", "participant_access",
        "training_reports", "chief_trainer_feedback", "coordinator_feedback",
    ]
    
    for collection_name in collections_to_clean:
        result = await db[collection_name].delete_many({})
        total_deleted += result.deleted_count
    
    return {
        "message": "All sessions and related data deleted successfully",
        "sessions_deleted": len(session_ids),
        "total_records_deleted": total_deleted
    }


@router.post("/{session_id}/participants/{participant_id}/attendance")
async def mark_participant_attendance(
    session_id: str,
    participant_id: str,
    status: str,
    current_user: User = Depends(get_current_user)
):
    """Mark participant as present or absent for a session"""
    if current_user.role not in ["coordinator", "admin"]:
        raise HTTPException(status_code=403, detail="Only coordinators and admins can mark attendance")
    
    if status not in ["present", "absent"]:
        raise HTTPException(status_code=400, detail="Status must be 'present' or 'absent'")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if participant_id not in session.get("participant_ids", []):
        raise HTTPException(status_code=400, detail="Participant not enrolled in this session")
    
    await db.participant_attendance.update_one(
        {"session_id": session_id, "participant_id": participant_id},
        {"$set": {"status": status, "marked_by": current_user.id, "marked_at": get_malaysia_time().isoformat()}},
        upsert=True
    )
    
    return {"message": f"Participant marked as {status}", "status": status}


@router.post("/{session_id}/mark-completed")
async def mark_session_completed(session_id: str, current_user: User = Depends(get_current_user)):
    """Mark session as completed by coordinator - archives session and pushes report to supervisors"""
    if current_user.role not in ["coordinator", "admin"]:
        raise HTTPException(status_code=403, detail="Only coordinators and admins can mark sessions as completed")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    training_report = await db.training_reports.find_one({"session_id": session_id}, {"_id": 0})
    
    if not training_report or not training_report.get("final_pdf_filename"):
        raise HTTPException(
            status_code=400,
            detail="Training report must be uploaded before marking session as completed. Please upload the final PDF report first."
        )
    
    await db.sessions.update_one(
        {"id": session_id},
        {"$set": {
            "completion_status": "completed",
            "completed_by_coordinator": True,
            "completed_date": get_malaysia_time().isoformat(),
            "report_available_to_supervisors": True
        }}
    )
    
    await db.training_reports.update_one(
        {"session_id": session_id},
        {"$set": {"available_to_supervisors": True, "pushed_to_supervisors_at": get_malaysia_time().isoformat()}}
    )
    
    # Notify admin, coordinator, trainers
    try:
        updated_session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
        await notify_session_completed(updated_session)
    except:
        pass
    
    return {
        "message": "Session marked as completed successfully. Report is now available to supervisors.",
        "session_archived": True,
        "report_pushed_to_supervisors": True
    }


# ============================================================
# Admin Session Complete + Excel Import/Export
# ============================================================

@router.post("/{session_id}/admin-complete")
async def admin_mark_session_complete(session_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """Admin marks session as completed (bypasses coordinator workflow)"""
    if current_user.role not in ["admin", "super_admin", "assistant_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can force-complete sessions")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    reason = data.get("reason", "Admin override")
    
    await db.sessions.update_one(
        {"id": session_id},
        {"$set": {
            "completion_status": "completed",
            "completed_by_coordinator": True,
            "completed_date": get_malaysia_time().isoformat(),
            "admin_completed": True,
            "admin_complete_reason": reason,
            "admin_completed_by": current_user.id,
            "is_archived": True
        }}
    )
    
    # Notify coordinator & trainers
    try:
        updated_session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
        await notify_session_completed(updated_session)
    except:
        pass
    
    return {
        "message": f"Session marked as completed by admin",
        "session_id": session_id,
        "reason": reason
    }


@router.post("/{session_id}/admin-revert-complete")
async def admin_revert_session_complete(session_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """Admin reverts a completed session back to ongoing"""
    if current_user.role not in ["admin", "super_admin", "assistant_admin"]:
        raise HTTPException(status_code=403, detail="Only admins can revert session completion")
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    reason = data.get("reason", "Admin revert")
    
    await db.sessions.update_one(
        {"id": session_id},
        {"$set": {
            "completion_status": "ongoing",
            "completed_by_coordinator": False,
            "is_archived": False,
            "revert_reason": reason,
            "reverted_by": current_user.id,
            "reverted_at": get_malaysia_time().isoformat()
        },
        "$unset": {
            "completed_date": "",
            "admin_completed": "",
            "admin_complete_reason": "",
            "admin_completed_by": ""
        }}
    )
    
    return {"message": "Session reverted to ongoing", "session_id": session_id}


@router.get("/{session_id}/export-template")
async def export_session_template(session_id: str, current_user: User = Depends(get_current_user)):
    """Export Excel template pre-populated with session participants"""
    if current_user.role not in ["admin", "super_admin", "assistant_admin", "coordinator", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get program for pass_percentage
    program = None
    if session.get("program_id"):
        program = await db.programs.find_one({"id": session["program_id"]}, {"_id": 0})
    pass_pct = program.get("pass_percentage", 70.0) if program else 70.0
    
    # Get participants (use correct field names: full_name, id_number)
    participant_ids = session.get("participant_ids", [])
    participants = []
    for pid in participant_ids:
        user = await db.users.find_one({"id": pid}, {"_id": 0, "id": 1, "full_name": 1, "id_number": 1, "email": 1, "phone_number": 1})
        if user:
            participants.append(user)
    
    # Get existing data
    existing_attendance = await db.attendance.find({"session_id": session_id}, {"_id": 0}).to_list(1000)
    existing_tests = await db.test_results.find({"session_id": session_id}, {"_id": 0}).to_list(1000)
    existing_checklists = await db.vehicle_checklists.find({"session_id": session_id}, {"_id": 0}).to_list(1000)
    existing_vehicle_details = await db.vehicle_details.find({"session_id": session_id}, {"_id": 0}).to_list(1000)
    
    att_map = {}
    for a in existing_attendance:
        key = a.get("participant_id")
        if key not in att_map:
            att_map[key] = []
        att_map[key].append(a)
    
    test_map = {}
    for t in existing_tests:
        key = (t.get("participant_id"), t.get("test_type"))
        test_map[key] = t
    
    checklist_map = {}
    for c in existing_checklists:
        checklist_map[c.get("participant_id")] = c
    
    vehicle_map = {}
    for v in existing_vehicle_details:
        vehicle_map[v.get("participant_id")] = v
    
    start = session.get("start_date", "")
    end = session.get("end_date", start)
    
    wb = Workbook()
    
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    info_font = Font(size=11, italic=True)
    
    # === Sheet 1: Pre/Post Tests (Raw Marks) ===
    ws1 = wb.active
    ws1.title = "Pre-Post Tests"
    
    ws1.merge_cells('A1:H1')
    ws1['A1'] = f"Session: {session.get('company_name', 'N/A')} - {session.get('program_name', 'N/A')}"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A2:H2')
    ws1['A2'] = f"Dates: {start} to {end} | Pass Mark: {pass_pct}% | Session ID: {session_id}"
    ws1['A2'].font = info_font
    
    test_headers = [
        "No", "Participant Name", "IC Number",
        "Pre-Test (Marks)", "Pre-Test (Total)",
        "Post-Test (Marks)", "Post-Test (Total)",
        "Remarks"
    ]
    for col, h in enumerate(test_headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    for i, p in enumerate(participants, 1):
        row = i + 4
        ws1.cell(row=row, column=1, value=i).border = thin_border
        ws1.cell(row=row, column=2, value=p.get("full_name", "")).border = thin_border
        ws1.cell(row=row, column=3, value=p.get("id_number", "")).border = thin_border
        
        pre = test_map.get((p["id"], "pre")) or test_map.get((p["id"], "pre_test"))
        post = test_map.get((p["id"], "post")) or test_map.get((p["id"], "post_test"))
        
        # Pre-existing marks (if imported before with raw marks, show them; otherwise show score as marks with total=100)
        if pre:
            ws1.cell(row=row, column=4, value=pre.get("marks_obtained", pre.get("score"))).border = thin_border
            ws1.cell(row=row, column=5, value=pre.get("total_marks", 100)).border = thin_border
        else:
            ws1.cell(row=row, column=4, value=None).border = thin_border
            ws1.cell(row=row, column=5, value=None).border = thin_border
        
        if post:
            ws1.cell(row=row, column=6, value=post.get("marks_obtained", post.get("score"))).border = thin_border
            ws1.cell(row=row, column=7, value=post.get("total_marks", 100)).border = thin_border
        else:
            ws1.cell(row=row, column=6, value=None).border = thin_border
            ws1.cell(row=row, column=7, value=None).border = thin_border
        
        ws1.cell(row=row, column=8, value="").border = thin_border
    
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 18
    ws1.column_dimensions['G'].width = 16
    ws1.column_dimensions['H'].width = 25
    
    # === Sheet 2: Attendance ===
    ws2 = wb.create_sheet("Attendance")
    ws2.merge_cells('A1:F1')
    ws2['A1'] = f"Attendance Record - {session.get('company_name', 'N/A')}"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2.merge_cells('A2:F2')
    ws2['A2'] = "Enter 'P' for Present, 'A' for Absent, 'L' for Late"
    ws2['A2'].font = Font(size=11, italic=True, color="666666")
    
    att_headers = ["No", "Participant Name", "IC Number"]
    from datetime import date as dt_date, timedelta
    day_dates = []
    if start:
        try:
            s_date = dt_date.fromisoformat(start)
            e_date = dt_date.fromisoformat(end) if end else s_date
            current = s_date
            while current <= e_date:
                day_dates.append(current.isoformat())
                att_headers.append(f"Day {len(day_dates)} ({current.strftime('%d/%m')})")
                current += timedelta(days=1)
        except Exception:
            att_headers.append("Day 1")
            day_dates.append(start)
    
    for col, h in enumerate(att_headers, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for i, p in enumerate(participants, 1):
        row = i + 4
        ws2.cell(row=row, column=1, value=i).border = thin_border
        ws2.cell(row=row, column=2, value=p.get("full_name", "")).border = thin_border
        ws2.cell(row=row, column=3, value=p.get("id_number", "")).border = thin_border
        
        p_att = att_map.get(p["id"], [])
        for d_idx, d_date in enumerate(day_dates):
            day_att = next((a for a in p_att if a.get("date") == d_date), None)
            val = "P" if day_att and day_att.get("clock_in") else ""
            ws2.cell(row=row, column=4 + d_idx, value=val).border = thin_border
    
    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 18
    for d_idx in range(len(day_dates)):
        col_letter = get_column_letter(4 + d_idx)
        ws2.column_dimensions[col_letter].width = 16
    
    # === Sheet 3: Vehicle Checklist ===
    ws3 = wb.create_sheet("Vehicle Checklist")
    ws3.merge_cells('A1:J1')
    ws3['A1'] = f"Vehicle Checklist - {session.get('company_name', 'N/A')}"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3.merge_cells('A2:J2')
    ws3['A2'] = "Enter vehicle details and checklist items. Status: Good / Needs Repair / Satisfactory / N/A"
    ws3['A2'].font = Font(size=11, italic=True, color="666666")
    
    # Determine checklist items from existing data, or use defaults
    all_checklist_item_names = set()
    for c in existing_checklists:
        for item in c.get("checklist_items", []):
            all_checklist_item_names.add(item.get("item", "").lower().strip())
    
    checklist_items_list = sorted(all_checklist_item_names) if all_checklist_item_names else ["helmet", "tires", "safety vest", "lights", "side mirror"]
    
    vc_headers = ["No", "Participant Name", "IC Number", "Vehicle Model", "Registration No", "Road Tax Expiry"]
    for ci_name in checklist_items_list:
        vc_headers.append(ci_name.title())
    vc_headers.append("Remarks")
    remarks_col = len(vc_headers)
    
    for col, h in enumerate(vc_headers, 1):
        cell = ws3.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    for i, p in enumerate(participants, 1):
        row = i + 4
        ws3.cell(row=row, column=1, value=i).border = thin_border
        ws3.cell(row=row, column=2, value=p.get("full_name", "")).border = thin_border
        ws3.cell(row=row, column=3, value=p.get("id_number", "")).border = thin_border
        
        # Pre-fill vehicle details if they exist
        vd = vehicle_map.get(p["id"])
        ws3.cell(row=row, column=4, value=vd.get("vehicle_model", "") if vd else "").border = thin_border
        ws3.cell(row=row, column=5, value=vd.get("registration_number", "") if vd else "").border = thin_border
        ws3.cell(row=row, column=6, value=vd.get("roadtax_expiry", "") if vd else "").border = thin_border
        
        # Pre-fill checklist items if they exist
        vc = checklist_map.get(p["id"])
        vc_items = {}
        if vc:
            for item in vc.get("checklist_items", []):
                vc_items[item.get("item", "").lower()] = item.get("status", "")
        
        for ci_idx, ci_name in enumerate(checklist_items_list):
            status = vc_items.get(ci_name, "")
            ws3.cell(row=row, column=7 + ci_idx, value=status).border = thin_border
        
        # Remarks column (collect comments from checklist items)
        remarks_text = ""
        if vc:
            comments = [f"{item['item']}: {item['comments']}" for item in vc.get("checklist_items", []) if item.get("comments")]
            remarks_text = "; ".join(comments)
        ws3.cell(row=row, column=remarks_col, value=remarks_text).border = thin_border
    
    ws3.column_dimensions['A'].width = 6
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 18
    ws3.column_dimensions['F'].width = 16
    for ci_idx in range(len(checklist_items_list)):
        col_letter = get_column_letter(7 + ci_idx)
        ws3.column_dimensions[col_letter].width = 16
    ws3.column_dimensions[get_column_letter(remarks_col)].width = 30
    
    # === Sheet 4: Feedback ===
    ws4 = wb.create_sheet("Feedback")
    ws4.merge_cells('A1:J1')
    ws4['A1'] = f"Participant Feedback - {session.get('company_name', 'N/A')}"
    ws4['A1'].font = Font(bold=True, size=14)
    ws4.merge_cells('A2:J2')
    ws4['A2'] = "Rating questions: Enter 1-5 (1=Poor, 5=Excellent). Text questions: Enter free text."
    ws4['A2'].font = Font(size=11, italic=True, color="666666")
    
    # Determine feedback questions:
    # 1. From feedback_questions collection (admin-configured)
    # 2. From existing feedback for this session
    # 3. Default set
    feedback_questions_list = []
    configured_qs = await db.feedback_questions.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    if configured_qs:
        feedback_questions_list = [{"question": q.get("question_text", q.get("question", "")), "type": q.get("question_type", "rating")} for q in configured_qs]
    else:
        # Discover from existing feedback for this session
        existing_feedback = await db.course_feedback.find({"session_id": session_id}, {"_id": 0}).to_list(1000)
        seen_questions = []
        for fb in existing_feedback:
            for r in fb.get("responses", []):
                q = r.get("question", "")
                if q and q not in seen_questions:
                    seen_questions.append(q)
                    q_type = "rating" if isinstance(r.get("answer"), (int, float)) else "text"
                    feedback_questions_list.append({"question": q, "type": q_type})
    
    if not feedback_questions_list:
        feedback_questions_list = [
            {"question": "Overall Training Experience", "type": "rating"},
            {"question": "Training Content Quality", "type": "rating"},
            {"question": "Trainer Effectiveness", "type": "rating"},
            {"question": "Venue & Facilities", "type": "rating"},
            {"question": "Suggestions for Improvement", "type": "text"},
            {"question": "Additional Comments", "type": "text"},
        ]
    
    # Build feedback lookup
    existing_feedback = await db.course_feedback.find({"session_id": session_id}, {"_id": 0}).to_list(1000)
    feedback_map = {}
    for fb in existing_feedback:
        pid = fb.get("participant_id")
        resp_map = {}
        for r in fb.get("responses", []):
            resp_map[r.get("question", "")] = r.get("answer", "")
        feedback_map[pid] = resp_map
    
    # Headers
    fb_headers = ["No", "Participant Name", "IC Number"]
    for fq in feedback_questions_list:
        label = fq["question"]
        if fq["type"] == "rating":
            label += " (1-5)"
        fb_headers.append(label)
    
    for col, h in enumerate(fb_headers, 1):
        cell = ws4.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    for i, p in enumerate(participants, 1):
        row = i + 4
        ws4.cell(row=row, column=1, value=i).border = thin_border
        ws4.cell(row=row, column=2, value=p.get("full_name", "")).border = thin_border
        ws4.cell(row=row, column=3, value=p.get("id_number", "")).border = thin_border
        
        p_feedback = feedback_map.get(p["id"], {})
        for fq_idx, fq in enumerate(feedback_questions_list):
            answer = p_feedback.get(fq["question"], "")
            ws4.cell(row=row, column=4 + fq_idx, value=answer if answer != "" else None).border = thin_border
    
    ws4.column_dimensions['A'].width = 6
    ws4.column_dimensions['B'].width = 30
    ws4.column_dimensions['C'].width = 18
    for fq_idx in range(len(feedback_questions_list)):
        col_letter = get_column_letter(4 + fq_idx)
        ws4.column_dimensions[col_letter].width = 25
    
    # === Sheet 5: Instructions ===
    ws5 = wb.create_sheet("Instructions")
    instructions = [
        ("MDDRC Session Data Import Template", Font(bold=True, size=16)),
        ("", None),
        ("Sheet 1: Pre/Post Tests", Font(bold=True, size=12)),
        (f"- Enter raw marks: 'Marks Obtained' and 'Total Marks' for each test", None),
        (f"- System will auto-calculate percentage and pass/fail (pass mark: {pass_pct}%)", None),
        ("- Do NOT modify the IC Number column - it's used for matching", None),
        ("- Leave blank if no score available", None),
        ("", None),
        ("Sheet 2: Attendance", Font(bold=True, size=12)),
        ("- Enter 'P' for Present, 'A' for Absent, 'L' for Late", None),
        ("- Each Day column corresponds to the session date shown in the header", None),
        ("- Leave blank if no data", None),
        ("", None),
        ("Sheet 3: Vehicle Checklist", Font(bold=True, size=12)),
        ("- Enter vehicle details: Model, Registration No, Road Tax Expiry (YYYY-MM-DD)", None),
        ("- For each checklist item, enter: 'good', 'needs_repair', 'satisfactory', or 'n/a'", None),
        ("- Use the Remarks column for any additional notes per participant", None),
        ("- Leave blank if no data", None),
        ("", None),
        ("Sheet 4: Feedback", Font(bold=True, size=12)),
        ("- For rating questions (marked 1-5): Enter a number from 1 to 5", None),
        ("  1 = Poor, 2 = Below Average, 3 = Average, 4 = Good, 5 = Excellent", None),
        ("- For text questions: Enter free text responses", None),
        ("- Leave blank if no feedback available", None),
        ("", None),
        ("IMPORTANT:", Font(bold=True, size=12, color="FF0000")),
        ("- Do NOT add/remove rows or change the order of participants", None),
        ("- Do NOT modify IC Numbers - they are used to match participants", None),
        ("- Save as .xlsx format before uploading", None),
    ]
    for i, (text, font) in enumerate(instructions, 1):
        cell = ws5.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
    ws5.column_dimensions['A'].width = 70
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    company = (session.get("company_name") or "session").replace(" ", "_")[:30]
    filename = f"MDDRC_Template_{company}_{start}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/{session_id}/import-data")
async def import_session_data(session_id: str, file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Import session data (test scores, attendance, vehicle checklist) from Excel"""
    if current_user.role not in ["admin", "super_admin", "assistant_admin", "coordinator"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from openpyxl import load_workbook
    from io import BytesIO
    
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get program for pass_percentage
    program = None
    if session.get("program_id"):
        program = await db.programs.find_one({"id": session["program_id"]}, {"_id": 0})
    pass_pct = program.get("pass_percentage", 70.0) if program else 70.0
    
    content = await file.read()
    wb = load_workbook(BytesIO(content), read_only=True)
    
    results = {"test_scores_imported": 0, "attendance_imported": 0, "vehicle_checklists_imported": 0, "feedback_imported": 0, "errors": [], "skipped": []}
    
    # Build participant lookup by IC number (use id_number field)
    participant_ids = session.get("participant_ids", [])
    ic_to_participant = {}
    for pid in participant_ids:
        user = await db.users.find_one({"id": pid}, {"_id": 0, "id": 1, "full_name": 1, "id_number": 1})
        if user and user.get("id_number"):
            ic_to_participant[str(user["id_number"]).strip()] = user
    
    # === Process Pre/Post Tests (Sheet 1) — Raw Marks ===
    test_sheet_name = None
    for name in ["Pre-Post Tests", "Test Scores"]:
        if name in wb.sheetnames:
            test_sheet_name = name
            break
    
    if test_sheet_name:
        ws = wb[test_sheet_name]
        for row in ws.iter_rows(min_row=5, values_only=False):
            try:
                ic = str(row[2].value or "").strip()
                if not ic or ic not in ic_to_participant:
                    continue
                
                participant = ic_to_participant[ic]
                pid = participant["id"]
                
                # Determine column layout based on sheet name
                if test_sheet_name == "Pre-Post Tests":
                    # New format: col 3=Pre Marks, col 4=Pre Total, col 5=Post Marks, col 6=Post Total
                    pre_marks = row[3].value
                    pre_total = row[4].value
                    post_marks = row[5].value
                    post_total = row[6].value
                else:
                    # Legacy format: col 3=Pre Score (%), col 4=Post Score (%)
                    pre_marks = row[3].value
                    pre_total = 100
                    post_marks = row[4].value
                    post_total = 100
                
                # Import pre-test score from raw marks
                if pre_marks is not None and str(pre_marks).strip() != "":
                    marks = float(pre_marks)
                    total = float(pre_total) if pre_total and str(pre_total).strip() != "" else 100
                    if total <= 0:
                        total = 100
                    score_pct = round((marks / total) * 100, 2)
                    passed = score_pct >= pass_pct
                    
                    test_data = {
                        "score": score_pct,
                        "marks_obtained": marks,
                        "total_marks": total,
                        "passed": passed,
                        "pass_percentage_used": pass_pct,
                        "imported": True,
                        "imported_at": get_malaysia_time().isoformat()
                    }
                    
                    existing = await db.test_results.find_one({"session_id": session_id, "participant_id": pid, "test_type": "pre"})
                    if existing:
                        await db.test_results.update_one(
                            {"session_id": session_id, "participant_id": pid, "test_type": "pre"},
                            {"$set": test_data}
                        )
                    else:
                        await db.test_results.insert_one({
                            "id": str(uuid.uuid4()),
                            "test_id": f"import-pre-{session_id}",
                            "participant_id": pid,
                            "session_id": session_id,
                            "test_type": "pre",
                            "answers": [],
                            "total_questions": 0,
                            "correct_answers": 0,
                            "submitted_at": get_malaysia_time().isoformat(),
                            **test_data
                        })
                    results["test_scores_imported"] += 1
                
                # Import post-test score from raw marks
                if post_marks is not None and str(post_marks).strip() != "":
                    marks = float(post_marks)
                    total = float(post_total) if post_total and str(post_total).strip() != "" else 100
                    if total <= 0:
                        total = 100
                    score_pct = round((marks / total) * 100, 2)
                    passed = score_pct >= pass_pct
                    
                    test_data = {
                        "score": score_pct,
                        "marks_obtained": marks,
                        "total_marks": total,
                        "passed": passed,
                        "pass_percentage_used": pass_pct,
                        "imported": True,
                        "imported_at": get_malaysia_time().isoformat()
                    }
                    
                    existing = await db.test_results.find_one({"session_id": session_id, "participant_id": pid, "test_type": "post"})
                    if existing:
                        await db.test_results.update_one(
                            {"session_id": session_id, "participant_id": pid, "test_type": "post"},
                            {"$set": test_data}
                        )
                    else:
                        await db.test_results.insert_one({
                            "id": str(uuid.uuid4()),
                            "test_id": f"import-post-{session_id}",
                            "participant_id": pid,
                            "session_id": session_id,
                            "test_type": "post",
                            "answers": [],
                            "total_questions": 0,
                            "correct_answers": 0,
                            "submitted_at": get_malaysia_time().isoformat(),
                            **test_data
                        })
                    results["test_scores_imported"] += 1
                    
            except Exception as e:
                results["errors"].append(f"Test row error: {str(e)}")
    
    # === Process Attendance (Sheet 2) ===
    if "Attendance" in wb.sheetnames:
        ws = wb["Attendance"]
        
        day_dates = []
        from datetime import date as dt_date, timedelta
        start = session.get("start_date", "")
        end = session.get("end_date", start)
        if start:
            try:
                s_date = dt_date.fromisoformat(start)
                e_date = dt_date.fromisoformat(end) if end else s_date
                current = s_date
                while current <= e_date:
                    day_dates.append(current.isoformat())
                    current += timedelta(days=1)
            except Exception:
                day_dates = [start]
        
        for row in ws.iter_rows(min_row=5, values_only=False):
            try:
                ic = str(row[2].value or "").strip()
                if not ic or ic not in ic_to_participant:
                    continue
                
                participant = ic_to_participant[ic]
                pid = participant["id"]
                
                for d_idx, d_date in enumerate(day_dates):
                    col_idx = 3 + d_idx
                    if col_idx >= len(row):
                        break
                    val = str(row[col_idx].value or "").strip().upper()
                    
                    if val in ["P", "PRESENT", "L", "LATE"]:
                        existing = await db.attendance.find_one({"session_id": session_id, "participant_id": pid, "date": d_date})
                        clock_in = "09:00:00" if val != "L" else "09:30:00"
                        if existing:
                            await db.attendance.update_one(
                                {"session_id": session_id, "participant_id": pid, "date": d_date},
                                {"$set": {"clock_in": clock_in, "imported": True}}
                            )
                        else:
                            await db.attendance.insert_one({
                                "id": str(uuid.uuid4()),
                                "participant_id": pid,
                                "session_id": session_id,
                                "date": d_date,
                                "clock_in": clock_in,
                                "clock_out": None,
                                "imported": True,
                                "created_at": get_malaysia_time().isoformat()
                            })
                        results["attendance_imported"] += 1
                    elif val in ["A", "ABSENT"]:
                        await db.attendance.delete_one({"session_id": session_id, "participant_id": pid, "date": d_date})
                        
            except Exception as e:
                results["errors"].append(f"Attendance row error: {str(e)}")
    
    # === Process Vehicle Checklist (Sheet 3) ===
    if "Vehicle Checklist" in wb.sheetnames:
        ws = wb["Vehicle Checklist"]
        
        # Read checklist item names from header row (columns starting from 7th)
        # Stop at "Remarks" column which is the last column
        header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=False))[0]
        checklist_items_list = []
        remarks_header_col = None
        for col_idx in range(6, len(header_row)):
            h_val = header_row[col_idx].value
            if h_val and str(h_val).strip():
                h_lower = str(h_val).strip().lower()
                if h_lower == "remarks":
                    remarks_header_col = col_idx
                else:
                    checklist_items_list.append(str(h_val).strip().lower())
        
        for row in ws.iter_rows(min_row=5, values_only=False):
            try:
                ic = str(row[2].value or "").strip()
                if not ic or ic not in ic_to_participant:
                    continue
                
                participant = ic_to_participant[ic]
                pid = participant["id"]
                
                # Vehicle details (cols 3-5: model, reg no, road tax)
                vehicle_model = str(row[3].value or "").strip()
                registration_no = str(row[4].value or "").strip()
                roadtax_expiry = str(row[5].value or "").strip()
                
                has_vehicle_data = vehicle_model or registration_no
                has_checklist_data = any(
                    str(row[6 + ci_idx].value or "").strip()
                    for ci_idx in range(len(checklist_items_list))
                    if 6 + ci_idx < len(row)
                )
                
                if not has_vehicle_data and not has_checklist_data:
                    continue
                
                # Upsert vehicle details
                if has_vehicle_data:
                    vd_data = {
                        "vehicle_model": vehicle_model,
                        "registration_number": registration_no,
                        "roadtax_expiry": roadtax_expiry,
                        "imported": True
                    }
                    existing_vd = await db.vehicle_details.find_one({"session_id": session_id, "participant_id": pid})
                    if existing_vd:
                        await db.vehicle_details.update_one(
                            {"session_id": session_id, "participant_id": pid},
                            {"$set": vd_data}
                        )
                    else:
                        await db.vehicle_details.insert_one({
                            "id": str(uuid.uuid4()),
                            "participant_id": pid,
                            "session_id": session_id,
                            "created_at": get_malaysia_time().isoformat(),
                            **vd_data
                        })
                
                # Upsert vehicle checklist
                if has_checklist_data:
                    items = []
                    for ci_idx, ci_name in enumerate(checklist_items_list):
                        col_idx = 6 + ci_idx
                        if col_idx < len(row):
                            status = str(row[col_idx].value or "").strip().lower()
                            if status in ["good", "needs_repair", "needs repair", "satisfactory", "n/a", "na"]:
                                status = status.replace("needs repair", "needs_repair").replace("na", "n/a")
                                items.append({"item": ci_name, "status": status, "comments": "", "photo_url": None})
                    
                    # Check for remarks column
                    remarks_text = ""
                    if remarks_header_col is not None and remarks_header_col < len(row):
                        remarks_text = str(row[remarks_header_col].value or "").strip()
                    
                    if items:
                        existing_vc = await db.vehicle_checklists.find_one({"session_id": session_id, "participant_id": pid})
                        vc_data = {
                            "checklist_items": items,
                            "interval": "imported",
                            "submitted_at": get_malaysia_time().isoformat(),
                            "verification_status": "imported",
                            "imported": True,
                            "remarks": remarks_text
                        }
                        if existing_vc:
                            await db.vehicle_checklists.update_one(
                                {"session_id": session_id, "participant_id": pid},
                                {"$set": vc_data}
                            )
                        else:
                            await db.vehicle_checklists.insert_one({
                                "id": str(uuid.uuid4()),
                                "participant_id": pid,
                                "session_id": session_id,
                                **vc_data
                            })
                        results["vehicle_checklists_imported"] += 1
                
            except Exception as e:
                results["errors"].append(f"Vehicle checklist row error: {str(e)}")
    
    # === Process Feedback (Sheet 4) ===
    if "Feedback" in wb.sheetnames:
        ws = wb["Feedback"]
        
        # Read question names from header row (columns starting from 4th, 0-indexed col 3)
        header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=False))[0]
        feedback_questions_from_header = []
        for col_idx in range(3, len(header_row)):
            h_val = header_row[col_idx].value
            if h_val and str(h_val).strip():
                # Strip " (1-5)" suffix if present
                q_text = str(h_val).strip()
                if q_text.endswith("(1-5)"):
                    q_text = q_text[:-5].strip()
                feedback_questions_from_header.append(q_text)
        
        for row in ws.iter_rows(min_row=5, values_only=False):
            try:
                ic = str(row[2].value or "").strip()
                if not ic or ic not in ic_to_participant:
                    continue
                
                participant = ic_to_participant[ic]
                pid = participant["id"]
                
                # Build responses array
                responses = []
                has_any_response = False
                for fq_idx, fq_text in enumerate(feedback_questions_from_header):
                    col_idx = 3 + fq_idx
                    if col_idx >= len(row):
                        break
                    answer = row[col_idx].value
                    if answer is not None and str(answer).strip() != "":
                        # Try to convert to number for rating questions
                        try:
                            answer_val = int(float(str(answer)))
                            if 1 <= answer_val <= 5:
                                answer = answer_val
                            else:
                                answer = str(answer).strip()
                        except (ValueError, TypeError):
                            answer = str(answer).strip()
                        responses.append({"question": fq_text, "answer": answer})
                        has_any_response = True
                    else:
                        responses.append({"question": fq_text, "answer": ""})
                
                if has_any_response:
                    existing_fb = await db.course_feedback.find_one({"session_id": session_id, "participant_id": pid})
                    fb_data = {
                        "responses": responses,
                        "submitted_at": get_malaysia_time().isoformat(),
                        "imported": True,
                        "imported_at": get_malaysia_time().isoformat()
                    }
                    if existing_fb:
                        await db.course_feedback.update_one(
                            {"session_id": session_id, "participant_id": pid},
                            {"$set": fb_data}
                        )
                    else:
                        await db.course_feedback.insert_one({
                            "id": str(uuid.uuid4()),
                            "participant_id": pid,
                            "session_id": session_id,
                            "program_id": session.get("program_id", ""),
                            **fb_data
                        })
                    results["feedback_imported"] += 1
                    
            except Exception as e:
                results["errors"].append(f"Feedback row error: {str(e)}")
    
    wb.close()
    return results


@router.get("/{session_id}/export-feedback-excel")
async def export_session_feedback_excel(session_id: str, current_user: User = Depends(get_current_user)):
    """Export all feedback for a session as Excel file"""
    if current_user.role not in ["admin", "super_admin", "coordinator"]:
        raise HTTPException(status_code=403, detail="Permission denied")

    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    feedbacks = await db.course_feedback.find({"session_id": session_id}, {"_id": 0}).to_list(None)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Session Feedback"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    if not feedbacks:
        ws.append(["No feedback submitted for this session"])
    else:
        all_question_ids = set()
        for fb in feedbacks:
            for response in fb.get("responses", []):
                all_question_ids.add(response.get("question_id", ""))

        question_ids = sorted(all_question_ids)

        headers = ["Participant Name", "IC Number", "Submitted At"]
        for qid in question_ids:
            headers.append(f"Q: {qid}")
        headers.append("Comments")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for idx, fb in enumerate(feedbacks, 2):
            participant = await db.users.find_one({"id": fb.get("participant_id")}, {"_id": 0, "full_name": 1, "id_number": 1})
            ws.cell(row=idx, column=1, value=participant.get("full_name") if participant else "Unknown")
            ws.cell(row=idx, column=2, value=participant.get("id_number") if participant else "")
            ws.cell(row=idx, column=3, value=fb.get("submitted_at", ""))

            response_map = {r.get("question_id"): r.get("rating") or r.get("text", "") for r in fb.get("responses", [])}
            for qcol, qid in enumerate(question_ids, 4):
                ws.cell(row=idx, column=qcol, value=response_map.get(qid, ""))

            ws.cell(row=idx, column=len(headers), value=fb.get("comments", ""))

        for col in ws.columns:
            max_length = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = session.get("name", "session").replace(" ", "_")[:30]
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Feedback_{safe_name}.xlsx"}
    )



@router.get("/{session_id}/notes")
async def get_session_notes(session_id: str, current_user: User = Depends(get_current_user)):
    """Get trainer notes for a session"""
    if current_user.role not in ["trainer", "admin", "super_admin", "coordinator"]:
        raise HTTPException(status_code=403, detail="Unauthorized")
    notes = await db.session_notes.find(
        {"session_id": session_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return notes


@router.post("/{session_id}/notes")
async def add_session_note(session_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """Add a note to a session"""
    if current_user.role not in ["trainer", "admin", "super_admin", "coordinator"]:
        raise HTTPException(status_code=403, detail="Unauthorized")
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    note = {
        "id": str(uuid.uuid4()),
        "session_id": session_id,
        "trainer_id": current_user.id,
        "trainer_name": current_user.full_name,
        "content": data.get("content", "").strip(),
        "created_at": get_malaysia_time().isoformat()
    }
    if not note["content"]:
        raise HTTPException(status_code=400, detail="Note content is required")
    await db.session_notes.insert_one(note)
    note.pop("_id", None)
    return note


@router.delete("/{session_id}/notes/{note_id}")
async def delete_session_note(session_id: str, note_id: str, current_user: User = Depends(get_current_user)):
    """Delete a session note (only the author or admin)"""
    note = await db.session_notes.find_one({"id": note_id, "session_id": session_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    if note["trainer_id"] != current_user.id and current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="You can only delete your own notes")
    await db.session_notes.delete_one({"id": note_id})
    return {"message": "Note deleted"}
