"""
Accounting Engine - Phase 1: Foundation
Double-Entry Accounting System for MDDRC Training Management

Collections:
- chart_of_accounts: Chart of Accounts (COA)
- journal_entries: Double-entry journal entries
- accounting_periods: Fiscal period control
- accounting_settings: System configuration
- journal_entry_counters: Atomic journal numbering
- accounting_audit_log: Audit trail

Endpoints: ~30 (including Excel exports)
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pydantic import BaseModel, Field, ConfigDict, validator
from pymongo import ReturnDocument
import uuid
import io

from core import db, get_current_user, get_malaysia_time
from models import User


router = APIRouter(prefix="/accounting", tags=["accounting"])


# ============ PYDANTIC MODELS ============

class ChartOfAccountCreate(BaseModel):
    """Create a new account in Chart of Accounts"""
    account_code: str = Field(..., min_length=3, max_length=10)
    account_name: str = Field(..., min_length=2, max_length=100)
    account_type: str = Field(..., pattern="^(Asset|Liability|Equity|Income|Expense)$")
    account_category: str = Field(..., min_length=2, max_length=50)
    parent_code: Optional[str] = None
    description: Optional[str] = None
    normal_balance: str = Field(default="debit", pattern="^(debit|credit)$")
    statement_type: Optional[str] = Field(default=None, pattern="^(profit_and_loss|balance_sheet)$")
    pnl_section: Optional[str] = Field(default=None, pattern="^(revenue|cost_of_sales|operating_expense|other_income|other_expense)$")


class ChartOfAccountUpdate(BaseModel):
    """Update an existing account"""
    account_name: Optional[str] = None
    account_category: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None
    statement_type: Optional[str] = None
    pnl_section: Optional[str] = None


class JournalLine(BaseModel):
    """Single line in a journal entry (debit or credit)"""
    account_code: str
    debit: float = 0.0
    credit: float = 0.0
    memo: Optional[str] = None
    
    @validator('debit', 'credit')
    def round_to_two_decimals(cls, v):
        return round(float(v), 2)


class JournalEntryCreate(BaseModel):
    """Create a manual journal entry"""
    date: str = Field(..., pattern="^\\d{4}-\\d{2}-\\d{2}$")
    description: str = Field(..., min_length=5, max_length=500)
    lines: List[JournalLine] = Field(..., min_items=2)
    
    @validator('lines')
    def validate_balanced(cls, v):
        total_debit = sum(line.debit for line in v)
        total_credit = sum(line.credit for line in v)
        if abs(total_debit - total_credit) > 0.01:
            raise ValueError(f"Journal entry must be balanced. Debit: {total_debit}, Credit: {total_credit}")
        return v


class PeriodCloseRequest(BaseModel):
    """Request to close an accounting period"""
    reason: Optional[str] = None


class PeriodReopenRequest(BaseModel):
    """Request to reopen a closed period"""
    reason: str = Field(..., min_length=10)



@router.get("/diagnose-journal-references")
async def diagnose_journal_references(current_user: User = Depends(get_current_user)):
    """Diagnostic: Show sample journal entry references to help debug migration issues."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    total = await db.journal_entries.count_documents({})
    
    # Get distinct source_reference patterns
    all_refs = await db.journal_entries.find(
        {},
        {"_id": 0, "id": 1, "source_reference": 1, "source_module": 1, "source_id": 1, "description": 1}
    ).to_list(5000)
    
    # Categorize
    with_inv = [r for r in all_refs if "INV" in (r.get("source_reference") or "")]
    tf_cf_mc = [r for r in all_refs if (r.get("source_reference") or "").startswith(("TF-", "CF-", "MC-"))]
    tf_cf_mc_no_inv = [r for r in tf_cf_mc if "INV" not in (r.get("source_reference") or "")]
    no_source_id = [r for r in tf_cf_mc_no_inv if not r.get("source_id")]
    
    return {
        "total_journal_entries": total,
        "with_invoice_ref": len(with_inv),
        "tf_cf_mc_total": len(tf_cf_mc),
        "tf_cf_mc_needing_fix": len(tf_cf_mc_no_inv),
        "tf_cf_mc_missing_source_id": len(no_source_id),
        "sample_needing_fix": [
            {"ref": r.get("source_reference"), "source_id": r.get("source_id"), "module": r.get("source_module"), "desc": (r.get("description") or "")[:60]}
            for r in tf_cf_mc_no_inv[:15]
        ],
        "sample_already_fixed": [r.get("source_reference") for r in with_inv[:5]],
        "all_unique_ref_prefixes": list(set([(r.get("source_reference") or "?")[:3] for r in all_refs]))[:20]
    }


@router.post("/migrate-journal-references")
@router.get("/migrate-journal-references")
async def migrate_journal_references(current_user: User = Depends(get_current_user)):
    """One-time migration: Update TF-xxx, CF-xxx, MC-xxx references to include invoice numbers.
    Finds ALL entries starting with TF-, CF-, MC- that do NOT already contain an invoice number (INV/)."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    updated = 0
    skipped_no_session = 0
    skipped_no_invoice = 0
    errors = []
    sample_refs = []
    
    # Find ALL journal entries with TF-/CF-/MC- prefix, then filter out already-migrated ones
    all_tf_cf_mc = await db.journal_entries.find(
        {"source_reference": {"$regex": "^(TF|CF|MC)-"}},
        {"_id": 0, "id": 1, "source_reference": 1, "source_module": 1, "source_id": 1}
    ).to_list(5000)
    sample_refs = [e.get("source_reference", "?") for e in all_tf_cf_mc[:15]]
    entries = [e for e in all_tf_cf_mc if "INV" not in (e.get("source_reference") or "")]
    
    for entry in entries:
        try:
            ref = entry.get("source_reference", "")
            prefix = ref[:2]  # TF, CF, or MC
            source_id = entry.get("source_id")
            session_id = None
            
            # Find the session_id from the fee/commission record
            if prefix == "TF" and source_id:
                fee = await db.trainer_fees.find_one({"id": source_id}, {"_id": 0, "session_id": 1})
                if fee:
                    session_id = fee.get("session_id")
            elif prefix == "CF" and source_id:
                fee = await db.coordinator_fees.find_one({"id": source_id}, {"_id": 0, "session_id": 1})
                if fee:
                    session_id = fee.get("session_id")
            elif prefix == "MC" and source_id:
                comm = await db.marketing_commissions.find_one({"id": source_id}, {"_id": 0, "session_id": 1})
                if comm:
                    session_id = comm.get("session_id")
            
            if not session_id:
                skipped_no_session += 1
                continue
            
            # Find ANY invoice for this session (broader status match)
            invoice = await db.invoices.find_one(
                {"session_id": session_id},
                {"_id": 0, "invoice_number": 1}
            )
            if not invoice:
                # Fallback: check session.invoice_id
                sess = await db.sessions.find_one({"id": session_id}, {"_id": 0, "invoice_id": 1})
                if sess and sess.get("invoice_id"):
                    invoice = await db.invoices.find_one({"id": sess["invoice_id"]}, {"_id": 0, "invoice_number": 1})
            
            if invoice and invoice.get("invoice_number"):
                new_ref = f"{prefix}-{invoice['invoice_number']}"
                await db.journal_entries.update_one(
                    {"id": entry["id"]},
                    {"$set": {"source_reference": new_ref}}
                )
                updated += 1
            else:
                skipped_no_invoice += 1
        except Exception as e:
            errors.append(f"{entry.get('id')}: {str(e)}")
    
    return {
        "message": f"Migration complete. Updated {updated} of {len(entries)} entries.",
        "updated": updated,
        "total_found": len(entries),
        "skipped_no_session": skipped_no_session,
        "skipped_no_invoice": skipped_no_invoice,
        "sample_existing_refs": sample_refs[:10],
        "errors": errors[:10]
    }


# ============ HELPER FUNCTIONS ============

def round_money(value: float) -> float:
    """Round money to 2 decimal places using banker's rounding"""
    return float(Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))


async def get_next_journal_number(entry_date: str) -> str:
    """
    Generate atomic, unique journal entry number.
    Format: JE-YYYY-MM-XXXX
    Resets monthly.
    """
    year = entry_date[:4]
    month = entry_date[5:7]
    counter_key = f"JE-{year}-{month}"
    
    result = await db.journal_entry_counters.find_one_and_update(
        {"_id": counter_key},
        {"$inc": {"seq": 1}},
        return_document=ReturnDocument.AFTER,
        upsert=True
    )
    
    sequence = result["seq"]
    return f"{counter_key}-{str(sequence).zfill(4)}"


async def check_period_open(entry_date: str) -> bool:
    """Check if the accounting period is open for the given date"""
    year = int(entry_date[:4])
    month = int(entry_date[5:7])
    
    period = await db.accounting_periods.find_one(
        {"year": year, "month": month},
        {"_id": 0}
    )
    
    # If no period record exists, assume open (will be created on first use)
    if not period:
        return True
    
    return period.get("status") == "open"


async def get_account_by_code(account_code: str) -> dict:
    """Get account details by code"""
    account = await db.chart_of_accounts.find_one(
        {"account_code": account_code, "is_active": True},
        {"_id": 0}
    )
    return account


async def validate_journal_accounts(lines: List[JournalLine]) -> List[dict]:
    """Validate all accounts exist and are active, return enriched lines"""
    enriched_lines = []
    for i, line in enumerate(lines):
        account = await get_account_by_code(line.account_code)
        if not account:
            raise HTTPException(
                status_code=400, 
                detail=f"Account {line.account_code} not found or inactive"
            )
        enriched_lines.append({
            "line_no": i + 1,
            "account_code": line.account_code,
            "account_name": account["account_name"],
            "account_type": account["account_type"],
            "debit": round_money(line.debit),
            "credit": round_money(line.credit),
            "memo": line.memo or ""
        })
    return enriched_lines


# ============ PHASE 2: AUTO-POSTING FUNCTIONS ============

async def get_accounting_settings():
    """Get accounting settings with defaults"""
    settings = await db.accounting_settings.find_one({"id": "accounting_settings"}, {"_id": 0})
    if not settings:
        # Return defaults if not initialized
        return {
            "use_deferred_revenue": True,
            "auto_post_invoices": True,
            "auto_post_payments": True,
            "auto_post_expenses": True,
            "auto_post_payroll": True,
            "default_bank_account": "1000",
            "default_ar_account": "1100",
            "default_ap_account": "2100",
            "default_revenue_account": "4000",
            "default_deferred_revenue_account": "2300",
            "default_sst_account": "2200",
            "accounting_start_date": "2026-01-01"
        }
    return settings


async def create_auto_journal_entry(
    entry_date: str,
    description: str,
    source_module: str,
    source_id: str,
    source_reference: str,
    lines: list,
    created_by_id: str = "system",
    created_by_name: str = "System Auto-Post"
) -> dict:
    """
    Create and post a journal entry automatically from a source module.
    
    Idempotent: Checks if entry already exists for source_id to prevent duplicates.
    
    Args:
        entry_date: Date in YYYY-MM-DD format
        description: Journal description
        source_module: invoice, payment, credit_note, expense, payroll
        source_id: ID of the source document
        source_reference: Human-readable reference (e.g., invoice number)
        lines: List of dicts with account_code, debit, credit, memo
        created_by_id: User ID who triggered the action
        created_by_name: User name for display
    
    Returns:
        Created journal entry dict or existing entry if duplicate
    """
    # IDEMPOTENCY CHECK: Don't create duplicate entries
    existing = await db.journal_entries.find_one({
        "source_module": source_module,
        "source_id": source_id,
        "status": {"$ne": "voided"}
    }, {"_id": 0})
    
    if existing:
        return {"message": "Journal entry already exists", "journal_entry": existing, "is_duplicate": True}
    
    # Check if period is open
    if not await check_period_open(entry_date):
        return {"error": f"Cannot post to closed period: {entry_date[:7]}", "journal_entry": None}
    
    # Check accounting start date
    settings = await get_accounting_settings()
    start_date = settings.get("accounting_start_date", "2026-01-01")
    if entry_date < start_date:
        return {"error": f"Date {entry_date} is before accounting start date {start_date}", "journal_entry": None}
    
    # Validate and enrich lines
    enriched_lines = []
    for i, line in enumerate(lines):
        account = await get_account_by_code(line["account_code"])
        if not account:
            return {"error": f"Account {line['account_code']} not found", "journal_entry": None}
        
        enriched_lines.append({
            "line_no": i + 1,
            "account_code": line["account_code"],
            "account_name": account["account_name"],
            "account_type": account["account_type"],
            "debit": round_money(line.get("debit", 0)),
            "credit": round_money(line.get("credit", 0)),
            "memo": line.get("memo", "")
        })
    
    # Calculate totals
    total_debit = round_money(sum(line["debit"] for line in enriched_lines))
    total_credit = round_money(sum(line["credit"] for line in enriched_lines))
    
    # Validate balanced
    if abs(total_debit - total_credit) > 0.01:
        return {"error": f"Unbalanced entry: DR {total_debit} != CR {total_credit}", "journal_entry": None}
    
    # Generate journal number
    journal_no = await get_next_journal_number(entry_date)
    
    now = get_malaysia_time().isoformat()
    journal_entry = {
        "id": str(uuid.uuid4()),
        "journal_no": journal_no,
        "date": entry_date,
        "description": description,
        "source_module": source_module,
        "source_id": source_id,
        "source_reference": source_reference,
        "status": "posted",  # Auto-posted
        "lines": enriched_lines,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "is_balanced": True,
        "created_by": created_by_id,
        "created_by_name": created_by_name,
        "posted_by": "system",
        "posted_by_name": "System Auto-Post",
        "posted_at": now,
        "voided_by": None,
        "voided_at": None,
        "void_reason": None,
        "created_at": now,
        "updated_at": now
    }
    
    await db.journal_entries.insert_one(journal_entry)
    
    # Log the action
    await log_accounting_action(
        action="journal_auto_posted",
        entity_type="journal_entry",
        entity_id=journal_entry["id"],
        performed_by_id=created_by_id,
        performed_by_name=created_by_name,
        after_value={
            "journal_no": journal_no, 
            "source_module": source_module,
            "source_reference": source_reference,
            "total_debit": total_debit
        }
    )
    
    journal_entry.pop("_id", None)
    return {"message": "Journal entry auto-posted", "journal_entry": journal_entry, "is_duplicate": False}


async def log_accounting_action(
    action: str,
    entity_type: str,
    entity_id: str,
    performed_by_id: str = None,
    performed_by_name: str = None,
    performed_by: "User" = None,
    before_value: dict = None,
    after_value: dict = None,
    reason: str = None
):
    """Log accounting actions for audit trail - supports both User object and ID/name"""
    if performed_by:
        performed_by_id = performed_by.id
        performed_by_name = performed_by.full_name
    
    log_entry = {
        "id": str(uuid.uuid4()),
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "before_value": before_value,
        "after_value": after_value,
        "performed_by": performed_by_id or "system",
        "performed_by_name": performed_by_name or "System",
        "reason": reason,
        "timestamp": get_malaysia_time().isoformat()
    }
    await db.accounting_audit_log.insert_one(log_entry)
    return log_entry


# ============ AUTO-POSTING: INVOICE ISSUED ============

async def post_invoice_issued(
    invoice: dict,
    session: dict = None,
    user_id: str = "system",
    user_name: str = "System"
) -> dict:
    """
    Create journal entry when invoice is issued.
    
    Revenue Recognition Rules:
    - If session.completion_status == "completed": Post to Revenue
    - If session not completed: Post to Deferred Revenue
    
    Journal Entry:
    DR 1100 Accounts Receivable     [total_amount]
    CR 4000 Training Revenue        [total_amount - tax_amount] (if completed)
    CR 2300 Deferred Revenue        [total_amount - tax_amount] (if not completed)
    CR 2200 SST Payable             [tax_amount]
    """
    settings = await get_accounting_settings()
    if not settings.get("auto_post_invoices", True):
        return {"message": "Auto-posting invoices is disabled", "journal_entry": None}
    
    invoice_id = invoice.get("id")
    invoice_number = invoice.get("invoice_number", "Unknown")
    total_amount = round_money(float(invoice.get("total_amount", 0)))
    tax_amount = round_money(float(invoice.get("tax_amount") or invoice.get("sst_amount") or 0))
    revenue_amount = round_money(total_amount - tax_amount)
    
    # Determine invoice date
    invoice_date = invoice.get("invoice_date") or invoice.get("created_at", "")[:10]
    if not invoice_date or len(invoice_date) < 10:
        invoice_date = get_malaysia_time().strftime("%Y-%m-%d")
    
    # Get session completion status for revenue recognition
    is_completed = True  # Default to completed if no session
    if session:
        completion_status = session.get("completion_status", "completed")
        is_completed = completion_status == "completed"
    elif invoice.get("session_id"):
        # Fetch session if not provided
        session = await db.sessions.find_one({"id": invoice.get("session_id")}, {"_id": 0, "completion_status": 1})
        if session:
            is_completed = session.get("completion_status", "completed") == "completed"
    
    # Determine revenue account based on completion status
    if is_completed:
        revenue_account = settings.get("default_revenue_account", "4000")
        revenue_description = "Training Revenue"
    else:
        revenue_account = settings.get("default_deferred_revenue_account", "2300")
        revenue_description = "Deferred Revenue (Pending Session Completion)"
    
    # Get client/company name for memo
    company_name = invoice.get("bill_to_name") or invoice.get("company_name") or "Unknown Client"
    
    # Build journal lines
    lines = [
        {
            "account_code": settings.get("default_ar_account", "1100"),
            "debit": total_amount,
            "credit": 0,
            "memo": f"AR - {company_name}"
        },
        {
            "account_code": revenue_account,
            "debit": 0,
            "credit": revenue_amount,
            "memo": revenue_description
        }
    ]
    
    # Add SST line if applicable
    if tax_amount > 0:
        lines.append({
            "account_code": settings.get("default_sst_account", "2200"),
            "debit": 0,
            "credit": tax_amount,
            "memo": "SST Payable (6%)"
        })
    
    result = await create_auto_journal_entry(
        entry_date=invoice_date,
        description=f"Invoice {invoice_number} issued to {company_name}",
        source_module="invoice",
        source_id=invoice_id,
        source_reference=invoice_number,
        lines=lines,
        created_by_id=user_id,
        created_by_name=user_name
    )
    
    # Store deferred revenue flag on invoice for later recognition
    if not is_completed and result.get("journal_entry"):
        await db.invoices.update_one(
            {"id": invoice_id},
            {"$set": {"has_deferred_revenue": True, "deferred_revenue_amount": revenue_amount}}
        )
    
    return result


# ============ AUTO-POSTING: SESSION COMPLETED (Revenue Recognition) ============

async def post_session_completed_revenue(
    session_id: str,
    user_id: str = "system",
    user_name: str = "System"
) -> list:
    """
    Recognize deferred revenue when session is completed.
    
    Finds all invoices with deferred revenue and creates journal entries:
    DR 2300 Deferred Revenue        [deferred_amount]
    CR 4000 Training Revenue        [deferred_amount]
    """
    settings = await get_accounting_settings()
    
    # Find invoices with deferred revenue for this session
    invoices = await db.invoices.find({
        "session_id": session_id,
        "has_deferred_revenue": True,
        "status": {"$in": ["issued", "partial", "paid"]}
    }, {"_id": 0}).to_list(100)
    
    if not invoices:
        return []
    
    results = []
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0, "name": 1, "start_date": 1})
    session_name = session.get("name", "Unknown") if session else "Unknown"
    recognition_date = get_malaysia_time().strftime("%Y-%m-%d")
    
    for invoice in invoices:
        deferred_amount = round_money(float(invoice.get("deferred_revenue_amount", 0)))
        if deferred_amount <= 0:
            continue
        
        invoice_number = invoice.get("invoice_number", "Unknown")
        
        lines = [
            {
                "account_code": settings.get("default_deferred_revenue_account", "2300"),
                "debit": deferred_amount,
                "credit": 0,
                "memo": f"Deferred Revenue recognized - {invoice_number}"
            },
            {
                "account_code": settings.get("default_revenue_account", "4000"),
                "debit": 0,
                "credit": deferred_amount,
                "memo": f"Training Revenue - Session {session_name}"
            }
        ]
        
        result = await create_auto_journal_entry(
            entry_date=recognition_date,
            description=f"Revenue recognition for {invoice_number} - Session completed: {session_name}",
            source_module="revenue_recognition",
            source_id=f"{invoice.get('id')}_recognition",
            source_reference=f"RR-{invoice_number}",
            lines=lines,
            created_by_id=user_id,
            created_by_name=user_name
        )
        
        if result.get("journal_entry") and not result.get("is_duplicate"):
            # Mark invoice as revenue recognized
            await db.invoices.update_one(
                {"id": invoice.get("id")},
                {"$set": {"has_deferred_revenue": False, "revenue_recognized_at": recognition_date}}
            )
        
        results.append(result)
    
    return results


# ============ AUTO-POSTING: PAYMENT RECEIVED ============

async def post_payment_received(
    payment: dict,
    invoice: dict = None,
    user_id: str = "system",
    user_name: str = "System"
) -> dict:
    """
    Create journal entry when payment is received.
    
    Journal Entry:
    DR 1000 Cash at Bank            [amount]
    CR 1100 Accounts Receivable     [amount]
    """
    settings = await get_accounting_settings()
    if not settings.get("auto_post_payments", True):
        return {"message": "Auto-posting payments is disabled", "journal_entry": None}
    
    payment_id = payment.get("id")
    payment_reference = payment.get("reference_number") or payment.get("id", "")[:8]
    amount = round_money(float(payment.get("amount", 0)))
    
    # Get payment date
    payment_date = payment.get("payment_date") or payment.get("created_at", "")[:10]
    if not payment_date or len(payment_date) < 10:
        payment_date = get_malaysia_time().strftime("%Y-%m-%d")
    
    # Get invoice details if not provided
    if not invoice and payment.get("invoice_id"):
        invoice = await db.invoices.find_one({"id": payment.get("invoice_id")}, {"_id": 0})
    
    invoice_number = invoice.get("invoice_number", "Unknown") if invoice else "Unknown"
    company_name = invoice.get("bill_to_name", "Unknown") if invoice else "Unknown"
    
    lines = [
        {
            "account_code": settings.get("default_bank_account", "1000"),
            "debit": amount,
            "credit": 0,
            "memo": f"Payment received - {payment_reference}"
        },
        {
            "account_code": settings.get("default_ar_account", "1100"),
            "debit": 0,
            "credit": amount,
            "memo": f"AR reduction - {invoice_number}"
        }
    ]
    
    result = await create_auto_journal_entry(
        entry_date=payment_date,
        description=f"Payment received for {invoice_number} from {company_name}",
        source_module="payment",
        source_id=payment_id,
        source_reference=f"PMT-{payment_reference}",
        lines=lines,
        created_by_id=user_id,
        created_by_name=user_name
    )
    
    return result


# ============ AUTO-POSTING: CREDIT NOTE ISSUED ============

async def post_credit_note_issued(
    credit_note: dict,
    user_id: str = "system",
    user_name: str = "System"
) -> dict:
    """
    Create journal entry when credit note is issued.
    
    Journal Entry (reverses invoice):
    DR 4000 Training Revenue        [amount - tax]
    DR 2200 SST Payable             [tax]
    CR 1100 Accounts Receivable     [amount]
    """
    settings = await get_accounting_settings()
    
    cn_id = credit_note.get("id")
    cn_number = credit_note.get("credit_note_number", "Unknown")
    amount = round_money(float(credit_note.get("amount", 0)))
    tax_amount = round_money(float(credit_note.get("tax_amount") or 0))
    revenue_amount = round_money(amount - tax_amount)
    
    # Get CN date
    cn_date = credit_note.get("date") or credit_note.get("created_at", "")[:10]
    if not cn_date or len(cn_date) < 10:
        cn_date = get_malaysia_time().strftime("%Y-%m-%d")
    
    company_name = credit_note.get("company_name", "Unknown")
    reason = credit_note.get("reason", "")
    
    lines = [
        {
            "account_code": settings.get("default_revenue_account", "4000"),
            "debit": revenue_amount,
            "credit": 0,
            "memo": f"Revenue reversal - {reason}"
        },
        {
            "account_code": settings.get("default_ar_account", "1100"),
            "debit": 0,
            "credit": amount,
            "memo": f"AR reduction - CN to {company_name}"
        }
    ]
    
    # Add SST reversal if applicable
    if tax_amount > 0:
        lines.insert(1, {
            "account_code": settings.get("default_sst_account", "2200"),
            "debit": tax_amount,
            "credit": 0,
            "memo": "SST reversal"
        })
    
    result = await create_auto_journal_entry(
        entry_date=cn_date,
        description=f"Credit Note {cn_number} issued to {company_name} - {reason}",
        source_module="credit_note",
        source_id=cn_id,
        source_reference=cn_number,
        lines=lines,
        created_by_id=user_id,
        created_by_name=user_name
    )
    
    return result


# ============ AUTO-POSTING: EXPENSE RECORDED ============

async def post_expense_recorded(
    expense: dict,
    expense_type: str = "session",  # session, petty_cash, manual
    user_id: str = "system",
    user_name: str = "System"
) -> dict:
    """
    Create journal entry when expense is recorded.
    
    Journal Entry:
    DR 5xxx Expense Account         [amount]
    CR 1000 Cash at Bank            [amount] (if paid)
    CR 2100 Accounts Payable        [amount] (if not paid)
    """
    settings = await get_accounting_settings()
    if not settings.get("auto_post_expenses", True):
        return {"message": "Auto-posting expenses is disabled", "journal_entry": None}
    
    expense_id = expense.get("id")
    amount = round_money(float(expense.get("actual_amount") or expense.get("amount") or 0))
    
    if amount <= 0:
        return {"message": "No amount to post", "journal_entry": None}
    
    # Get expense date
    expense_date = expense.get("date") or expense.get("created_at", "")[:10]
    if not expense_date or len(expense_date) < 10:
        expense_date = get_malaysia_time().strftime("%Y-%m-%d")
    
    # Determine expense account based on type/category
    expense_category = expense.get("category", "").lower()
    expense_account = "6400"  # Default: Office Expenses
    
    category_mapping = {
        "trainer": "5000",
        "coordinator": "5100",
        "marketing": "5200",
        "materials": "5300",
        "training materials": "5300",
        "venue": "5400",
        "logistics": "5400",
        "transport": "5500",
        "transportation": "5500",
        "salary": "6000",
        "wages": "6000",
        "office": "6400",
        "utilities": "6500",
        "petty cash": "6600",
    }
    
    for key, code in category_mapping.items():
        if key in expense_category:
            expense_account = code
            break
    
    # Determine if paid (use bank) or unpaid (use AP)
    is_paid = expense.get("status") == "paid" or expense.get("is_paid", True)
    credit_account = settings.get("default_bank_account", "1000") if is_paid else settings.get("default_ap_account", "2100")
    credit_memo = "Cash payment" if is_paid else "Accounts Payable"
    
    description_text = expense.get("description") or expense.get("name") or "Expense"
    
    lines = [
        {
            "account_code": expense_account,
            "debit": amount,
            "credit": 0,
            "memo": description_text[:100]
        },
        {
            "account_code": credit_account,
            "debit": 0,
            "credit": amount,
            "memo": credit_memo
        }
    ]
    
    # Build source reference
    if expense_type == "petty_cash":
        source_ref = f"PC-{expense_id[:8]}"
    else:
        source_ref = f"EXP-{expense_id[:8]}"
    
    result = await create_auto_journal_entry(
        entry_date=expense_date,
        description=f"Expense: {description_text[:50]}",
        source_module=f"expense_{expense_type}",
        source_id=expense_id,
        source_reference=source_ref,
        lines=lines,
        created_by_id=user_id,
        created_by_name=user_name
    )
    
    return result



async def _get_invoice_ref(prefix: str, record_id: str, trainer_fee: dict = None, coordinator_fee: dict = None, commission: dict = None, session: dict = None) -> str:
    """Build a journal reference that includes the invoice number for traceability.
    Format: TF-INV/MDDRC/2026/01/0001 or CF-INV/MDDRC/2026/01/0001
    Falls back to prefix-{id[:8]} if no invoice found."""
    session_id = None
    if session:
        session_id = session.get("id")
    elif trainer_fee:
        session_id = trainer_fee.get("session_id")
    elif coordinator_fee:
        session_id = coordinator_fee.get("session_id")
    elif commission:
        session_id = commission.get("session_id")
    
    if session_id:
        # Find invoices linked to this session
        invoice = await db.invoices.find_one(
            {"session_id": session_id, "status": {"$in": ["issued", "paid", "partial", "approved"]}},
            {"_id": 0, "invoice_number": 1}
        )
        if not invoice:
            # Fallback: check session.invoice_id
            sess = session or await db.sessions.find_one({"id": session_id}, {"_id": 0, "invoice_id": 1})
            if sess and sess.get("invoice_id"):
                invoice = await db.invoices.find_one({"id": sess["invoice_id"]}, {"_id": 0, "invoice_number": 1})
        
        if invoice and invoice.get("invoice_number"):
            return f"{prefix}-{invoice['invoice_number']}"
    
    return f"{prefix}-{record_id[:8]}"


# ============ AUTO-POSTING: TRAINER FEE RECORDED ============

async def post_trainer_fee(
    trainer_fee: dict,
    session: dict = None,
    user_id: str = "system",
    user_name: str = "System"
) -> dict:
    """
    Create journal entry when trainer fee is recorded.
    
    Journal Entry:
    DR 5000 Trainer Fees            [amount]
    CR 2100 Accounts Payable        [amount]
    """
    settings = await get_accounting_settings()
    
    fee_id = trainer_fee.get("id")
    amount = round_money(float(trainer_fee.get("fee_amount") or 0))
    
    if amount <= 0:
        return {"message": "No amount to post", "journal_entry": None}
    
    # Get session date
    if session:
        fee_date = session.get("start_date", "")[:10]
    elif trainer_fee.get("session_id"):
        session = await db.sessions.find_one({"id": trainer_fee.get("session_id")}, {"_id": 0, "start_date": 1, "name": 1})
        fee_date = session.get("start_date", "")[:10] if session else get_malaysia_time().strftime("%Y-%m-%d")
    else:
        fee_date = get_malaysia_time().strftime("%Y-%m-%d")
    
    trainer_name = trainer_fee.get("trainer_name", "Unknown Trainer")
    session_name = session.get("name", "Unknown Session") if session else "Unknown Session"
    
    lines = [
        {
            "account_code": "5000",  # Trainer Fees
            "debit": amount,
            "credit": 0,
            "memo": f"Trainer: {trainer_name}"
        },
        {
            "account_code": settings.get("default_ap_account", "2100"),
            "debit": 0,
            "credit": amount,
            "memo": f"Payable to {trainer_name}"
        }
    ]
    
    result = await create_auto_journal_entry(
        entry_date=fee_date,
        description=f"Trainer fee - {trainer_name} for {session_name}",
        source_module="trainer_fee",
        source_id=fee_id,
        source_reference=await _get_invoice_ref("TF", fee_id, trainer_fee=trainer_fee, session=session),
        lines=lines,
        created_by_id=user_id,
        created_by_name=user_name
    )
    
    return result


# ============ AUTO-POSTING: COORDINATOR FEE RECORDED ============

async def post_coordinator_fee(
    coordinator_fee: dict,
    session: dict = None,
    user_id: str = "system",
    user_name: str = "System"
) -> dict:
    """
    Create journal entry when coordinator fee is recorded.
    
    Journal Entry:
    DR 5100 Coordinator Fees        [amount]
    CR 2100 Accounts Payable        [amount]
    """
    settings = await get_accounting_settings()
    
    fee_id = coordinator_fee.get("id")
    amount = round_money(float(coordinator_fee.get("total_fee") or 0))
    
    if amount <= 0:
        return {"message": "No amount to post", "journal_entry": None}
    
    # Get session date
    if session:
        fee_date = session.get("start_date", "")[:10]
    elif coordinator_fee.get("session_id"):
        session = await db.sessions.find_one({"id": coordinator_fee.get("session_id")}, {"_id": 0, "start_date": 1, "name": 1})
        fee_date = session.get("start_date", "")[:10] if session else get_malaysia_time().strftime("%Y-%m-%d")
    else:
        fee_date = get_malaysia_time().strftime("%Y-%m-%d")
    
    coordinator_name = coordinator_fee.get("coordinator_name", "Unknown Coordinator")
    session_name = session.get("name", "Unknown Session") if session else "Unknown Session"
    
    lines = [
        {
            "account_code": "5100",  # Coordinator Fees
            "debit": amount,
            "credit": 0,
            "memo": f"Coordinator: {coordinator_name}"
        },
        {
            "account_code": settings.get("default_ap_account", "2100"),
            "debit": 0,
            "credit": amount,
            "memo": f"Payable to {coordinator_name}"
        }
    ]
    
    result = await create_auto_journal_entry(
        entry_date=fee_date,
        description=f"Coordinator fee - {coordinator_name} for {session_name}",
        source_module="coordinator_fee",
        source_id=fee_id,
        source_reference=await _get_invoice_ref("CF", fee_id, coordinator_fee=coordinator_fee, session=session),
        lines=lines,
        created_by_id=user_id,
        created_by_name=user_name
    )
    
    return result


# ============ AUTO-POSTING: MARKETING COMMISSION ============

async def post_marketing_commission(
    commission: dict,
    session: dict = None,
    user_id: str = "system",
    user_name: str = "System"
) -> dict:
    """
    Create journal entry when marketing commission is recorded/approved.
    
    Journal Entry:
    DR 5200 Marketing Commission    [amount]
    CR 2100 Accounts Payable        [amount]
    """
    settings = await get_accounting_settings()
    
    comm_id = commission.get("id")
    amount = round_money(float(commission.get("calculated_amount") or commission.get("amount") or 0))
    
    if amount <= 0:
        return {"message": "No amount to post", "journal_entry": None}
    
    # Get session date
    if session:
        comm_date = session.get("start_date", "")[:10]
    elif commission.get("session_id"):
        session = await db.sessions.find_one({"id": commission.get("session_id")}, {"_id": 0, "start_date": 1, "name": 1})
        comm_date = session.get("start_date", "")[:10] if session else get_malaysia_time().strftime("%Y-%m-%d")
    else:
        comm_date = get_malaysia_time().strftime("%Y-%m-%d")
    
    marketer_name = commission.get("marketing_user_name", "Unknown Marketer")
    session_name = session.get("name", "Unknown Session") if session else "Unknown Session"
    
    lines = [
        {
            "account_code": "5200",  # Marketing Commission
            "debit": amount,
            "credit": 0,
            "memo": f"Commission: {marketer_name}"
        },
        {
            "account_code": settings.get("default_ap_account", "2100"),
            "debit": 0,
            "credit": amount,
            "memo": f"Payable to {marketer_name}"
        }
    ]
    
    result = await create_auto_journal_entry(
        entry_date=comm_date,
        description=f"Marketing commission - {marketer_name} for {session_name}",
        source_module="marketing_commission",
        source_id=comm_id,
        source_reference=await _get_invoice_ref("MC", comm_id, commission=commission, session=session),
        lines=lines,
        created_by_id=user_id,
        created_by_name=user_name
    )
    
    return result


# ============ AUTO-POSTING: PAYROLL ============

async def post_payroll(
    payslip: dict,
    user_id: str = "system",
    user_name: str = "System"
) -> dict:
    """
    Create journal entry when payroll is processed.
    
    Journal Entry:
    DR 6000 Salary & Wages          [gross_salary]
    DR 6100 EPF Employer            [epf_employer]
    DR 6200 SOCSO Employer          [socso_employer]
    DR 6300 EIS Employer            [eis_employer]
    CR 2400 EPF Payable             [epf_employee + epf_employer]
    CR 2450 SOCSO Payable           [socso_employee + socso_employer]
    CR 2460 EIS Payable             [eis_employee + eis_employer]
    CR 2470 PCB Payable             [pcb]
    CR 1000 Cash at Bank            [net_pay]
    """
    settings = await get_accounting_settings()
    if not settings.get("auto_post_payroll", True):
        return {"message": "Auto-posting payroll is disabled", "journal_entry": None}
    
    payslip_id = payslip.get("id")
    
    # Get payslip amounts (nett_pay with double-t is the field name in payslips collection)
    gross_salary = round_money(float(payslip.get("gross_salary") or payslip.get("basic_salary") or 0))
    net_pay = round_money(float(payslip.get("nett_pay") or payslip.get("net_pay") or 0))
    
    epf_employee = round_money(float(payslip.get("epf_employee") or 0))
    epf_employer = round_money(float(payslip.get("epf_employer") or 0))
    socso_employee = round_money(float(payslip.get("socso_employee") or 0))
    socso_employer = round_money(float(payslip.get("socso_employer") or 0))
    eis_employee = round_money(float(payslip.get("eis_employee") or 0))
    eis_employer = round_money(float(payslip.get("eis_employer") or 0))
    pcb = round_money(float(payslip.get("pcb") or payslip.get("mtd") or 0))
    
    if gross_salary <= 0:
        return {"message": "No salary to post", "journal_entry": None}
    
    # Get payroll date (use payment date or month-end)
    year = payslip.get("year", 2026)
    month = payslip.get("month", 1)
    payroll_date = f"{year}-{str(month).zfill(2)}-28"  # Use 28th as standard payroll date
    
    employee_name = payslip.get("full_name") or payslip.get("employee_name") or "Unknown Employee"
    
    lines = []
    
    # Debit: Expenses
    lines.append({
        "account_code": "6000",  # Salary & Wages
        "debit": gross_salary,
        "credit": 0,
        "memo": f"Gross salary - {employee_name}"
    })
    
    if epf_employer > 0:
        lines.append({
            "account_code": "6100",  # EPF Employer
            "debit": epf_employer,
            "credit": 0,
            "memo": "EPF employer contribution"
        })
    
    if socso_employer > 0:
        lines.append({
            "account_code": "6200",  # SOCSO Employer
            "debit": socso_employer,
            "credit": 0,
            "memo": "SOCSO employer contribution"
        })
    
    if eis_employer > 0:
        lines.append({
            "account_code": "6300",  # EIS Employer
            "debit": eis_employer,
            "credit": 0,
            "memo": "EIS employer contribution"
        })
    
    # Credit: Liabilities
    total_epf = round_money(epf_employee + epf_employer)
    if total_epf > 0:
        lines.append({
            "account_code": "2400",  # EPF Payable
            "debit": 0,
            "credit": total_epf,
            "memo": "EPF payable (employee + employer)"
        })
    
    total_socso = round_money(socso_employee + socso_employer)
    if total_socso > 0:
        lines.append({
            "account_code": "2450",  # SOCSO Payable
            "debit": 0,
            "credit": total_socso,
            "memo": "SOCSO payable (employee + employer)"
        })
    
    total_eis = round_money(eis_employee + eis_employer)
    if total_eis > 0:
        lines.append({
            "account_code": "2460",  # EIS Payable
            "debit": 0,
            "credit": total_eis,
            "memo": "EIS payable (employee + employer)"
        })
    
    if pcb > 0:
        lines.append({
            "account_code": "2470",  # PCB Payable
            "debit": 0,
            "credit": pcb,
            "memo": "PCB/MTD payable"
        })
    
    # Credit: Bank (net pay)
    lines.append({
        "account_code": settings.get("default_bank_account", "1000"),
        "debit": 0,
        "credit": net_pay,
        "memo": f"Net pay to {employee_name}"
    })
    
    result = await create_auto_journal_entry(
        entry_date=payroll_date,
        description=f"Payroll - {employee_name} for {month}/{year}",
        source_module="payroll",
        source_id=payslip_id,
        source_reference=f"PAY-{year}{str(month).zfill(2)}-{payslip_id[:6]}",
        lines=lines,
        created_by_id=user_id,
        created_by_name=user_name
    )
    
    return result


# ============ END PHASE 2 AUTO-POSTING FUNCTIONS ============


# ============ INITIALIZATION ============

async def initialize_accounting_system():
    """Initialize accounting system with default COA and settings"""
    
    # Check if already initialized
    existing = await db.chart_of_accounts.count_documents({})
    if existing > 0:
        return {"message": "Accounting system already initialized", "accounts": existing}
    
    # Default Chart of Accounts
    default_coa = [
        # ASSETS (1000-1999) — Balance Sheet
        {"account_code": "1000", "account_name": "Cash at Bank", "account_type": "Asset", "account_category": "Bank", "normal_balance": "debit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "1001", "account_name": "Petty Cash", "account_type": "Asset", "account_category": "Bank", "normal_balance": "debit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "1100", "account_name": "Accounts Receivable", "account_type": "Asset", "account_category": "AR", "normal_balance": "debit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "1200", "account_name": "Prepaid Expenses", "account_type": "Asset", "account_category": "Current Asset", "normal_balance": "debit", "is_system": False, "statement_type": "balance_sheet", "pnl_section": None},
        
        # LIABILITIES (2000-2999) — Balance Sheet
        {"account_code": "2100", "account_name": "Accounts Payable", "account_type": "Liability", "account_category": "AP", "normal_balance": "credit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "2200", "account_name": "SST Payable", "account_type": "Liability", "account_category": "Tax Liability", "normal_balance": "credit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "2300", "account_name": "Deferred Revenue", "account_type": "Liability", "account_category": "Deferred", "normal_balance": "credit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "2400", "account_name": "EPF Payable", "account_type": "Liability", "account_category": "Payroll Liability", "normal_balance": "credit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "2450", "account_name": "SOCSO Payable", "account_type": "Liability", "account_category": "Payroll Liability", "normal_balance": "credit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "2460", "account_name": "EIS Payable", "account_type": "Liability", "account_category": "Payroll Liability", "normal_balance": "credit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "2470", "account_name": "PCB Payable", "account_type": "Liability", "account_category": "Payroll Liability", "normal_balance": "credit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "2500", "account_name": "Accrued Expenses", "account_type": "Liability", "account_category": "Current Liability", "normal_balance": "credit", "is_system": False, "statement_type": "balance_sheet", "pnl_section": None},
        
        # EQUITY (3000-3999) — Balance Sheet
        {"account_code": "3000", "account_name": "Opening Balance Equity", "account_type": "Equity", "account_category": "Equity", "normal_balance": "credit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        {"account_code": "3100", "account_name": "Retained Earnings", "account_type": "Equity", "account_category": "Equity", "normal_balance": "credit", "is_system": True, "statement_type": "balance_sheet", "pnl_section": None},
        
        # REVENUE (4000-4099) — P&L: Revenue
        {"account_code": "4000", "account_name": "Training Revenue", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "is_system": True, "statement_type": "profit_and_loss", "pnl_section": "revenue"},
        {"account_code": "4001", "account_name": "Defensive Driving Training Income", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "revenue"},
        {"account_code": "4002", "account_name": "Defensive Riding Training Income", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "revenue"},
        {"account_code": "4003", "account_name": "Safety Talk Income", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "revenue"},
        {"account_code": "4004", "account_name": "Consultancy Income", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "revenue"},
        {"account_code": "4100", "account_name": "Other Income", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "other_income"},
        
        # COST OF SALES / DIRECT COSTS (5000-5999) — P&L: Cost of Sales
        {"account_code": "5000", "account_name": "Trainer Fees", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": True, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5001", "account_name": "Assistant Trainer Fees", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5100", "account_name": "Coordinator Fees", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": True, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5200", "account_name": "Marketing Commission", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": True, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5300", "account_name": "Training Materials", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5400", "account_name": "Venue Rental - Training", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5500", "account_name": "Fuel and Vehicle Usage", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5510", "account_name": "Toll and Travel - Direct", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5600", "account_name": "Accommodation - Direct", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5700", "account_name": "Meals - Direct", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        
        # OPERATING EXPENSES (6000-6999) — P&L: Operating Expense
        {"account_code": "6000", "account_name": "Office Salaries", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": True, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6100", "account_name": "EPF Employer Contribution", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": True, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6200", "account_name": "SOCSO Employer Contribution", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": True, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6300", "account_name": "EIS Employer Contribution", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": True, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6400", "account_name": "Office Expenses", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6410", "account_name": "Office Rental", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6500", "account_name": "Utilities", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6510", "account_name": "Internet and Phone", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6600", "account_name": "Petty Cash Expenses", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6700", "account_name": "Marketing Expenses", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6800", "account_name": "Software Subscription", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6900", "account_name": "Bank Charges", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6910", "account_name": "Professional Fees", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6920", "account_name": "Repairs and Maintenance", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6930", "account_name": "General Transport", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6999", "account_name": "Other Operating Expenses", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "is_system": False, "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
    ]
    
    now = get_malaysia_time().isoformat()
    for account in default_coa:
        account["id"] = str(uuid.uuid4())
        account["is_active"] = True
        account["description"] = ""
        account["parent_code"] = None
        account["created_by"] = "system"
        account["created_at"] = now
        account["updated_at"] = now
    
    await db.chart_of_accounts.insert_many(default_coa)
    
    # Create indexes
    await db.chart_of_accounts.create_index("account_code", unique=True)
    await db.chart_of_accounts.create_index("account_type")
    await db.chart_of_accounts.create_index("is_active")
    
    await db.journal_entries.create_index("journal_no", unique=True)
    await db.journal_entries.create_index("date")
    await db.journal_entries.create_index("source_module")
    await db.journal_entries.create_index("source_id")
    await db.journal_entries.create_index("status")
    await db.journal_entries.create_index([("date", 1), ("status", 1)])
    await db.journal_entries.create_index("lines.account_code")
    
    await db.accounting_periods.create_index([("year", 1), ("month", 1)], unique=True)
    await db.accounting_audit_log.create_index("timestamp")
    await db.accounting_audit_log.create_index("entity_type")
    
    # Default accounting settings
    settings = {
        "id": "accounting_settings",
        "fiscal_year_start_month": 1,
        "accounting_start_date": "2026-01-01",
        "default_currency": "MYR",
        "use_deferred_revenue": True,
        "auto_post_invoices": True,
        "auto_post_payments": True,
        "auto_post_expenses": True,
        "auto_post_payroll": True,
        "default_bank_account": "1000",
        "default_ar_account": "1100",
        "default_ap_account": "2100",
        "default_revenue_account": "4000",
        "default_deferred_revenue_account": "2300",
        "default_sst_account": "2200",
        "created_at": now,
        "updated_at": now
    }
    await db.accounting_settings.update_one(
        {"id": "accounting_settings"},
        {"$set": settings},
        upsert=True
    )
    
    # Create Jan and Feb 2026 periods as open
    for month in [1, 2]:
        period = {
            "id": str(uuid.uuid4()),
            "year": 2026,
            "month": month,
            "period_name": f"{'January' if month == 1 else 'February'} 2026",
            "status": "open",
            "closed_by": None,
            "closed_at": None,
            "created_at": now
        }
        await db.accounting_periods.update_one(
            {"year": 2026, "month": month},
            {"$setOnInsert": period},
            upsert=True
        )
    
    return {"message": "Accounting system initialized", "accounts": len(default_coa)}



@router.post("/upgrade-coa")
async def upgrade_coa(current_user: User = Depends(get_current_user)):
    """Upgrade existing COA: add statement_type, pnl_section fields, and new accounts.
    Safe to run multiple times - only adds what's missing."""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Admin only")
    
    now = get_malaysia_time().isoformat()
    updated = 0
    added = 0
    
    # 1. Auto-assign statement_type and pnl_section to existing accounts
    pnl_mapping = {
        # Revenue
        "4000": ("profit_and_loss", "revenue"), "4001": ("profit_and_loss", "revenue"),
        "4002": ("profit_and_loss", "revenue"), "4003": ("profit_and_loss", "revenue"),
        "4004": ("profit_and_loss", "revenue"),
        "4100": ("profit_and_loss", "other_income"),
        # Cost of Sales (5xxx)
        "5000": ("profit_and_loss", "cost_of_sales"), "5001": ("profit_and_loss", "cost_of_sales"),
        "5100": ("profit_and_loss", "cost_of_sales"), "5200": ("profit_and_loss", "cost_of_sales"),
        "5300": ("profit_and_loss", "cost_of_sales"), "5400": ("profit_and_loss", "cost_of_sales"),
        "5500": ("profit_and_loss", "cost_of_sales"), "5510": ("profit_and_loss", "cost_of_sales"),
        "5600": ("profit_and_loss", "cost_of_sales"), "5700": ("profit_and_loss", "cost_of_sales"),
        # Operating Expenses (6xxx)
        "6000": ("profit_and_loss", "operating_expense"), "6100": ("profit_and_loss", "operating_expense"),
        "6200": ("profit_and_loss", "operating_expense"), "6300": ("profit_and_loss", "operating_expense"),
        "6400": ("profit_and_loss", "operating_expense"), "6410": ("profit_and_loss", "operating_expense"),
        "6500": ("profit_and_loss", "operating_expense"), "6510": ("profit_and_loss", "operating_expense"),
        "6600": ("profit_and_loss", "operating_expense"), "6700": ("profit_and_loss", "operating_expense"),
        "6800": ("profit_and_loss", "operating_expense"), "6900": ("profit_and_loss", "operating_expense"),
        "6910": ("profit_and_loss", "operating_expense"), "6920": ("profit_and_loss", "operating_expense"),
        "6930": ("profit_and_loss", "operating_expense"), "6999": ("profit_and_loss", "operating_expense"),
    }
    
    all_accounts = await db.chart_of_accounts.find({}, {"_id": 0}).to_list(500)
    for acc in all_accounts:
        code = acc.get("account_code", "")
        needs_update = {}
        
        # Assign from mapping if known
        if code in pnl_mapping:
            st, ps = pnl_mapping[code]
            if acc.get("statement_type") != st:
                needs_update["statement_type"] = st
            if acc.get("pnl_section") != ps:
                needs_update["pnl_section"] = ps
        else:
            # Auto-assign by code range
            if not acc.get("statement_type"):
                if code.startswith(("1", "2", "3")):
                    needs_update["statement_type"] = "balance_sheet"
                elif code.startswith(("4", "5", "6")):
                    needs_update["statement_type"] = "profit_and_loss"
            if not acc.get("pnl_section") and code.startswith(("4", "5", "6")):
                if code.startswith("4") and code >= "4100":
                    needs_update["pnl_section"] = "other_income"
                elif code.startswith("4"):
                    needs_update["pnl_section"] = "revenue"
                elif code.startswith("5"):
                    needs_update["pnl_section"] = "cost_of_sales"
                elif code.startswith("6"):
                    needs_update["pnl_section"] = "operating_expense"
        
        if needs_update:
            needs_update["updated_at"] = now
            await db.chart_of_accounts.update_one({"account_code": code}, {"$set": needs_update})
            updated += 1
    
    # 2. Add new accounts that don't exist yet
    new_accounts = [
        {"account_code": "4001", "account_name": "Defensive Driving Training Income", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "statement_type": "profit_and_loss", "pnl_section": "revenue"},
        {"account_code": "4002", "account_name": "Defensive Riding Training Income", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "statement_type": "profit_and_loss", "pnl_section": "revenue"},
        {"account_code": "4003", "account_name": "Safety Talk Income", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "statement_type": "profit_and_loss", "pnl_section": "revenue"},
        {"account_code": "4004", "account_name": "Consultancy Income", "account_type": "Income", "account_category": "Revenue", "normal_balance": "credit", "statement_type": "profit_and_loss", "pnl_section": "revenue"},
        {"account_code": "5001", "account_name": "Assistant Trainer Fees", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5510", "account_name": "Toll and Travel - Direct", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5600", "account_name": "Accommodation - Direct", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "5700", "account_name": "Meals - Direct", "account_type": "Expense", "account_category": "Direct Cost", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "cost_of_sales"},
        {"account_code": "6410", "account_name": "Office Rental", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6510", "account_name": "Internet and Phone", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6700", "account_name": "Marketing Expenses", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6800", "account_name": "Software Subscription", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6900", "account_name": "Bank Charges", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6910", "account_name": "Professional Fees", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6920", "account_name": "Repairs and Maintenance", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6930", "account_name": "General Transport", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
        {"account_code": "6999", "account_name": "Other Operating Expenses", "account_type": "Expense", "account_category": "Operating Expense", "normal_balance": "debit", "statement_type": "profit_and_loss", "pnl_section": "operating_expense"},
    ]
    
    existing_codes = {a.get("account_code") for a in all_accounts}
    for acc in new_accounts:
        if acc["account_code"] not in existing_codes:
            acc["id"] = str(uuid.uuid4())
            acc["is_active"] = True
            acc["is_system"] = False
            acc["description"] = ""
            acc["parent_code"] = None
            acc["created_by"] = current_user.id
            acc["created_at"] = now
            acc["updated_at"] = now
            await db.chart_of_accounts.insert_one(acc)
            added += 1
    
    return {"message": f"COA upgraded. Updated {updated} existing accounts, added {added} new accounts.", "updated": updated, "added": added}


# ============ CHART OF ACCOUNTS ENDPOINTS ============

@router.post("/initialize")
async def initialize_accounting(current_user: User = Depends(get_current_user)):
    """Initialize the accounting system with default COA and settings (Super Admin only)"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin/Super Admin can initialize accounting")
    
    result = await initialize_accounting_system()
    
    await log_accounting_action(
        action="system_initialized",
        entity_type="system",
        entity_id="accounting",
        performed_by=current_user,
        after_value=result
    )
    
    return result


@router.get("/chart-of-accounts")
async def get_chart_of_accounts(
    account_type: Optional[str] = None,
    is_active: Optional[bool] = True,
    current_user: User = Depends(get_current_user)
):
    """Get all accounts in the Chart of Accounts"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if account_type:
        query["account_type"] = account_type
    if is_active is not None:
        query["is_active"] = is_active
    
    accounts = await db.chart_of_accounts.find(query, {"_id": 0}).sort("account_code", 1).to_list(500)
    
    # Group by account type for easier display
    grouped = {
        "Asset": [],
        "Liability": [],
        "Equity": [],
        "Income": [],
        "Expense": []
    }
    for account in accounts:
        acc_type = account.get("account_type", "Asset")
        if acc_type in grouped:
            grouped[acc_type].append(account)
    
    return {
        "accounts": accounts,
        "grouped": grouped,
        "total": len(accounts)
    }


@router.post("/chart-of-accounts")
async def create_account(
    account_data: ChartOfAccountCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new account in the Chart of Accounts (Super Admin only)"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin/Super Admin can create accounts")
    
    # Check if account code already exists
    existing = await db.chart_of_accounts.find_one({"account_code": account_data.account_code})
    if existing:
        raise HTTPException(status_code=400, detail=f"Account code {account_data.account_code} already exists")
    
    # Validate parent account if provided
    if account_data.parent_code:
        parent = await db.chart_of_accounts.find_one({"account_code": account_data.parent_code})
        if not parent:
            raise HTTPException(status_code=400, detail=f"Parent account {account_data.parent_code} not found")
    
    now = get_malaysia_time().isoformat()
    account = {
        "id": str(uuid.uuid4()),
        "account_code": account_data.account_code,
        "account_name": account_data.account_name,
        "account_type": account_data.account_type,
        "account_category": account_data.account_category,
        "parent_code": account_data.parent_code,
        "description": account_data.description or "",
        "normal_balance": account_data.normal_balance,
        "is_system": False,
        "is_active": True,
        "created_by": current_user.id,
        "created_at": now,
        "updated_at": now
    }
    
    await db.chart_of_accounts.insert_one(account)
    
    await log_accounting_action(
        action="account_created",
        entity_type="chart_of_accounts",
        entity_id=account["id"],
        performed_by=current_user,
        after_value={"account_code": account["account_code"], "account_name": account["account_name"]}
    )
    
    account.pop("_id", None)
    return {"message": "Account created", "account": account}


@router.put("/chart-of-accounts/{account_code}")
async def update_account(
    account_code: str,
    update_data: ChartOfAccountUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update an account in the Chart of Accounts (Super Admin only)"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin/Super Admin can update accounts")
    
    account = await db.chart_of_accounts.find_one({"account_code": account_code}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Cannot change type of system accounts
    if account.get("is_system") and update_data.is_active is False:
        # Check if account has been used in journal entries
        used = await db.journal_entries.find_one({"lines.account_code": account_code, "status": "posted"})
        if used:
            raise HTTPException(status_code=400, detail="Cannot deactivate system account that has been used")
    
    update_fields = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_fields["updated_at"] = get_malaysia_time().isoformat()
    
    before_value = {"account_name": account.get("account_name"), "is_active": account.get("is_active")}
    
    await db.chart_of_accounts.update_one(
        {"account_code": account_code},
        {"$set": update_fields}
    )
    
    await log_accounting_action(
        action="account_updated",
        entity_type="chart_of_accounts",
        entity_id=account["id"],
        performed_by=current_user,
        before_value=before_value,
        after_value=update_fields
    )
    
    updated = await db.chart_of_accounts.find_one({"account_code": account_code}, {"_id": 0})
    return {"message": "Account updated", "account": updated}


# ============ JOURNAL ENTRY ENDPOINTS ============

@router.get("/journal-entries")
async def get_journal_entries(
    year: Optional[int] = None,
    month: Optional[int] = None,
    status: Optional[str] = None,
    source_module: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    """Get journal entries with optional filters"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    
    if year and month:
        # Filter by specific month
        start_date = f"{year}-{str(month).zfill(2)}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{str(month + 1).zfill(2)}-01"
        query["date"] = {"$gte": start_date, "$lt": end_date}
    elif year:
        # Filter by year
        query["date"] = {"$gte": f"{year}-01-01", "$lt": f"{year + 1}-01-01"}
    
    if status:
        query["status"] = status
    
    if source_module:
        query["source_module"] = source_module
    
    entries = await db.journal_entries.find(query, {"_id": 0}).sort("date", -1).to_list(limit)
    
    return {
        "entries": entries,
        "count": len(entries),
        "filters": {"year": year, "month": month, "status": status, "source_module": source_module}
    }


@router.get("/journal-entries/{journal_id}")
async def get_journal_entry(journal_id: str, current_user: User = Depends(get_current_user)):
    """Get a single journal entry by ID"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    entry = await db.journal_entries.find_one({"id": journal_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    
    return entry


@router.post("/journal-entries")
async def create_journal_entry(
    entry_data: JournalEntryCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a manual journal entry (draft status)"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin/Finance can create journal entries")
    
    # Check if period is open
    if not await check_period_open(entry_data.date):
        raise HTTPException(status_code=400, detail="Cannot create journal entry in a closed period")
    
    # Validate and enrich lines with account names
    enriched_lines = await validate_journal_accounts(entry_data.lines)
    
    # Calculate totals
    total_debit = round_money(sum(line["debit"] for line in enriched_lines))
    total_credit = round_money(sum(line["credit"] for line in enriched_lines))
    
    # Generate journal number
    journal_no = await get_next_journal_number(entry_data.date)
    
    now = get_malaysia_time().isoformat()
    journal_entry = {
        "id": str(uuid.uuid4()),
        "journal_no": journal_no,
        "date": entry_data.date,
        "description": entry_data.description,
        "source_module": "manual",
        "source_id": None,
        "source_reference": None,
        "status": "draft",
        "lines": enriched_lines,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "is_balanced": abs(total_debit - total_credit) < 0.01,
        "created_by": current_user.id,
        "created_by_name": current_user.full_name,
        "posted_by": None,
        "posted_at": None,
        "voided_by": None,
        "voided_at": None,
        "void_reason": None,
        "created_at": now,
        "updated_at": now
    }
    
    await db.journal_entries.insert_one(journal_entry)
    
    await log_accounting_action(
        action="journal_created",
        entity_type="journal_entry",
        entity_id=journal_entry["id"],
        performed_by=current_user,
        after_value={"journal_no": journal_no, "total_debit": total_debit, "total_credit": total_credit}
    )
    
    journal_entry.pop("_id", None)
    return {"message": "Journal entry created (draft)", "journal_entry": journal_entry}


@router.post("/journal-entries/{journal_id}/post")
async def post_journal_entry(journal_id: str, current_user: User = Depends(get_current_user)):
    """Post a draft journal entry"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Only Admin/Finance can post journal entries")
    
    entry = await db.journal_entries.find_one({"id": journal_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    
    if entry.get("status") != "draft":
        raise HTTPException(status_code=400, detail="Only draft entries can be posted")
    
    if not entry.get("is_balanced"):
        raise HTTPException(status_code=400, detail="Cannot post unbalanced journal entry")
    
    # Check if period is still open
    if not await check_period_open(entry["date"]):
        raise HTTPException(status_code=400, detail="Cannot post to a closed period")
    
    now = get_malaysia_time().isoformat()
    await db.journal_entries.update_one(
        {"id": journal_id},
        {"$set": {
            "status": "posted",
            "posted_by": current_user.id,
            "posted_by_name": current_user.full_name,
            "posted_at": now,
            "updated_at": now
        }}
    )
    
    await log_accounting_action(
        action="journal_posted",
        entity_type="journal_entry",
        entity_id=journal_id,
        performed_by=current_user,
        before_value={"status": "draft"},
        after_value={"status": "posted"}
    )
    
    return {"message": "Journal entry posted", "journal_no": entry["journal_no"]}


@router.post("/journal-entries/{journal_id}/void")
async def void_journal_entry(
    journal_id: str,
    reason: str,
    current_user: User = Depends(get_current_user)
):
    """Void a posted journal entry"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin/Super Admin can void journal entries")
    
    if not reason or len(reason) < 10:
        raise HTTPException(status_code=400, detail="Void reason must be at least 10 characters")
    
    entry = await db.journal_entries.find_one({"id": journal_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    
    if entry.get("status") != "posted":
        raise HTTPException(status_code=400, detail="Only posted entries can be voided")
    
    now = get_malaysia_time().isoformat()
    await db.journal_entries.update_one(
        {"id": journal_id},
        {"$set": {
            "status": "voided",
            "voided_by": current_user.id,
            "voided_by_name": current_user.full_name,
            "voided_at": now,
            "void_reason": reason,
            "updated_at": now
        }}
    )
    
    await log_accounting_action(
        action="journal_voided",
        entity_type="journal_entry",
        entity_id=journal_id,
        performed_by=current_user,
        before_value={"status": "posted"},
        after_value={"status": "voided"},
        reason=reason
    )
    
    return {"message": "Journal entry voided", "journal_no": entry["journal_no"]}


# ============ ACCOUNTING PERIODS ============

@router.get("/periods")
async def get_accounting_periods(
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all accounting periods"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {}
    if year:
        query["year"] = year
    
    periods = await db.accounting_periods.find(query, {"_id": 0}).sort([("year", -1), ("month", -1)]).to_list(100)
    
    return {"periods": periods, "count": len(periods)}


@router.post("/periods/{year}/{month}/close")
async def close_period(
    year: int,
    month: int,
    request: PeriodCloseRequest,
    current_user: User = Depends(get_current_user)
):
    """Close an accounting period (Super Admin only)"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin/Super Admin can close periods")
    
    period = await db.accounting_periods.find_one({"year": year, "month": month}, {"_id": 0})
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")
    
    if period.get("status") == "closed":
        raise HTTPException(status_code=400, detail="Period is already closed")
    
    # Check for unposted journal entries in this period
    start_date = f"{year}-{str(month).zfill(2)}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{str(month + 1).zfill(2)}-01"
    
    unposted = await db.journal_entries.count_documents({
        "date": {"$gte": start_date, "$lt": end_date},
        "status": "draft"
    })
    
    if unposted > 0:
        raise HTTPException(status_code=400, detail=f"Cannot close period with {unposted} unposted journal entries")
    
    now = get_malaysia_time().isoformat()
    await db.accounting_periods.update_one(
        {"year": year, "month": month},
        {"$set": {
            "status": "closed",
            "closed_by": current_user.id,
            "closed_by_name": current_user.full_name,
            "closed_at": now,
            "close_reason": request.reason
        }}
    )
    
    await log_accounting_action(
        action="period_closed",
        entity_type="accounting_period",
        entity_id=period["id"],
        performed_by=current_user,
        after_value={"year": year, "month": month, "status": "closed"},
        reason=request.reason
    )
    
    month_name = ["", "January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"][month]
    
    return {"message": f"Period {month_name} {year} closed"}


@router.post("/periods/{year}/{month}/reopen")
async def reopen_period(
    year: int,
    month: int,
    request: PeriodReopenRequest,
    current_user: User = Depends(get_current_user)
):
    """Reopen a closed period (Super Admin only, with reason)"""
    if current_user.role not in ["super_admin"]:
        raise HTTPException(status_code=403, detail="Only Super Admin can reopen periods")
    
    period = await db.accounting_periods.find_one({"year": year, "month": month}, {"_id": 0})
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")
    
    if period.get("status") == "open":
        raise HTTPException(status_code=400, detail="Period is already open")
    
    now = get_malaysia_time().isoformat()
    await db.accounting_periods.update_one(
        {"year": year, "month": month},
        {"$set": {
            "status": "open",
            "reopened_by": current_user.id,
            "reopened_by_name": current_user.full_name,
            "reopened_at": now,
            "reopen_reason": request.reason
        }}
    )
    
    await log_accounting_action(
        action="period_reopened",
        entity_type="accounting_period",
        entity_id=period["id"],
        performed_by=current_user,
        before_value={"status": "closed"},
        after_value={"status": "open"},
        reason=request.reason
    )
    
    month_name = ["", "January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"][month]
    
    return {"message": f"Period {month_name} {year} reopened", "reason": request.reason}


# ============ SETTINGS ============

@router.get("/settings")
async def get_accounting_settings_endpoint(current_user: User = Depends(get_current_user)):
    """Get accounting settings"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    settings = await db.accounting_settings.find_one({"id": "accounting_settings"}, {"_id": 0})
    if not settings:
        return {"message": "Accounting system not initialized. Call /accounting/initialize first."}
    
    return settings


@router.put("/settings")
async def update_accounting_settings(
    settings_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update accounting settings (Super Admin only)"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin/Super Admin can update settings")
    
    # Protected fields that cannot be changed
    protected = ["id", "accounting_start_date", "created_at"]
    update_fields = {k: v for k, v in settings_data.items() if k not in protected}
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    update_fields["updated_at"] = get_malaysia_time().isoformat()
    update_fields["updated_by"] = current_user.id
    
    await db.accounting_settings.update_one(
        {"id": "accounting_settings"},
        {"$set": update_fields}
    )
    
    await log_accounting_action(
        action="settings_updated",
        entity_type="accounting_settings",
        entity_id="accounting_settings",
        performed_by=current_user,
        after_value=update_fields
    )
    
    return {"message": "Settings updated", "updated_fields": list(update_fields.keys())}


# ============ TRIAL BALANCE ============

@router.get("/trial-balance")
async def get_trial_balance(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    """Generate Trial Balance for a specific month"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get all posted journal entries up to end of specified month
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{str(month + 1).zfill(2)}-01"
    
    # Aggregate debits and credits by account
    pipeline = [
        {"$match": {"status": "posted", "date": {"$lt": end_date}}},
        {"$unwind": "$lines"},
        {"$group": {
            "_id": "$lines.account_code",
            "account_name": {"$first": "$lines.account_name"},
            "total_debit": {"$sum": "$lines.debit"},
            "total_credit": {"$sum": "$lines.credit"}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    results = await db.journal_entries.aggregate(pipeline).to_list(500)
    
    # Enrich with account details and calculate balances
    accounts = await db.chart_of_accounts.find({}, {"_id": 0}).to_list(500)
    account_map = {a["account_code"]: a for a in accounts}
    
    trial_balance = []
    total_debit = 0
    total_credit = 0
    
    for row in results:
        account_code = row["_id"]
        account = account_map.get(account_code, {})
        
        debit_total = round_money(row["total_debit"])
        credit_total = round_money(row["total_credit"])
        net_balance = debit_total - credit_total
        
        # Determine which column to show balance in based on normal balance
        normal_balance = account.get("normal_balance", "debit")
        if normal_balance == "debit":
            debit_balance = net_balance if net_balance >= 0 else 0
            credit_balance = abs(net_balance) if net_balance < 0 else 0
        else:
            credit_balance = abs(net_balance) if net_balance <= 0 else 0
            debit_balance = net_balance if net_balance > 0 else 0
        
        trial_balance.append({
            "account_code": account_code,
            "account_name": row["account_name"],
            "account_type": account.get("account_type", ""),
            "debit_balance": round_money(debit_balance),
            "credit_balance": round_money(credit_balance)
        })
        
        total_debit += debit_balance
        total_credit += credit_balance
    
    # Group by account type
    grouped = {"Asset": [], "Liability": [], "Equity": [], "Income": [], "Expense": []}
    for row in trial_balance:
        acc_type = row.get("account_type", "Asset")
        if acc_type in grouped:
            grouped[acc_type].append(row)
    
    month_name = ["", "January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"][month]
    
    return {
        "period": f"{month_name} {year}",
        "as_of_date": end_date,
        "trial_balance": trial_balance,
        "grouped": grouped,
        "totals": {
            "total_debit": round_money(total_debit),
            "total_credit": round_money(total_credit),
            "is_balanced": abs(total_debit - total_credit) < 0.01
        },
        "generated_at": get_malaysia_time().isoformat()
    }


# ============ GENERAL LEDGER ============

@router.get("/general-ledger/{account_code}")
async def get_general_ledger(
    account_code: str,
    year: int,
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Get General Ledger for a specific account"""
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Verify account exists
    account = await db.chart_of_accounts.find_one({"account_code": account_code}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Build date range
    if month:
        start_date = f"{year}-{str(month).zfill(2)}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{str(month + 1).zfill(2)}-01"
    else:
        start_date = f"{year}-01-01"
        end_date = f"{year + 1}-01-01"
    
    # Calculate opening balance (all entries before start_date)
    opening_pipeline = [
        {"$match": {"status": "posted", "date": {"$lt": start_date}}},
        {"$unwind": "$lines"},
        {"$match": {"lines.account_code": account_code}},
        {"$group": {
            "_id": None,
            "total_debit": {"$sum": "$lines.debit"},
            "total_credit": {"$sum": "$lines.credit"}
        }}
    ]
    
    opening_result = await db.journal_entries.aggregate(opening_pipeline).to_list(1)
    opening_debit = opening_result[0]["total_debit"] if opening_result else 0
    opening_credit = opening_result[0]["total_credit"] if opening_result else 0
    opening_balance = round_money(opening_debit - opening_credit)
    
    # Get transactions in the period
    entries = await db.journal_entries.find(
        {"status": "posted", "date": {"$gte": start_date, "$lt": end_date}, "lines.account_code": account_code},
        {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    # Build ledger with running balance
    ledger_entries = []
    running_balance = opening_balance
    
    for entry in entries:
        for line in entry.get("lines", []):
            if line["account_code"] == account_code:
                debit = round_money(line.get("debit", 0))
                credit = round_money(line.get("credit", 0))
                running_balance = round_money(running_balance + debit - credit)
                
                ledger_entries.append({
                    "date": entry["date"],
                    "journal_no": entry["journal_no"],
                    "description": entry["description"],
                    "memo": line.get("memo", ""),
                    "debit": debit,
                    "credit": credit,
                    "balance": running_balance
                })
    
    period_debit = round_money(sum(e["debit"] for e in ledger_entries))
    period_credit = round_money(sum(e["credit"] for e in ledger_entries))
    
    return {
        "account": account,
        "period": f"{year}" if not month else f"{year}-{str(month).zfill(2)}",
        "opening_balance": opening_balance,
        "entries": ledger_entries,
        "totals": {
            "period_debit": period_debit,
            "period_credit": period_credit,
            "closing_balance": running_balance
        },
        "generated_at": get_malaysia_time().isoformat()
    }


# ============ AUDIT LOG ============

@router.get("/audit-log")
async def get_accounting_audit_log(
    entity_type: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    """Get accounting audit log"""
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin/Super Admin can view audit log")
    
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    
    logs = await db.accounting_audit_log.find(query, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    
    return {"logs": logs, "count": len(logs)}


# ============ BALANCE SHEET ============

@router.get("/balance-sheet")
async def get_balance_sheet(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    """Generate Balance Sheet for a specific month
    
    Assets = Liabilities + Equity
    """
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # End date for balance calculation
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{str(month + 1).zfill(2)}-01"
    
    # Aggregate balances by account
    pipeline = [
        {"$match": {"status": "posted", "date": {"$lt": end_date}}},
        {"$unwind": "$lines"},
        {"$group": {
            "_id": "$lines.account_code",
            "account_name": {"$first": "$lines.account_name"},
            "account_type": {"$first": "$lines.account_type"},
            "total_debit": {"$sum": "$lines.debit"},
            "total_credit": {"$sum": "$lines.credit"}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    results = await db.journal_entries.aggregate(pipeline).to_list(500)
    
    # Get account details
    accounts = await db.chart_of_accounts.find({}, {"_id": 0}).to_list(500)
    account_map = {a["account_code"]: a for a in accounts}
    
    # Separate by type
    assets = {"accounts": [], "total": 0}
    liabilities = {"accounts": [], "total": 0}
    equity = {"accounts": [], "total": 0}
    income_total = 0
    expense_total = 0
    
    for row in results:
        account_code = row["_id"]
        account = account_map.get(account_code, {})
        account_type = account.get("account_type", row.get("account_type", ""))
        
        debit = round_money(row["total_debit"])
        credit = round_money(row["total_credit"])
        
        # Calculate balance based on normal balance
        normal_balance = account.get("normal_balance", "debit")
        if normal_balance == "debit":
            balance = debit - credit
        else:
            balance = credit - debit
        
        account_data = {
            "account_code": account_code,
            "account_name": row["account_name"],
            "balance": round_money(abs(balance)) if balance >= 0 else round_money(-abs(balance))
        }
        
        if account_type == "Asset":
            account_data["balance"] = round_money(debit - credit)
            assets["accounts"].append(account_data)
            assets["total"] += account_data["balance"]
        elif account_type == "Liability":
            account_data["balance"] = round_money(credit - debit)
            liabilities["accounts"].append(account_data)
            liabilities["total"] += account_data["balance"]
        elif account_type == "Equity":
            account_data["balance"] = round_money(credit - debit)
            equity["accounts"].append(account_data)
            equity["total"] += account_data["balance"]
        elif account_type == "Income":
            income_total += round_money(credit - debit)
        elif account_type == "Expense":
            expense_total += round_money(debit - credit)
    
    # Calculate current year earnings (Net Income)
    current_year_earnings = round_money(income_total - expense_total)
    equity["current_year_earnings"] = current_year_earnings
    equity["total"] = round_money(equity["total"] + current_year_earnings)
    
    assets["total"] = round_money(assets["total"])
    liabilities["total"] = round_money(liabilities["total"])
    
    total_liabilities_equity = round_money(liabilities["total"] + equity["total"])
    
    month_name = ["", "January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"][month]
    
    return {
        "period": f"As of {month_name} {year}",
        "assets": assets,
        "liabilities": liabilities,
        "equity": equity,
        "total_liabilities_equity": total_liabilities_equity,
        "is_balanced": abs(assets["total"] - total_liabilities_equity) < 0.01,
        "generated_at": get_malaysia_time().isoformat()
    }


# ============ ACCOUNTING PROFIT & LOSS ============

@router.get("/profit-loss")
async def get_accounting_profit_loss(
    year: int,
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Generate Profit & Loss from journal entries
    
    Revenue - Expenses = Net Profit/Loss
    """
    if current_user.role not in ["admin", "super_admin", "finance"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Build date range
    if month:
        start_date = f"{year}-{str(month).zfill(2)}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{str(month + 1).zfill(2)}-01"
        period_name = ["", "January", "February", "March", "April", "May", "June",
                      "July", "August", "September", "October", "November", "December"][month]
        period = f"{period_name} {year}"
    else:
        start_date = f"{year}-01-01"
        end_date = f"{year + 1}-01-01"
        period = f"Year {year}"
    
    # Aggregate by account for the period
    pipeline = [
        {"$match": {"status": "posted", "date": {"$gte": start_date, "$lt": end_date}}},
        {"$unwind": "$lines"},
        {"$group": {
            "_id": "$lines.account_code",
            "account_name": {"$first": "$lines.account_name"},
            "account_type": {"$first": "$lines.account_type"},
            "total_debit": {"$sum": "$lines.debit"},
            "total_credit": {"$sum": "$lines.credit"}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    results = await db.journal_entries.aggregate(pipeline).to_list(500)
    
    # Get account details
    accounts = await db.chart_of_accounts.find({}, {"_id": 0}).to_list(500)
    account_map = {a["account_code"]: a for a in accounts}
    
    revenue = {"accounts": [], "total": 0}
    expenses = {"accounts": [], "total": 0}
    
    for row in results:
        account_code = row["_id"]
        account = account_map.get(account_code, {})
        account_type = account.get("account_type", row.get("account_type", ""))
        
        debit = round_money(row["total_debit"])
        credit = round_money(row["total_credit"])
        
        if account_type == "Income":
            amount = round_money(credit - debit)  # Income has credit normal balance
            if amount != 0:
                revenue["accounts"].append({
                    "account_code": account_code,
                    "account_name": row["account_name"],
                    "amount": amount
                })
                revenue["total"] += amount
        elif account_type == "Expense":
            amount = round_money(debit - credit)  # Expense has debit normal balance
            if amount != 0:
                expenses["accounts"].append({
                    "account_code": account_code,
                    "account_name": row["account_name"],
                    "amount": amount
                })
                expenses["total"] += amount
    
    revenue["total"] = round_money(revenue["total"])
    expenses["total"] = round_money(expenses["total"])
    net_profit = round_money(revenue["total"] - expenses["total"])
    
    return {
        "period": period,
        "revenue": revenue,
        "expenses": expenses,
        "net_profit": net_profit,
        "generated_at": get_malaysia_time().isoformat()
    }


# ============ MIGRATION ENDPOINT (Phase 3) ============

@router.post("/migrate/2026")
async def migrate_2026_data(current_user: User = Depends(get_current_user)):
    """
    One-time migration to create journal entries from Jan-Feb 2026 transactions.
    
    Migrates:
    - Invoices (issued, partial, paid)
    - Payments received
    - Credit notes issued
    - Session expenses (with actuals)
    - Trainer fees
    - Coordinator fees
    - Marketing commissions
    """
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin/Super Admin can run migration")
    
    results = {
        "invoices": {"processed": 0, "created": 0, "skipped": 0, "errors": []},
        "payments": {"processed": 0, "created": 0, "skipped": 0, "errors": []},
        "credit_notes": {"processed": 0, "created": 0, "skipped": 0, "errors": []},
        "expenses": {"processed": 0, "created": 0, "skipped": 0, "errors": []},
        "trainer_fees": {"processed": 0, "created": 0, "skipped": 0, "errors": []},
        "coordinator_fees": {"processed": 0, "created": 0, "skipped": 0, "errors": []},
        "marketing_commissions": {"processed": 0, "created": 0, "skipped": 0, "errors": []}
    }
    
    settings = await get_accounting_settings()
    start_date = settings.get("accounting_start_date", "2026-01-01")
    
    # Import auto-posting functions
    from routes.accounting import (
        post_invoice_issued, post_payment_received, post_credit_note_issued,
        post_expense_recorded, post_trainer_fee, post_coordinator_fee, post_marketing_commission
    )
    
    # 1. Migrate Invoices
    invoices = await db.invoices.find({
        "status": {"$in": ["issued", "partial", "paid"]},
        "$or": [
            {"invoice_date": {"$gte": start_date}},
            {"created_at": {"$gte": start_date}}
        ]
    }, {"_id": 0}).to_list(1000)
    
    for inv in invoices:
        results["invoices"]["processed"] += 1
        try:
            session = None
            if inv.get("session_id"):
                session = await db.sessions.find_one({"id": inv["session_id"]}, {"_id": 0})
            
            result = await post_invoice_issued(inv, session, current_user.id, current_user.full_name)
            if result.get("is_duplicate"):
                results["invoices"]["skipped"] += 1
            elif result.get("journal_entry"):
                results["invoices"]["created"] += 1
            elif result.get("error"):
                results["invoices"]["errors"].append(f"{inv.get('invoice_number')}: {result['error']}")
        except Exception as e:
            results["invoices"]["errors"].append(f"{inv.get('invoice_number')}: {str(e)}")
    
    # 2. Migrate Payments
    payments = await db.payments.find({
        "created_at": {"$gte": start_date}
    }, {"_id": 0}).to_list(1000)
    
    for pmt in payments:
        results["payments"]["processed"] += 1
        try:
            invoice = await db.invoices.find_one({"id": pmt.get("invoice_id")}, {"_id": 0})
            result = await post_payment_received(pmt, invoice, current_user.id, current_user.full_name)
            if result.get("is_duplicate"):
                results["payments"]["skipped"] += 1
            elif result.get("journal_entry"):
                results["payments"]["created"] += 1
            elif result.get("error"):
                results["payments"]["errors"].append(f"{pmt.get('id')[:8]}: {result['error']}")
        except Exception as e:
            results["payments"]["errors"].append(f"{pmt.get('id')[:8]}: {str(e)}")
    
    # 3. Migrate Credit Notes
    credit_notes = await db.credit_notes.find({
        "status": "issued",
        "created_at": {"$gte": start_date}
    }, {"_id": 0}).to_list(1000)
    
    for cn in credit_notes:
        results["credit_notes"]["processed"] += 1
        try:
            result = await post_credit_note_issued(cn, current_user.id, current_user.full_name)
            if result.get("is_duplicate"):
                results["credit_notes"]["skipped"] += 1
            elif result.get("journal_entry"):
                results["credit_notes"]["created"] += 1
            elif result.get("error"):
                results["credit_notes"]["errors"].append(f"{cn.get('cn_number')}: {result['error']}")
        except Exception as e:
            results["credit_notes"]["errors"].append(f"{cn.get('cn_number')}: {str(e)}")
    
    # 4. Migrate Session Expenses
    # Get sessions from 2026
    sessions_2026 = await db.sessions.find({
        "start_date": {"$gte": start_date}
    }, {"_id": 0, "id": 1}).to_list(1000)
    session_ids_2026 = [s["id"] for s in sessions_2026]
    
    expenses = await db.session_expenses.find({
        "session_id": {"$in": session_ids_2026},
        "actual_amount": {"$gt": 0}
    }, {"_id": 0}).to_list(1000)
    
    for exp in expenses:
        results["expenses"]["processed"] += 1
        try:
            result = await post_expense_recorded(exp, "session", current_user.id, current_user.full_name)
            if result.get("is_duplicate"):
                results["expenses"]["skipped"] += 1
            elif result.get("journal_entry"):
                results["expenses"]["created"] += 1
            elif result.get("error"):
                results["expenses"]["errors"].append(f"{exp.get('id')[:8]}: {result['error']}")
        except Exception as e:
            results["expenses"]["errors"].append(f"{exp.get('id')[:8]}: {str(e)}")
    
    # 5. Migrate Trainer Fees
    trainer_fees = await db.trainer_fees.find({
        "session_id": {"$in": session_ids_2026}
    }, {"_id": 0}).to_list(1000)
    
    for tf in trainer_fees:
        results["trainer_fees"]["processed"] += 1
        try:
            session = await db.sessions.find_one({"id": tf.get("session_id")}, {"_id": 0})
            result = await post_trainer_fee(tf, session, current_user.id, current_user.full_name)
            if result.get("is_duplicate"):
                results["trainer_fees"]["skipped"] += 1
            elif result.get("journal_entry"):
                results["trainer_fees"]["created"] += 1
            elif result.get("error"):
                results["trainer_fees"]["errors"].append(f"{tf.get('id')[:8]}: {result['error']}")
        except Exception as e:
            results["trainer_fees"]["errors"].append(f"{tf.get('id')[:8]}: {str(e)}")
    
    # 6. Migrate Coordinator Fees
    coordinator_fees = await db.coordinator_fees.find({
        "session_id": {"$in": session_ids_2026}
    }, {"_id": 0}).to_list(1000)
    
    for cf in coordinator_fees:
        results["coordinator_fees"]["processed"] += 1
        try:
            session = await db.sessions.find_one({"id": cf.get("session_id")}, {"_id": 0})
            result = await post_coordinator_fee(cf, session, current_user.id, current_user.full_name)
            if result.get("is_duplicate"):
                results["coordinator_fees"]["skipped"] += 1
            elif result.get("journal_entry"):
                results["coordinator_fees"]["created"] += 1
            elif result.get("error"):
                results["coordinator_fees"]["errors"].append(f"{cf.get('id')[:8]}: {result['error']}")
        except Exception as e:
            results["coordinator_fees"]["errors"].append(f"{cf.get('id')[:8]}: {str(e)}")
    
    # 7. Migrate Marketing Commissions
    commissions = await db.marketing_commissions.find({
        "session_id": {"$in": session_ids_2026},
        "status": {"$in": ["approved", "paid"]}
    }, {"_id": 0}).to_list(1000)
    
    for mc in commissions:
        results["marketing_commissions"]["processed"] += 1
        try:
            session = await db.sessions.find_one({"id": mc.get("session_id")}, {"_id": 0})
            result = await post_marketing_commission(mc, session, current_user.id, current_user.full_name)
            if result.get("is_duplicate"):
                results["marketing_commissions"]["skipped"] += 1
            elif result.get("journal_entry"):
                results["marketing_commissions"]["created"] += 1
            elif result.get("error"):
                results["marketing_commissions"]["errors"].append(f"{mc.get('id')[:8]}: {result['error']}")
        except Exception as e:
            results["marketing_commissions"]["errors"].append(f"{mc.get('id')[:8]}: {str(e)}")
    
    # Calculate totals
    total_processed = sum(r["processed"] for r in results.values())
    total_created = sum(r["created"] for r in results.values())
    total_skipped = sum(r["skipped"] for r in results.values())
    
    await log_accounting_action(
        action="migration_completed",
        entity_type="migration",
        entity_id="2026",
        performed_by=current_user,
        after_value={
            "total_processed": total_processed,
            "total_created": total_created,
            "total_skipped": total_skipped
        }
    )
    
    return {
        "message": "Migration completed",
        "summary": {
            "total_processed": total_processed,
            "total_created": total_created,
            "total_skipped": total_skipped
        },
        "details": results
    }


# ============ OPENING BALANCE (Phase 4) ============

@router.post("/opening-balance")
async def create_opening_balance(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Create opening balance journal entry as of accounting start date.
    
    Expected data:
    {
        "date": "2026-01-01",
        "cash_at_bank": 0,
        "petty_cash": 0,
        "accounts_receivable": 0,  # Will be calculated from unpaid invoices if not provided
        "accounts_payable": 0      # Will be calculated from unpaid expenses if not provided
    }
    """
    if current_user.role not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Only Admin/Super Admin can create opening balance")
    
    entry_date = data.get("date", "2026-01-01")
    
    # Check if opening balance already exists
    existing = await db.journal_entries.find_one({
        "source_module": "opening_balance",
        "status": {"$ne": "voided"}
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(status_code=400, detail="Opening balance already exists. Void it first to create a new one.")
    
    settings = await get_accounting_settings()
    
    # Get values from input or calculate
    cash_at_bank = round_money(float(data.get("cash_at_bank", 0)))
    petty_cash = round_money(float(data.get("petty_cash", 0)))
    
    # Calculate AR from unpaid invoices before accounting start date
    if "accounts_receivable" in data:
        accounts_receivable = round_money(float(data["accounts_receivable"]))
    else:
        # Calculate from invoices issued before start date that are not fully paid
        ar_invoices = await db.invoices.find({
            "status": {"$in": ["issued", "partial"]},
            "created_at": {"$lt": entry_date}
        }, {"_id": 0, "total_amount": 1}).to_list(1000)
        accounts_receivable = round_money(sum(inv.get("total_amount", 0) for inv in ar_invoices))
    
    # Calculate AP from unpaid expenses before start date
    if "accounts_payable" in data:
        accounts_payable = round_money(float(data["accounts_payable"]))
    else:
        accounts_payable = 0  # Default to 0 if not provided
    
    # Build journal lines
    lines = []
    total_debit = 0
    total_credit = 0
    
    # Debit: Assets
    if cash_at_bank > 0:
        lines.append({
            "account_code": settings.get("default_bank_account", "1000"),
            "debit": cash_at_bank,
            "credit": 0,
            "memo": "Opening balance - Cash at Bank"
        })
        total_debit += cash_at_bank
    
    if petty_cash > 0:
        lines.append({
            "account_code": "1001",
            "debit": petty_cash,
            "credit": 0,
            "memo": "Opening balance - Petty Cash"
        })
        total_debit += petty_cash
    
    if accounts_receivable > 0:
        lines.append({
            "account_code": settings.get("default_ar_account", "1100"),
            "debit": accounts_receivable,
            "credit": 0,
            "memo": "Opening balance - Accounts Receivable"
        })
        total_debit += accounts_receivable
    
    # Credit: Liabilities
    if accounts_payable > 0:
        lines.append({
            "account_code": settings.get("default_ap_account", "2100"),
            "debit": 0,
            "credit": accounts_payable,
            "memo": "Opening balance - Accounts Payable"
        })
        total_credit += accounts_payable
    
    # Calculate balancing equity
    equity_balance = round_money(total_debit - total_credit)
    
    if equity_balance != 0:
        lines.append({
            "account_code": "3000",  # Opening Balance Equity
            "debit": 0 if equity_balance > 0 else abs(equity_balance),
            "credit": equity_balance if equity_balance > 0 else 0,
            "memo": "Opening balance - Equity (balancing)"
        })
    
    # Create journal entry
    result = await create_auto_journal_entry(
        entry_date=entry_date,
        description=f"Opening Balance as of {entry_date}",
        source_module="opening_balance",
        source_id="opening_balance_2026",
        source_reference="OB-2026",
        lines=lines,
        created_by_id=current_user.id,
        created_by_name=current_user.full_name
    )
    
    return {
        "message": "Opening balance created",
        "opening_balance": {
            "cash_at_bank": cash_at_bank,
            "petty_cash": petty_cash,
            "accounts_receivable": accounts_receivable,
            "accounts_payable": accounts_payable,
            "equity_balance": equity_balance
        },
        "journal_entry": result.get("journal_entry")
    }


# ============ EXCEL EXPORT ENDPOINTS ============

@router.get("/journal-entries/export/excel")
async def export_journal_entries_excel(
    year: int = 2026,
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Export journal entries to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Build query
    query = {"$expr": {"$eq": [{"$year": {"$dateFromString": {"dateString": "$date"}}}, year]}}
    if month:
        query["$expr"] = {"$and": [
            {"$eq": [{"$year": {"$dateFromString": {"dateString": "$date"}}}, year]},
            {"$eq": [{"$month": {"$dateFromString": {"dateString": "$date"}}}, month]}
        ]}
    
    entries = await db.journal_entries.find(
        query, {"_id": 0}
    ).sort("date", 1).to_list(5000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Entries"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Journal #", "Date", "Description", "Source", "Reference", "Account Code", "Account Name", "Debit (RM)", "Credit (RM)", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 2
    for entry in entries:
        for line in entry.get("lines", []):
            ws.cell(row=row, column=1, value=entry.get("journal_no", "")).border = thin_border
            ws.cell(row=row, column=2, value=entry.get("date", "")).border = thin_border
            ws.cell(row=row, column=3, value=entry.get("description", "")).border = thin_border
            ws.cell(row=row, column=4, value=entry.get("source_module", "")).border = thin_border
            ws.cell(row=row, column=5, value=entry.get("source_reference", "")).border = thin_border
            ws.cell(row=row, column=6, value=line.get("account_code", "")).border = thin_border
            ws.cell(row=row, column=7, value=line.get("account_name", "")).border = thin_border
            
            debit_cell = ws.cell(row=row, column=8, value=float(line.get("debit", 0)))
            debit_cell.number_format = '#,##0.00'
            debit_cell.border = thin_border
            
            credit_cell = ws.cell(row=row, column=9, value=float(line.get("credit", 0)))
            credit_cell.number_format = '#,##0.00'
            credit_cell.border = thin_border
            
            ws.cell(row=row, column=10, value=entry.get("status", "")).border = thin_border
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"journal_entries_{year}_{month or 'all'}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/trial-balance/export/excel")
async def export_trial_balance_excel(
    year: int = 2026,
    month: int = 2,
    current_user: User = Depends(get_current_user)
):
    """Export trial balance to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Get trial balance data
    accounts = await db.chart_of_accounts.find(
        {"is_active": True}, {"_id": 0}
    ).sort("account_code", 1).to_list(500)
    
    if month == 12:
        period_end = f"{year+1}-01-01"
    else:
        period_end = f"{year}-{month+1:02d}-01"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="TRIAL BALANCE")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    period_cell = ws.cell(row=2, column=1, value=f"As of {year}-{month:02d}")
    period_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Account Code", "Account Name", "Account Type", "Debit (RM)", "Credit (RM)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 5
    total_debit = 0
    total_credit = 0
    
    for acc in accounts:
        # Get balance from posted journal entries
        pipeline = [
            {"$match": {"status": "posted", "date": {"$lt": period_end}}},
            {"$unwind": "$lines"},
            {"$match": {"lines.account_code": acc["account_code"]}},
            {"$group": {
                "_id": None,
                "total_debit": {"$sum": "$lines.debit"},
                "total_credit": {"$sum": "$lines.credit"}
            }}
        ]
        result = await db.journal_entries.aggregate(pipeline).to_list(1)
        
        debit = float(result[0]["total_debit"]) if result else 0
        credit = float(result[0]["total_credit"]) if result else 0
        
        if debit == 0 and credit == 0:
            continue
        
        # Calculate balance based on normal balance
        net = debit - credit
        show_debit = net if net > 0 else 0
        show_credit = abs(net) if net < 0 else 0
        
        ws.cell(row=row, column=1, value=acc["account_code"]).border = thin_border
        ws.cell(row=row, column=2, value=acc["account_name"]).border = thin_border
        ws.cell(row=row, column=3, value=acc["account_type"]).border = thin_border
        
        debit_cell = ws.cell(row=row, column=4, value=show_debit)
        debit_cell.number_format = '#,##0.00'
        debit_cell.border = thin_border
        
        credit_cell = ws.cell(row=row, column=5, value=show_credit)
        credit_cell.number_format = '#,##0.00'
        credit_cell.border = thin_border
        
        total_debit += show_debit
        total_credit += show_credit
        row += 1
    
    # Totals row
    total_font = Font(bold=True)
    ws.cell(row=row, column=1, value="").border = thin_border
    ws.cell(row=row, column=2, value="TOTAL").font = total_font
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=3, value="").border = thin_border
    
    total_debit_cell = ws.cell(row=row, column=4, value=total_debit)
    total_debit_cell.number_format = '#,##0.00'
    total_debit_cell.font = total_font
    total_debit_cell.border = thin_border
    
    total_credit_cell = ws.cell(row=row, column=5, value=total_credit)
    total_credit_cell.number_format = '#,##0.00'
    total_credit_cell.font = total_font
    total_credit_cell.border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=trial_balance_{year}_{month}.xlsx"}
    )


@router.get("/profit-loss/export/excel")
async def export_pl_excel(
    year: int = 2026,
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """Export P&L to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Get P&L data
    if month:
        period_start = f"{year}-{month:02d}-01"
        if month == 12:
            period_end = f"{year+1}-01-01"
        else:
            period_end = f"{year}-{month+1:02d}-01"
        period_label = f"{year}-{month:02d}"
    else:
        period_start = f"{year}-01-01"
        period_end = f"{year+1}-01-01"
        period_label = f"Full Year {year}"
    
    income_accounts = await db.chart_of_accounts.find(
        {"account_type": "Income", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    expense_accounts = await db.chart_of_accounts.find(
        {"account_type": "Expense", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    
    # Title
    ws.merge_cells('A1:C1')
    ws.cell(row=1, column=1, value="PROFIT & LOSS STATEMENT").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    ws.cell(row=2, column=1, value=period_label).alignment = Alignment(horizontal='center')
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    row = 4
    
    # Revenue section
    ws.cell(row=row, column=1, value="REVENUE").font = Font(bold=True, color="228B22")
    row += 1
    
    total_revenue = 0
    for acc in income_accounts:
        pipeline = [
            {"$match": {"status": "posted", "date": {"$gte": period_start, "$lt": period_end}}},
            {"$unwind": "$lines"},
            {"$match": {"lines.account_code": acc["account_code"]}},
            {"$group": {"_id": None, "credit": {"$sum": "$lines.credit"}, "debit": {"$sum": "$lines.debit"}}}
        ]
        result = await db.journal_entries.aggregate(pipeline).to_list(1)
        amount = float(result[0]["credit"] - result[0]["debit"]) if result else 0
        
        if amount != 0:
            ws.cell(row=row, column=1, value=acc["account_name"]).border = thin_border
            amt_cell = ws.cell(row=row, column=3, value=amount)
            amt_cell.number_format = '#,##0.00'
            amt_cell.border = thin_border
            total_revenue += amount
            row += 1
    
    ws.cell(row=row, column=1, value="Total Revenue").font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    total_rev_cell = ws.cell(row=row, column=3, value=total_revenue)
    total_rev_cell.number_format = '#,##0.00'
    total_rev_cell.font = Font(bold=True)
    total_rev_cell.border = thin_border
    row += 2
    
    # Expenses section
    ws.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, color="B22222")
    row += 1
    
    total_expenses = 0
    for acc in expense_accounts:
        pipeline = [
            {"$match": {"status": "posted", "date": {"$gte": period_start, "$lt": period_end}}},
            {"$unwind": "$lines"},
            {"$match": {"lines.account_code": acc["account_code"]}},
            {"$group": {"_id": None, "debit": {"$sum": "$lines.debit"}, "credit": {"$sum": "$lines.credit"}}}
        ]
        result = await db.journal_entries.aggregate(pipeline).to_list(1)
        amount = float(result[0]["debit"] - result[0]["credit"]) if result else 0
        
        if amount != 0:
            ws.cell(row=row, column=1, value=acc["account_name"]).border = thin_border
            amt_cell = ws.cell(row=row, column=3, value=amount)
            amt_cell.number_format = '#,##0.00'
            amt_cell.border = thin_border
            total_expenses += amount
            row += 1
    
    ws.cell(row=row, column=1, value="Total Expenses").font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    total_exp_cell = ws.cell(row=row, column=3, value=total_expenses)
    total_exp_cell.number_format = '#,##0.00'
    total_exp_cell.font = Font(bold=True)
    total_exp_cell.border = thin_border
    row += 2
    
    # Net profit
    net_profit = total_revenue - total_expenses
    ws.cell(row=row, column=1, value="NET PROFIT" if net_profit >= 0 else "NET LOSS").font = Font(bold=True, size=12)
    profit_cell = ws.cell(row=row, column=3, value=net_profit)
    profit_cell.number_format = '#,##0.00'
    profit_cell.font = Font(bold=True, size=12, color="228B22" if net_profit >= 0 else "B22222")
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['C'].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pl_{year}_{month or 'full'}.xlsx"}
    )


@router.get("/balance-sheet/export/excel")
async def export_balance_sheet_excel(
    year: int = 2026,
    month: int = 2,
    current_user: User = Depends(get_current_user)
):
    """Export Balance Sheet to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    if month == 12:
        period_end = f"{year+1}-01-01"
    else:
        period_end = f"{year}-{month+1:02d}-01"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    
    # Title
    ws.merge_cells('A1:C1')
    ws.cell(row=1, column=1, value="BALANCE SHEET").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    ws.cell(row=2, column=1, value=f"As of {year}-{month:02d}").alignment = Alignment(horizontal='center')
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    row = 4
    
    async def get_account_balance(account_code: str) -> float:
        pipeline = [
            {"$match": {"status": "posted", "date": {"$lt": period_end}}},
            {"$unwind": "$lines"},
            {"$match": {"lines.account_code": account_code}},
            {"$group": {"_id": None, "debit": {"$sum": "$lines.debit"}, "credit": {"$sum": "$lines.credit"}}}
        ]
        result = await db.journal_entries.aggregate(pipeline).to_list(1)
        if result:
            return float(result[0]["debit"] - result[0]["credit"])
        return 0
    
    # Assets
    ws.cell(row=row, column=1, value="ASSETS").font = Font(bold=True, size=12)
    row += 1
    
    asset_accounts = await db.chart_of_accounts.find(
        {"account_type": "Asset", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    total_assets = 0
    for acc in asset_accounts:
        balance = await get_account_balance(acc["account_code"])
        if balance != 0:
            ws.cell(row=row, column=1, value=acc["account_name"]).border = thin_border
            bal_cell = ws.cell(row=row, column=3, value=balance)
            bal_cell.number_format = '#,##0.00'
            bal_cell.border = thin_border
            total_assets += balance
            row += 1
    
    ws.cell(row=row, column=1, value="Total Assets").font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    total_assets_cell = ws.cell(row=row, column=3, value=total_assets)
    total_assets_cell.number_format = '#,##0.00'
    total_assets_cell.font = Font(bold=True)
    total_assets_cell.border = thin_border
    row += 2
    
    # Liabilities
    ws.cell(row=row, column=1, value="LIABILITIES").font = Font(bold=True, size=12)
    row += 1
    
    liability_accounts = await db.chart_of_accounts.find(
        {"account_type": "Liability", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    total_liabilities = 0
    for acc in liability_accounts:
        balance = await get_account_balance(acc["account_code"])
        balance = -balance  # Liabilities have credit balance
        if balance != 0:
            ws.cell(row=row, column=1, value=acc["account_name"]).border = thin_border
            bal_cell = ws.cell(row=row, column=3, value=balance)
            bal_cell.number_format = '#,##0.00'
            bal_cell.border = thin_border
            total_liabilities += balance
            row += 1
    
    ws.cell(row=row, column=1, value="Total Liabilities").font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    total_liab_cell = ws.cell(row=row, column=3, value=total_liabilities)
    total_liab_cell.number_format = '#,##0.00'
    total_liab_cell.font = Font(bold=True)
    total_liab_cell.border = thin_border
    row += 2
    
    # Equity
    ws.cell(row=row, column=1, value="EQUITY").font = Font(bold=True, size=12)
    row += 1
    
    equity_accounts = await db.chart_of_accounts.find(
        {"account_type": "Equity", "is_active": True}, {"_id": 0}
    ).to_list(100)
    
    total_equity = 0
    for acc in equity_accounts:
        balance = await get_account_balance(acc["account_code"])
        balance = -balance  # Equity has credit balance
        if balance != 0:
            ws.cell(row=row, column=1, value=acc["account_name"]).border = thin_border
            bal_cell = ws.cell(row=row, column=3, value=balance)
            bal_cell.number_format = '#,##0.00'
            bal_cell.border = thin_border
            total_equity += balance
            row += 1
    
    ws.cell(row=row, column=1, value="Total Equity").font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    total_eq_cell = ws.cell(row=row, column=3, value=total_equity)
    total_eq_cell.number_format = '#,##0.00'
    total_eq_cell.font = Font(bold=True)
    total_eq_cell.border = thin_border
    row += 2
    
    # Total Liabilities + Equity
    ws.cell(row=row, column=1, value="TOTAL LIABILITIES + EQUITY").font = Font(bold=True, size=12)
    total_le_cell = ws.cell(row=row, column=3, value=total_liabilities + total_equity)
    total_le_cell.number_format = '#,##0.00'
    total_le_cell.font = Font(bold=True, size=12)
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['C'].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=balance_sheet_{year}_{month}.xlsx"}
    )
